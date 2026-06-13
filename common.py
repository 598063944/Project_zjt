# -*- coding: utf-8 -*-
"""
common.py — 通用 UI 组件与共享函数
───────────────────────────────────
负责：
  - 筛选器组件（FilterBar / 条件行 / 预设管理）
  - 字段显示设置对话框（FieldSettingsDialog）
  - 日期选择器（QuickDatePickerDialog / DatePartPickerDialog）
  - 表格组件（CheckBoxHeader / 分页栏 / 右键菜单 / 复制粘贴 / 内联编辑）
  - 弹窗基类（CenteredPopupDialog / CustomMessageBox）
  - 通用小部件（BrowseLineEdit / PasswordEntry / CheckableOptionPopup）
  - UI 样式函数（玻璃效果 / 导航按钮 / 缩放 / 窗口居中 / QMessageBox 猴子补丁）
  - 全局状态读取（_get_current_user / _get_user_type）
依赖：core.py / network.py / auth.py / custom_report
被导入：全部 8 个 Mixin + 主程序
"""
from core import *
from network import *
from auth import *
import custom_report
_main = None  # lazy import，避免循环依赖
choice = None  # 全局功能选择标识，由 file_generation.py 在后台线程中设置

def _get_main():
    """延迟获取主模块引用（避免循环导入）。"""
    global _main
    if _main is None:
        import sys as _sys
        _main = _sys.modules.get('__main__')
    return _main
from PyQt6.QtWidgets import (QApplication, QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QAbstractItemView, QListWidget, QListWidgetItem, QLineEdit,
    QHeaderView, QMessageBox, QWidget, QFrame, QComboBox, QCheckBox, QTableWidget,
    QTableView, QTreeView, QListView, QMainWindow)
from PyQt6.QtCore import Qt, QRect, QSize, QTimer, QEvent, pyqtSignal, QPoint, QObject
from PyQt6.QtGui import QColor, QFont, QPalette, QGuiApplication, QIcon

# ── Shared global state (read from main module to stay in sync) ──
import sys as _sys
def _get_current_user():
    _m = _sys.modules.get('__main__')
    return getattr(_m, 'current_user', None) if _m else None
def _get_user_type():
    _m = _sys.modules.get('__main__')
    return getattr(_m, 'user_type', None) if _m else None
def load_config():
    """加载配置（延迟从 __main__ 获取，避免模块加载时序问题）"""
    import sys as _sys
    _m = _sys.modules.get('__main__')
    if _m is not None and hasattr(_m, 'load_config'):
        return _m.load_config()
    return {}

def __getattr__(name):
    """延迟从 __main__ 获取任何未导入的变量/函数（解决 from core import * 的时序问题）"""
    import sys as _s
    _m = _s.modules.get('__main__')
    if _m is not None and hasattr(_m, name):
        return getattr(_m, name)
    raise AttributeError(f"module 'common' has no attribute {name!r}")


# ===== Monkey-patch QMessageBox：弹窗内容同步打印到运行窗口 =====
_orig_msg_warning = QMessageBox.warning
_orig_msg_info = QMessageBox.information
_orig_msg_critical = QMessageBox.critical
_orig_msg_question = QMessageBox.question

def _msg_warning(parent, title, text, *args, **kwargs):
    print(f"[弹窗警告] {title}: {text}")
    return _orig_msg_warning(parent, title, text, *args, **kwargs)

def _msg_info(parent, title, text, *args, **kwargs):
    print(f"[弹窗提示] {title}: {text}")
    return _orig_msg_info(parent, title, text, *args, **kwargs)

def _msg_critical(parent, title, text, *args, **kwargs):
    print(f"[弹窗错误] {title}: {text}")
    return _orig_msg_critical(parent, title, text, *args, **kwargs)

def _msg_question(parent, title, text, *args, **kwargs):
    print(f"[弹窗确认] {title}: {text}")
    return _orig_msg_question(parent, title, text, *args, **kwargs)

QMessageBox.warning = _msg_warning
QMessageBox.information = _msg_info
QMessageBox.critical = _msg_critical
QMessageBox.question = _msg_question

_qt_app_instance = None


def ensure_qapplication():
    """确保当前进程存在QApplication实例。"""
    global _qt_app_instance
    app = QApplication.instance()
    if app is None:
        # 设置高DPI缩放策略，确保125%等非整数缩放比例正确渲染
        QGuiApplication.setHighDpiScaleFactorRoundingPolicy(
            Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
        )
        app = QApplication(sys.argv)
    _qt_app_instance = app
    return app


def calculate_centered_window_rect(widget=None, width_ratio=0.7, height_ratio=0.7, min_width=800, min_height=600):
    """按屏幕可用区域计算窗口大小并返回居中的矩形。"""
    app = QApplication.instance()
    screen = widget.screen() if widget and hasattr(widget, 'screen') else None
    if screen is None and app is not None:
        screen = app.primaryScreen()
    if screen is None:
        return 100, 100, max(min_width, 800), max(min_height, 600)

    available_geometry = screen.availableGeometry()
    width = min(available_geometry.width(), max(min_width, int(available_geometry.width() * width_ratio)))
    height = min(available_geometry.height(), max(min_height, int(available_geometry.height() * height_ratio)))
    x = available_geometry.left() + (available_geometry.width() - width) // 2
    y = available_geometry.top() + max(0, (available_geometry.height() - height) // 2)
    return x, y, width, height


def calculate_startup_window_rect(widget=None, saved_size=None, width_ratio=0.7, height_ratio=0.7, min_width=800, min_height=600):
    """启动时优先使用上次保存的窗口大小，并始终保持居中。"""
    app = QApplication.instance()
    screen = widget.screen() if widget and hasattr(widget, 'screen') else None
    if screen is None and app is not None:
        screen = app.primaryScreen()
    if screen is None:
        fallback_width = max(min_width, int(saved_size[0])) if isinstance(saved_size, (list, tuple)) and len(saved_size) >= 1 else max(min_width, 800)
        fallback_height = max(min_height, int(saved_size[1])) if isinstance(saved_size, (list, tuple)) and len(saved_size) >= 2 else max(min_height, 600)
        return 100, 100, fallback_width, fallback_height

    available_geometry = screen.availableGeometry()

    preferred_width = None
    preferred_height = None
    if isinstance(saved_size, (list, tuple)) and len(saved_size) >= 2:
        try:
            preferred_width = int(saved_size[0])
            preferred_height = int(saved_size[1])
        except (TypeError, ValueError):
            preferred_width = None
            preferred_height = None

    if preferred_width is None or preferred_height is None:
        return calculate_centered_window_rect(
            widget=widget,
            width_ratio=width_ratio,
            height_ratio=height_ratio,
            min_width=min_width,
            min_height=min_height,
        )

    # 屏幕够大用保存的尺寸，屏幕小于保存尺寸则用屏幕的94%
    if preferred_width <= available_geometry.width():
        width = max(min_width, preferred_width)
    else:
        width = max(min_width, int(available_geometry.width() * 0.96))
    if preferred_height <= available_geometry.height():
        height = max(min_height, preferred_height)
    else:
        height = max(min_height, int(available_geometry.height() * 0.96))
    x = available_geometry.left() + (available_geometry.width() - width) // 2
    y = available_geometry.top() + max(0, (available_geometry.height() - height) // 2)
    return x, y, width, height


def normalize_ui_scale_percent(value, default=100):
    """规范化界面缩放百分比。"""
    try:
        normalized = int(round(float(value)))
    except (TypeError, ValueError):
        normalized = default
    return max(70, min(200, normalized))


def get_current_ui_scale_percent():
    """读取当前应用级界面缩放百分比。"""
    app = QApplication.instance()
    if app is None:
        return 100
    return normalize_ui_scale_percent(app.property("_global_ui_scale_percent"), default=100)


def scale_ui_value(value, scale_percent=None, minimum=1):
    """按界面缩放百分比换算像素值。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else normalize_ui_scale_percent(scale_percent)
    return max(minimum, int(round(float(value) * effective_percent / 100.0)))

# QPalette颜色角色枚举
QPalette.Window = QPalette.ColorRole.Window
QPalette.WindowText = QPalette.ColorRole.WindowText
QPalette.Base = QPalette.ColorRole.Base
QPalette.AlternateBase = QPalette.ColorRole.AlternateBase
QPalette.ToolTipBase = QPalette.ColorRole.ToolTipBase
QPalette.ToolTipText = QPalette.ColorRole.ToolTipText
QPalette.Text = QPalette.ColorRole.Text
QPalette.Button = QPalette.ColorRole.Button
QPalette.ButtonText = QPalette.ColorRole.ButtonText
QPalette.BrightText = QPalette.ColorRole.BrightText
QPalette.Link = QPalette.ColorRole.Link
QPalette.Highlight = QPalette.ColorRole.Highlight
QPalette.HighlightedText = QPalette.ColorRole.HighlightedText

# QLineEdit回显模式枚举
QLineEdit.Normal = QLineEdit.EchoMode.Normal
QLineEdit.NoEcho = QLineEdit.EchoMode.NoEcho
QLineEdit.Password = QLineEdit.EchoMode.Password
QLineEdit.PasswordEchoOnEdit = QLineEdit.EchoMode.PasswordEchoOnEdit

# QAbstractItemView选择模式枚举
QAbstractItemView.SingleSelection = QAbstractItemView.SelectionMode.SingleSelection
QAbstractItemView.MultiSelection = QAbstractItemView.SelectionMode.MultiSelection
QAbstractItemView.ExtendedSelection = QAbstractItemView.SelectionMode.ExtendedSelection
QAbstractItemView.NoSelection = QAbstractItemView.SelectionMode.NoSelection
QAbstractItemView.SelectionMode.Single = QAbstractItemView.SelectionMode.SingleSelection
QAbstractItemView.SelectionMode.Multi = QAbstractItemView.SelectionMode.MultiSelection

# QHeaderView调整模式枚举
QHeaderView.Interactive = QHeaderView.ResizeMode.Interactive
QHeaderView.Stretch = QHeaderView.ResizeMode.Stretch
QHeaderView.Fixed = QHeaderView.ResizeMode.Fixed
QHeaderView.ResizeToContents = QHeaderView.ResizeMode.ResizeToContents

# QMessageBox标准按钮枚举
QMessageBox.Ok = QMessageBox.StandardButton.Ok
QMessageBox.Open = QMessageBox.StandardButton.Open
QMessageBox.Save = QMessageBox.StandardButton.Save
QMessageBox.Cancel = QMessageBox.StandardButton.Cancel
QMessageBox.Close = QMessageBox.StandardButton.Close
QMessageBox.Discard = QMessageBox.StandardButton.Discard
QMessageBox.Apply = QMessageBox.StandardButton.Apply
QMessageBox.Reset = QMessageBox.StandardButton.Reset
QMessageBox.RestoreDefaults = QMessageBox.StandardButton.RestoreDefaults
QMessageBox.Help = QMessageBox.StandardButton.Help
QMessageBox.SaveAll = QMessageBox.StandardButton.SaveAll
QMessageBox.Yes = QMessageBox.StandardButton.Yes
QMessageBox.YesToAll = QMessageBox.StandardButton.YesToAll
QMessageBox.No = QMessageBox.StandardButton.No
QMessageBox.NoToAll = QMessageBox.StandardButton.NoToAll
QMessageBox.Abort = QMessageBox.StandardButton.Abort
QMessageBox.Retry = QMessageBox.StandardButton.Retry
QMessageBox.Ignore = QMessageBox.StandardButton.Ignore

from PyQt6.QtWidgets import QTableWidget

# QTableWidget编辑触发器枚举
QTableWidget.NoEditTriggers = QTableWidget.EditTrigger.NoEditTriggers
QTableWidget.CurrentChanged = QTableWidget.EditTrigger.CurrentChanged
QTableWidget.DoubleClicked = QTableWidget.EditTrigger.DoubleClicked
QTableWidget.SelectedClicked = QTableWidget.EditTrigger.SelectedClicked
QTableWidget.EditKeyPressed = QTableWidget.EditTrigger.EditKeyPressed
QTableWidget.AnyKeyPressed = QTableWidget.EditTrigger.AnyKeyPressed
QTableWidget.AllEditTriggers = QTableWidget.EditTrigger.AllEditTriggers

from PyQt6.QtWidgets import QTableView, QTreeView, QListView

# QTableView/QTreeView/QListView编辑触发器枚举（继承自QAbstractItemView）
QTableView.NoEditTriggers = QAbstractItemView.EditTrigger.NoEditTriggers
QTableView.CurrentChanged = QAbstractItemView.EditTrigger.CurrentChanged
QTableView.DoubleClicked = QAbstractItemView.EditTrigger.DoubleClicked
QTableView.SelectedClicked = QAbstractItemView.EditTrigger.SelectedClicked
QTableView.EditKeyPressed = QAbstractItemView.EditTrigger.EditKeyPressed
QTableView.AnyKeyPressed = QAbstractItemView.EditTrigger.AnyKeyPressed
QTableView.AllEditTriggers = QAbstractItemView.EditTrigger.AllEditTriggers

QTreeView.NoEditTriggers = QAbstractItemView.EditTrigger.NoEditTriggers
QTreeView.DoubleClicked = QAbstractItemView.EditTrigger.DoubleClicked
QTreeView.EditKeyPressed = QAbstractItemView.EditTrigger.EditKeyPressed

QListView.NoEditTriggers = QAbstractItemView.EditTrigger.NoEditTriggers
QListView.DoubleClicked = QAbstractItemView.EditTrigger.DoubleClicked
QListView.EditKeyPressed = QAbstractItemView.EditTrigger.EditKeyPressed

# QTableView/QTreeView选择行为和模式
QTableView.SelectRows = QAbstractItemView.SelectionBehavior.SelectRows
QTableView.SelectColumns = QAbstractItemView.SelectionBehavior.SelectColumns
QTableView.SelectItems = QAbstractItemView.SelectionBehavior.SelectItems
QTreeView.SelectRows = QAbstractItemView.SelectionBehavior.SelectRows
QTreeView.SelectItems = QAbstractItemView.SelectionBehavior.SelectItems
QListView.SelectItems = QAbstractItemView.SelectionBehavior.SelectItems

from PyQt6.QtCore import QEvent, QObject

# QEvent类型枚举
QEvent.Wheel = QEvent.Type.Wheel
QEvent.KeyPress = QEvent.Type.KeyPress
QEvent.KeyRelease = QEvent.Type.KeyRelease
QEvent.MouseButtonPress = QEvent.Type.MouseButtonPress
QEvent.MouseButtonRelease = QEvent.Type.MouseButtonRelease
QEvent.MouseButtonDblClick = QEvent.Type.MouseButtonDblClick
QEvent.MouseMove = QEvent.Type.MouseMove
QEvent.Resize = QEvent.Type.Resize
QEvent.Show = QEvent.Type.Show
QEvent.Hide = QEvent.Type.Hide
QEvent.Close = QEvent.Type.Close
QEvent.FocusIn = QEvent.Type.FocusIn
QEvent.FocusOut = QEvent.Type.FocusOut
QEvent.Enter = QEvent.Type.Enter
QEvent.Leave = QEvent.Type.Leave

# 延迟导入其他大型模块

# saas_contract_config_2026.json
# user_type and _users_cache are now in core.py / auth.py
_users_last_modified = 0



def _color_to_rgba(color, alpha):
    """将 QColor 或颜色字符串转换为 RGBA CSS 字符串。"""
    qcolor = QColor(color)
    return f"rgba({qcolor.red()}, {qcolor.green()}, {qcolor.blue()}, {alpha})"


def build_modern_glass_stylesheet(bg_color, fg_color, btn_bg, entry_bg, entry_fg, select_color, border_color, scale_percent=100):
    """构建统一的毛玻璃扁平化样式。"""
    accent_color = QColor(select_color)
    if accent_color.lightness() < 120:
        accent_color = accent_color.lighter(180)

    panel_color = QColor(entry_bg).lighter(102)
    button_color = QColor(btn_bg).lighter(101)
    window_color = QColor(bg_color).lighter(104)
    border_tint = QColor(border_color)
    if border_tint.lightness() < 150:
        border_tint = border_tint.lighter(170)

    scale_percent = normalize_ui_scale_percent(scale_percent)
    font_size = scale_ui_value(13, scale_percent)
    group_radius = scale_ui_value(12, scale_percent)
    group_margin_top = scale_ui_value(16, scale_percent)
    group_title_left = scale_ui_value(16, scale_percent)
    group_title_padding = scale_ui_value(10, scale_percent)
    button_radius = scale_ui_value(6, scale_percent)
    button_padding_v = scale_ui_value(6, scale_percent)
    button_padding_h = scale_ui_value(14, scale_percent)
    input_radius = scale_ui_value(6, scale_percent)
    input_padding_v = scale_ui_value(6, scale_percent)
    input_padding_h = scale_ui_value(12, scale_percent)
    combo_drop_width = scale_ui_value(24, scale_percent)
    combo_popup_padding = scale_ui_value(4, scale_percent)
    header_padding_v = scale_ui_value(8, scale_percent)
    header_padding_h = scale_ui_value(10, scale_percent)
    menu_item_padding_v = scale_ui_value(8, scale_percent)
    menu_item_padding_h = scale_ui_value(16, scale_percent)
    menu_item_radius = scale_ui_value(6, scale_percent)
    scrollbar_size = scale_ui_value(10, scale_percent)
    scrollbar_margin = scale_ui_value(4, scale_percent)
    scrollbar_radius = max(3, scale_ui_value(5, scale_percent))
    scrollbar_min_handle = scale_ui_value(24, scale_percent)
    checkbox_size = scale_ui_value(12, scale_percent)
    checkbox_radius = max(2, scale_ui_value(3, scale_percent))

    return f"""
        QWidget {{
            color: {fg_color.name()};
            font-family: 'Microsoft YaHei UI', 'Segoe UI';
            font-size: {font_size}px;
        }}
        QMainWindow, QDialog {{
            background-color: {_color_to_rgba(window_color, 244)};
        }}
        QWidget#glassRoot {{
            background: qlineargradient(
                x1:0, y1:0, x2:1, y2:1,
                stop:0 {_color_to_rgba(window_color.lighter(108), 255)},
                stop:0.55 {_color_to_rgba(button_color.lighter(104), 248)},
                stop:1 {_color_to_rgba(window_color.darker(103), 255)}
            );
        }}
        QFrame#navPanel {{
            background-color: {_color_to_rgba(panel_color, 176)};
            border-right: 1px solid {_color_to_rgba(border_tint, 175)};
        }}
        QStackedWidget#contentStack {{
            background-color: {_color_to_rgba(panel_color, 184)};
        }}
        QGroupBox, QTreeWidget, QTableWidget, QTableView, QListWidget, QMenu, QTabWidget::pane {{
            background-color: {_color_to_rgba(panel_color, 184)};
            border: 1px solid {_color_to_rgba(border_tint, 170)};
            border-radius: {group_radius}px;
        }}
        QGroupBox {{
            margin-top: {group_margin_top}px;
            padding-top: {button_padding_v}px;
            font-weight: 600;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: {group_title_left}px;
            padding: 0 {group_title_padding}px;
            color: {fg_color.name()};
        }}
        QPushButton, QToolButton {{
            background-color: {_color_to_rgba(button_color, 198)};
            color: {fg_color.name()};
            border: 1px solid {_color_to_rgba(border_tint, 165)};
            border-radius: {button_radius}px;
            padding: {button_padding_v}px {button_padding_h}px;
        }}
        QPushButton:hover, QToolButton:hover {{
            background-color: {_color_to_rgba(button_color.lighter(105), 228)};
            border-color: {_color_to_rgba(accent_color, 188)};
        }}
        QPushButton:pressed, QToolButton:pressed {{
            background-color: {_color_to_rgba(button_color.darker(104), 236)};
        }}
        QPushButton:checked, QToolButton:checked {{
            background-color: {_color_to_rgba(accent_color, 232)};
            border-color: {_color_to_rgba(accent_color, 245)};
            color: #ffffff;
        }}
        QPushButton:disabled, QToolButton:disabled {{
            color: {_color_to_rgba(fg_color, 120)};
            background-color: {_color_to_rgba(button_color, 120)};
            border-color: {_color_to_rgba(border_tint, 110)};
        }}
        QLineEdit, QTextEdit, QPlainTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit, QTimeEdit {{
            background-color: {_color_to_rgba(panel_color, 212)};
            color: {entry_fg.name()};
            border: 1px solid {_color_to_rgba(border_tint, 160)};
            border-radius: {input_radius}px;
            padding: {input_padding_v}px {input_padding_h}px;
            selection-background-color: {_color_to_rgba(accent_color, 225)};
            selection-color: #ffffff;
        }}
        QLineEdit:focus, QTextEdit:focus, QPlainTextEdit:focus, QComboBox:focus, QSpinBox:focus, QDoubleSpinBox:focus, QDateEdit:focus, QTimeEdit:focus {{
            border: 1px solid {_color_to_rgba(accent_color, 232)};
            background-color: {_color_to_rgba(QColor(entry_bg), 236)};
        }}
        QComboBox::drop-down {{
            border: none;
            width: {combo_drop_width}px;
            background: transparent;
        }}
        QComboBox QAbstractItemView {{
            background-color: {_color_to_rgba(panel_color, 236)};
            border: 1px solid {_color_to_rgba(border_tint, 190)};
            selection-background-color: {_color_to_rgba(accent_color, 220)};
            selection-color: #ffffff;
            padding: {combo_popup_padding}px;
        }}
        QTableWidget, QTableView, QListWidget, QTreeWidget {{
            alternate-background-color: {_color_to_rgba(window_color.lighter(102), 116)};
            gridline-color: {_color_to_rgba(border_tint, 124)};
            selection-background-color: {_color_to_rgba(accent_color, 220)};
            selection-color: #ffffff;
            outline: none;
        }}
        QHeaderView::section {{
            background-color: {_color_to_rgba(button_color, 170)};
            color: {fg_color.name()};
            border: none;
            border-bottom: 1px solid {_color_to_rgba(border_tint, 170)};
            padding: {header_padding_v}px {header_padding_h}px;
            font-weight: 600;
        }}
        QTableCornerButton::section {{
            background: transparent;
            border: none;
        }}
        QMenu::item {{
            padding: {menu_item_padding_v}px {menu_item_padding_h}px;
            border-radius: {menu_item_radius}px;
        }}
        QMenu::item:selected {{
            background-color: {_color_to_rgba(accent_color, 215)};
            color: #ffffff;
        }}
        QScrollBar:vertical {{
            background: transparent;
            width: {scrollbar_size}px;
            margin: {scrollbar_margin}px 0 {scrollbar_margin}px 0;
        }}
        QScrollBar::handle:vertical {{
            background: {_color_to_rgba(border_tint, 182)};
            border-radius: {scrollbar_radius}px;
            min-height: {scrollbar_min_handle}px;
        }}
        QScrollBar:horizontal {{
            background: transparent;
            height: {scrollbar_size}px;
            margin: 0 {scrollbar_margin}px 0 {scrollbar_margin}px;
        }}
        QScrollBar::handle:horizontal {{
            background: {_color_to_rgba(border_tint, 182)};
            border-radius: {scrollbar_radius}px;
            min-width: {scrollbar_min_handle}px;
        }}
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical,
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal,
        QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical,
        QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
            background: transparent;
            border: none;
        }}
        QCheckBox::indicator {{
            width: {checkbox_size}px;
            height: {checkbox_size}px;
            border-radius: {checkbox_radius}px;
            border: 1px solid {_color_to_rgba(border_tint, 200)};
            background-color: {_color_to_rgba(QColor(entry_bg), 235)};
        }}
        QCheckBox::indicator:checked {{
            background-color: {_color_to_rgba(accent_color, 235)};
            border: 1px solid {_color_to_rgba(accent_color, 245)};
        }}
    """


def get_glass_groupbox_style(compact=False, scale_percent=None):
    """局部区域使用的统一玻璃卡片样式。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else scale_percent
    radius = scale_ui_value(14 if not compact else 10, effective_percent)
    margin_top = scale_ui_value(18 if not compact else 14, effective_percent)
    title_left = scale_ui_value(18 if not compact else 14, effective_percent)
    title_padding = scale_ui_value(10 if not compact else 8, effective_percent)
    return f"""
        QGroupBox {{
            background-color: rgba(255, 255, 255, 176);
            border: 1px solid rgba(193, 208, 224, 210);
            border-radius: {radius}px;
            margin-top: {margin_top}px;
            font-weight: 600;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: {title_left}px;
            padding: 0 {title_padding}px;
            color: #16324f;
        }}
    """


def get_glass_tree_style(scale_percent=None):
    """设置页左侧菜单的玻璃感样式。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else scale_percent
    radius = scale_ui_value(12, effective_percent)
    padding_v = scale_ui_value(8, effective_percent)
    padding_h = scale_ui_value(6, effective_percent)
    item_height = scale_ui_value(32, effective_percent)
    item_padding_v = scale_ui_value(4, effective_percent)
    item_padding_h = scale_ui_value(10, effective_percent)
    item_radius = scale_ui_value(6, effective_percent)
    return f"""
        QTreeWidget {{
            border: 1px solid rgba(193, 208, 224, 210);
            border-radius: {radius}px;
            background-color: rgba(255, 255, 255, 168);
            padding: {padding_v}px {padding_h}px;
            outline: none;
        }}
        QTreeWidget::item {{
            min-height: {item_height}px;
            padding: {item_padding_v}px {item_padding_h}px;
            border-radius: {item_radius}px;
            background: transparent;
        }}
        QTreeWidget::item:hover {{
            background-color: rgba(91, 141, 239, 0.10);
        }}
        QTreeWidget::item:selected {{
            background-color: rgba(91, 141, 239, 0.88);
            color: #ffffff;
            font-weight: 600;
        }}
        QTreeWidget::item:has-children {{
            font-weight: 600;
            color: #17324d;
        }}
        QTreeWidget::branch {{
            background: transparent;
            border: none;
        }}
        QTreeWidget::branch:has-children:closed,
        QTreeWidget::branch:has-children:open {{
            image: none;
        }}
    """


def get_nav_button_style(scale_percent=None):
    """左侧导航按钮样式。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else scale_percent
    padding_v = scale_ui_value(10, effective_percent)
    padding_h = scale_ui_value(16, effective_percent)
    radius = scale_ui_value(6, effective_percent)
    font_size = scale_ui_value(13, effective_percent)
    return f"""
        QPushButton {{
            text-align: left;
            padding: {padding_v}px {padding_h}px;
            border: 1px solid rgba(193, 208, 224, 0.90);
            border-radius: {radius}px;
            font-size: {font_size}px;
            font-weight: 500;
            background-color: rgba(255, 255, 255, 0.30);
            color: #17324d;
        }}
        QPushButton:hover {{
            background-color: rgba(255, 255, 255, 0.48);
            border-color: rgba(91, 141, 239, 0.50);
        }}
        QPushButton:checked {{
            background-color: rgba(91, 141, 239, 0.92);
            color: white;
            border-color: rgba(91, 141, 239, 1.0);
            font-weight: 600;
        }}
    """


def get_toggle_nav_button_style(scale_percent=None):
    """导航折叠按钮样式。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else scale_percent
    padding = scale_ui_value(8, effective_percent)
    radius = scale_ui_value(6, effective_percent)
    font_size = scale_ui_value(12, effective_percent)
    return f"""
        QPushButton {{
            padding: {padding}px;
            border: 1px solid rgba(193, 208, 224, 0.90);
            border-radius: {radius}px;
            background: rgba(255, 255, 255, 0.42);
            font-size: {font_size}px;
            color: #17324d;
        }}
        QPushButton:hover {{
            background-color: rgba(255, 255, 255, 0.58);
            border-color: rgba(91, 141, 239, 0.50);
        }}
    """


def get_password_input_frame_style(scale_percent=None):
    """密码输入容器样式。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else scale_percent
    radius = scale_ui_value(6, effective_percent)
    return f"""
        QFrame {{
            border: 1px solid rgba(193, 208, 224, 0.90);
            border-radius: {radius}px;
            background-color: rgba(255, 255, 255, 0.78);
        }}
    """


def apply_color_chip_style(button, color_hex, scale_percent=None):
    """统一颜色选择按钮的圆角玻璃样式。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else scale_percent
    radius = scale_ui_value(6, effective_percent)
    button.setProperty("_chip_color_hex", color_hex)
    button.setStyleSheet(
        f"""
        QPushButton {{
            background-color: {color_hex};
            border: 1px solid rgba(193, 208, 224, 0.90);
            border-radius: {radius}px;
        }}
        QPushButton:hover {{
            border-color: rgba(91, 141, 239, 0.70);
        }}
        """
    )


def get_compact_refresh_button_style(scale_percent=None):
    """文件生成和 CRM 页面使用的紧凑刷新按钮样式。"""
    effective_percent = get_current_ui_scale_percent() if scale_percent is None else scale_percent
    radius = scale_ui_value(4, effective_percent)
    font_size = scale_ui_value(13, effective_percent)
    padding_h = scale_ui_value(8, effective_percent)
    return f"""
        QPushButton {{
            background-color: #0078D7;
            color: white;
            border-radius: {radius}px;
            font-weight: 600;
            font-size: {font_size}px;
            padding: 0 {padding_h}px;
        }}
        QPushButton:hover {{
            background-color: #0063b1;
        }}
        QPushButton:pressed {{
            background-color: #005a9e;
        }}
    """


def set_dialog_opacity(dialog):
    """设置对话框的透明度和主题样式"""
    try:
        manual_stylesheet = dialog.property("_manual_dialog_stylesheet")
        if manual_stylesheet is None:
            manual_stylesheet = dialog.styleSheet() or ""
            dialog.setProperty("_manual_dialog_stylesheet", manual_stylesheet)

        # 从配置文件加载透明度设置
        config = load_config()
        
        # 判断是主窗口还是设置窗口，加载不同的透明度配置
        is_main_window = isinstance(dialog, QMainWindow)
        dialog_class_name = type(dialog).__name__
        is_settings_dialog = dialog_class_name == 'SettingsDialog' or dialog_class_name == 'SettingsDialogFromUI'
        
        if is_main_window:
             opacity_value = config.get('app_settings', {}).get('main_opacity', 100)
        elif is_settings_dialog:
             opacity_value = config.get('app_settings', {}).get('opacity', 100)
        else:
             # 其他所有弹窗（提醒弹窗、文件选择等）
             opacity_value = config.get('app_settings', {}).get('dialog_opacity', 100)
             
        # 检查值的范围，如果是百分比（0-100），转换为0-1
        if 0 <= opacity_value <= 100:
            opacity = opacity_value / 100.0
        elif 0 <= opacity_value <= 255:
            opacity = opacity_value / 255.0
        else:
            opacity = 1.0
        
        # 设置对话框透明度
        pass  # opacity disabled
        
        # 设置主题样式
        theme_mode = config.get('app_settings', {}).get('theme_mode', 'light')
        
        if theme_mode == 'custom':
            # 自定义主题设置
            custom_theme = config.get('app_settings', {}).get('custom_theme', {})
            bg_color = QColor(custom_theme.get('bg_color', '#ffffff'))    # 主背景
            fg_color = QColor(custom_theme.get('fg_color', '#333333'))    # 主文本
            btn_bg = QColor(custom_theme.get('btn_bg', '#ffffff'))      # 按钮背景
            entry_bg = QColor(custom_theme.get('entry_bg', '#ffffff'))    # 输入框背景
            entry_fg = QColor(custom_theme.get('entry_fg', '#333333'))    # 输入框文本
            select_color = QColor(custom_theme.get('select_color', '#f7f7f7'))# 选中态背景
            border_color = QColor(custom_theme.get('border_color', '#333333'))# 边框颜色
        else:
            # 浅色模式设置 (Trae风格)
            bg_color = QColor('#ffffff')    # 主背景
            fg_color = QColor('#333333')    # 主文本
            btn_bg = QColor('#ffffff')      # 按钮背景
            entry_bg = QColor('#ffffff')    # 输入框背景
            entry_fg = QColor('#333333')    # 输入框文本
            select_color = QColor('#5b8def')# 选中态背景
            border_color = QColor('#d7e0ec')# 边框颜色
            
        # 设置调色板
        palette = QPalette()
        palette.setColor(QPalette.Window, bg_color)
        palette.setColor(QPalette.WindowText, fg_color)
        palette.setColor(QPalette.Base, entry_bg)
        palette.setColor(QPalette.AlternateBase, bg_color)
        palette.setColor(QPalette.ToolTipBase, fg_color)
        palette.setColor(QPalette.ToolTipText, fg_color)
        palette.setColor(QPalette.Text, entry_fg)
        palette.setColor(QPalette.Button, btn_bg)
        palette.setColor(QPalette.ButtonText, fg_color)
        palette.setColor(QPalette.BrightText, QColor('#ffffff'))
        palette.setColor(QPalette.Link, QColor('#0078d7'))
        # 设置高亮颜色
        palette.setColor(QPalette.Highlight, select_color)
        # 根据背景颜色自动调整高亮文本颜色
        if bg_color.lightness() < 128:
            # 深色背景使用白色文本
            palette.setColor(QPalette.HighlightedText, QColor('#ffffff'))
        else:
            # 浅色背景使用黑色文本
            palette.setColor(QPalette.HighlightedText, QColor('#000000'))

        scale_percent = dialog.property("_ui_scale_percent")
        if scale_percent is None:
            app = QApplication.instance()
            scale_percent = app.property("_global_ui_scale_percent") if app is not None else 100
        scale_percent = normalize_ui_scale_percent(scale_percent, default=100)
        
        dialog.setPalette(palette)
        dialog.setStyleSheet(
            build_modern_glass_stylesheet(
                bg_color=bg_color,
                fg_color=fg_color,
                btn_bg=btn_bg,
                entry_bg=entry_bg,
                entry_fg=entry_fg,
                select_color=select_color,
                border_color=border_color,
                scale_percent=scale_percent,
            ) + (f"\n{manual_stylesheet}" if manual_stylesheet else "")
        )
        dialog.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        dialog.setProperty("_theme_mode", theme_mode)
        dialog.update()
    except Exception as e:
        logging.error(f"设置对话框透明度和样式失败: {str(e)}")


def get_standard_line_edit_height():
    """获取当前样式下标准输入框高度。"""
    probe = QLineEdit()
    probe.ensurePolished()
    return max(24, probe.sizeHint().height())


def apply_password_input_height(line_edit, container=None, toggle_button=None):
    """让密码输入框高度与普通用户名输入框保持一致。"""
    target_height = get_standard_line_edit_height()
    if container is not None:
        container.setFixedHeight(target_height)
        line_edit.setFixedHeight(max(18, target_height - 4))
    else:
        line_edit.setFixedHeight(target_height)

    if toggle_button is not None:
        if container is not None:
            button_size = max(18, target_height - 6)
            toggle_button.setFixedSize(button_size, button_size)
        else:
            toggle_button.setFixedHeight(target_height)


def _get_screen_for_widget(widget=None):
    """获取控件所在的屏幕对象。"""
    app = QApplication.instance()
    if widget is not None:
        try:
            screen = QApplication.screenAt(widget.mapToGlobal(widget.rect().center()))
            if screen is not None:
                return screen
        except Exception:
            pass
        if hasattr(widget, 'screen'):
            try:
                screen = widget.screen()
                if screen is not None:
                    return screen
            except Exception:
                pass
    return app.primaryScreen() if app is not None else None


def center_dialog(dialog, reference_widget=None):
    """将弹窗居中到参考窗口或当前屏幕。"""
    dialog.adjustSize()
    screen = _get_screen_for_widget(reference_widget or dialog.parentWidget() or dialog)
    if screen is None:
        return

    available_geometry = screen.availableGeometry()
    center_point = available_geometry.center()
    if reference_widget is not None:
        try:
            center_point = reference_widget.mapToGlobal(reference_widget.rect().center())
        except Exception:
            center_point = available_geometry.center()

    popup_width = dialog.width()
    popup_height = dialog.height()
    min_x = available_geometry.left()
    max_x = available_geometry.left() + max(0, available_geometry.width() - popup_width)
    min_y = available_geometry.top()
    max_y = available_geometry.top() + max(0, available_geometry.height() - popup_height)
    x = center_point.x() - popup_width // 2
    y = center_point.y() - popup_height // 2
    dialog.move(min(max(x, min_x), max_x), min(max(y, min_y), max_y))


def position_dialog_below_widget(dialog, anchor_widget, margin=6):
    """将弹窗定位到控件下方，超出屏幕时自动调整。"""
    if anchor_widget is None:
        center_dialog(dialog)
        return

    dialog.adjustSize()
    screen = _get_screen_for_widget(anchor_widget)
    if screen is None:
        center_dialog(dialog, anchor_widget)
        return

    available_geometry = screen.availableGeometry()
    popup_width = dialog.width()
    popup_height = dialog.height()

    anchor_rect = anchor_widget.rect()
    anchor_top_left = anchor_widget.mapToGlobal(anchor_rect.topLeft())
    anchor_bottom_left = anchor_widget.mapToGlobal(anchor_rect.bottomLeft())
    anchor_bottom_right = anchor_widget.mapToGlobal(anchor_rect.bottomRight())

    min_x = available_geometry.left()
    max_x = available_geometry.left() + max(0, available_geometry.width() - popup_width)
    min_y = available_geometry.top()
    max_y = available_geometry.top() + max(0, available_geometry.height() - popup_height)

    x = anchor_top_left.x()
    if x > max_x:
        x = anchor_bottom_right.x() - popup_width
    x = min(max(x, min_x), max_x)

    y = anchor_bottom_left.y() + margin
    if y > max_y:
        y = anchor_top_left.y() - popup_height - margin
    y = min(max(y, min_y), max_y)

    dialog.move(x, y)



# === Table utilities ===
class TableCellEditMenu(QObject):
    """为 QTableWidget 单元格添加右键编辑菜单"""

    def __init__(self, table, parent=None):
        super().__init__(parent)
        self._table = table
        table.installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.ContextMenu:
            # ✅ 转换坐标：event.pos() 是 table 坐标，itemAt 需要 viewport 坐标
            viewport = self._table.viewport()
            pos = viewport.mapFrom(obj, event.pos())
            item = self._table.itemAt(pos)
            if item is not None:
                if not (item.flags() & Qt.ItemFlag.ItemIsEditable):
                    return False
                row, col = item.row(), item.column()
                # ✅ 先选中当前单元格，避免编辑跨行
                self._table.setCurrentCell(row, col)
                self._table.clearSelection()
                self._table.selectRow(row)
                menu = QMenu(self._table)
                edit_action = menu.addAction("✏️ 编辑")
                action = menu.exec(event.globalPos())
                if action == edit_action:
                    self._table.editItem(item)
                    return True
        return False


def _copy_selected_cells(table):
    """复制 QTableWidget 中选中的单元格内容到剪贴板，Tab 分隔列，换行分隔行。"""
    from PyQt6.QtWidgets import QApplication, QComboBox
    ranges = table.selectedRanges()
    if not ranges:
        return
    selected = set()
    min_row, max_row = float('inf'), -1
    min_col, max_col = float('inf'), -1
    for rng in ranges:
        for row in range(rng.topRow(), rng.bottomRow() + 1):
            for col in range(rng.leftColumn(), rng.rightColumn() + 1):
                selected.add((row, col))
                min_row = min(min_row, row)
                max_row = max(max_row, row)
                min_col = min(min_col, col)
                max_col = max(max_col, col)

    lines = []
    for row in range(min_row, max_row + 1):
        parts = []
        for col in range(min_col, max_col + 1):
            if (row, col) in selected:
                item = table.item(row, col)
                if item:
                    parts.append(item.text())
                else:
                    w = table.cellWidget(row, col)
                    if w:
                        cb = w.findChild(QComboBox)
                        if cb:
                            parts.append(cb.currentText())
                        else:
                            parts.append('')
                    else:
                        parts.append('')
            else:
                parts.append('')
        lines.append('\t'.join(parts))
    QApplication.clipboard().setText('\n'.join(lines))


def install_table_copy_handler(table):
    """为 QTableWidget 安装 Ctrl+C 多选复制功能（防止重复安装）"""
    if hasattr(table, '_copy_handler_installed'):
        return
    table._copy_handler_installed = True
    orig_key_press = table.keyPressEvent

    def _handler(event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier and event.key() == Qt.Key.Key_C:
            _copy_selected_cells(table)
            return
        orig_key_press(event)

    table.keyPressEvent = _handler


def install_table_edit_context_menu(table):
    """为指定 QTableWidget 安装右键编辑菜单和 Ctrl+C 多选复制（防止重复安装）"""
    if not hasattr(table, '_edit_menu_installed'):
        table._edit_menu_installed = True
        TableCellEditMenu(table, table)
    install_table_copy_handler(table)


def install_header_alignment_menu(table, config_key, repopulate_callback=None):
    """为 QTableWidget 的水平表头安装右键对齐菜单。"""
    if hasattr(table, '_header_align_menu_installed'):
        return
    table._header_align_menu_installed = True
    header = table.horizontalHeader()

    # 保存原始的 contextMenuEvent
    _orig_context_menu_event = header.contextMenuEvent

    def _new_context_menu_event(event):
        from PyQt6.QtWidgets import QMenu
        pos = event.pos()
        logical_index = header.logicalIndexAt(pos)
        if logical_index < 0:
            _orig_context_menu_event(event)
            return
        header_item = table.horizontalHeaderItem(logical_index)

# === UI component classes ===
class _DialogOutsideCloseFilter(QObject):
    """用于点击弹层外部区域时关闭对话框的事件过滤器。"""
    def __init__(self, dialog):
        """初始化对话框外部关闭过滤。"""
        super().__init__(dialog)
        self._dialog = dialog

    def eventFilter(self, watched, event):
        """拦截并处理目标事件。"""
        dialog = self._dialog
        if dialog is None:
            return False
        try:
            if not dialog.isVisible() or not getattr(dialog, '_outside_close_armed', False):
                return False
        except RuntimeError:
            # C++ 对象已销毁
            return False
        if event.type() != QEvent.Type.MouseButtonPress:
            return False

        active_modal = QApplication.activeModalWidget()
        active_popup = QApplication.activePopupWidget()
        if active_modal is not None and active_modal is not dialog:
            return False
        if active_popup is not None and active_popup is not dialog:
            return False

        watched_widget = watched if hasattr(watched, 'parentWidget') else None
        if watched_widget is dialog or (watched_widget is not None and dialog.isAncestorOf(watched_widget)):
            return False

        global_pos = None
        if hasattr(event, 'globalPosition'):
            global_pos = event.globalPosition().toPoint()
        elif hasattr(event, 'globalPos'):
            global_pos = event.globalPos()

        if global_pos is not None and dialog.frameGeometry().contains(global_pos):
            return False

        try:
            dialog.reject()
        except RuntimeError:
            pass
        return False


class CenteredPopupDialog(QDialog):
    """
    统一处理居中显示和点外部关闭的弹窗基类。

    【重要】已知Bug及修复方案（2026-05-10）：
    ========================================
    问题描述：
    --------
    当继承此类的对话框（如ExcelFieldSettingsDialog）调用accept()后，
    对话框关闭时会触发WindowDeactivate事件，导致event()方法自动调用reject()。
    这会覆盖accept()的结果，使dialog.exec()返回Rejected而非Accepted，
    导致所有字段设置代码不执行，表现为"字段设置无法生效"。

    问题复现场景：
    ------------
    1. 用户点击"确定"按钮
    2. 触发 accept() 方法开始关闭对话框
    3. 窗口失活触发 QEvent.Type.WindowDeactivate 事件
    4. event() 捕获事件并自动调用 self.reject()
    5. reject() 覆盖了 accept() 的返回值
    6. dialog.exec() 返回 Rejected（错误）
    7. 字段设置代码被跳过，表格不更新

    解决方案：
    --------
    使用 _dialog_closed 标志位确保 accept()/reject() 只执行一次：
    - 在 __init__ 中初始化 _dialog_closed = False
    - 在 accept() 或 reject() 被调用时设置 _dialog_closed = True
    - 在 event() 中检查标志位，如果已关闭则跳过外部关闭逻辑

    影响范围：
    --------
    所有继承 CenteredPopupDialog 的对话框类：
    - ExcelFieldSettingsDialog（字段设置对话框）
    - CustomReportFilterDialog（自定义报表筛选弹窗）
    - 其他可能的子类

    测试验证：
    --------
    文件生成、CRM订单、对象查询三个模块的字段设置功能均已恢复正常。

    维护提醒：
    --------
    如果未来需要修改此类的关闭逻辑，请注意保持此标志位机制，
    否则可能导致同样的字段设置失效问题复发。
    """

    def __init__(self, parent=None, center_reference=None, close_on_outside=True):
        """
        初始化居中弹出对话框。

        参数：
        ------
        parent : QWidget, optional
            父窗口
        center_reference : QWidget, optional
            居中参考控件（默认使用parent）
        close_on_outside : bool, optional
            是否支持点击外部关闭（默认True）

        注意：
        ----
        _dialog_closed 标志位用于防止 accept()/reject() 被重复调用。
        这是修复"字段设置无法生效"问题的关键机制。
        """
        super().__init__(parent)
        self._center_reference_widget = center_reference or parent
        self._popup_anchor_widget = None
        self._popup_anchor_margin = 6
        self._close_on_outside = close_on_outside
        self._outside_close_armed = False
        self._outside_close_filter = _DialogOutsideCloseFilter(self) if close_on_outside else None
        self._outside_close_filter_installed = False
        # ✅【关键修复】防止重复关闭标志位
        # 用途：当 accept() 或 reject() 被显式调用后，设置为 True
        # 作用：防止 event() 中的 WindowDeactivate 处理逻辑再次调用 reject()
        # 问题：没有此标志时，点击确定后 event() 会调用 reject() 覆盖 accept() 结果
        # 结果：导致 dialog.exec() 返回错误状态，字段设置代码不执行
        self._dialog_closed = False

        # 浅色背景（防止默认黑色背景）
        self.setStyleSheet("QDialog { background-color: #FAFAFA; }")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)

    def set_center_reference_widget(self, widget):
        """设置centerreferencewidget。"""
        self._center_reference_widget = widget

    def set_popup_anchor_widget(self, widget, margin=6):
        """设置弹层anchorwidget。"""
        self._popup_anchor_widget = widget
        self._popup_anchor_margin = margin

    def clear_popup_anchor_widget(self):
        """清空弹层anchorwidget。"""
        self._popup_anchor_widget = None

    def showEvent(self, event):
        """处理控件显示事件。"""
        super().showEvent(event)
        if self._popup_anchor_widget is not None:
            position_dialog_below_widget(self, self._popup_anchor_widget, self._popup_anchor_margin)
        else:
            center_dialog(self, self._center_reference_widget or self.parentWidget())
        if self._close_on_outside:
            self._outside_close_armed = False
            self._install_outside_close_filter()
            QTimer.singleShot(0, self._arm_outside_close)

    def hideEvent(self, event):
        """处理控件隐藏事件。"""
        self._outside_close_armed = False
        self._remove_outside_close_filter()
        super().hideEvent(event)

    def event(self, event):
        """
        处理Qt事件（核心修复点）。

        【重要】此方法是"字段设置无法生效"Bug的关键位置：
        ------------------------------------------------
        原始问题：
        当用户点击对话框的"确定"按钮时，执行流程如下：
        1. 按钮点击 → 触发 accept() 方法
        2. accept() 调用 super().accept() 开始关闭对话框
        3. 对话框窗口失活 → 触发 QEvent.Type.WindowDeactivate 事件
        4. Qt事件循环调用本 event() 方法处理该事件
        5. 本方法检测到 WindowDeactivate + 外部关闭启用
        6. 自动调用 self.reject()
        7. reject() 将对话框结果设置为 Rejected，**覆盖** accept() 的 Accepted 结果
        8. dialog.exec() 返回 QDialog.DialogCode.Rejected（错误！）
        9. 调用方的 if dialog.exec() == Accepted: 条件不满足
        10. 所有字段设置代码被跳过不执行

        修复机制：
        通过 _dialog_closed 标志位打破上述恶性循环：
        - 当 accept() 或 reject() 被显式调用时，设置 _dialog_closed = True
        - 本方法在入口处检查标志位，如果已关闭则直接返回
        - 避免在 accept() 之后再次调用 reject()

        参数：
        ------
        event : QEvent
            Qt事件对象

        返回值：
        ------
        bool
            是否已处理事件
        """

        # ✅【关键防御】检查对话框是否已经通过 accept()/reject() 关闭
        # 使用 getattr 安全访问，防止 __init__ 未完成时触发事件导致 AttributeError
        if getattr(self, '_dialog_closed', False):
            return super().event(event)

        # 检查是否需要外部关闭逻辑
        if (
            self._close_on_outside                          # 启用了外部关闭功能
            and self.isVisible()                            # 对话框仍然可见
            and self._outside_close_armed                    # 外部关闭已激活（show后延迟激活）
            and event.type() == QEvent.Type.WindowDeactivate # 窗口失活事件
        ):
            # 检查是否有其他模态/弹出窗口活动
            active_modal = QApplication.activeModalWidget()
            active_popup = QApplication.activePopupWidget()
            if active_modal is not None and active_modal is not self:
                return super().event(event)
            if active_popup is not None and active_popup is not self:
                return super().event(event)

            # ✅【关键操作】标记为已关闭，然后调用 reject()
            # 注意：这里必须先设置标志再调用 reject()，
            # 否则 reject() 可能触发其他事件导致递归进入本方法
            self._dialog_closed = True
            self.reject()
            return True

        return super().event(event)

    def _arm_outside_close(self):
        """内部方法：处理arm外部关闭逻辑。"""
        if self.isVisible():
            self._outside_close_armed = True

    def _install_outside_close_filter(self):
        """内部方法：处理install外部关闭过滤逻辑。"""
        app = QApplication.instance()
        if app is None or self._outside_close_filter is None or self._outside_close_filter_installed:
            return
        app.installEventFilter(self._outside_close_filter)
        self._outside_close_filter_installed = True

    def _remove_outside_close_filter(self):
        """内部方法：处理移除外部关闭过滤逻辑。"""
        app = QApplication.instance()
        if app is None or self._outside_close_filter is None or not self._outside_close_filter_installed:
            return
        app.removeEventFilter(self._outside_close_filter)
        self._outside_close_filter_installed = False


class CustomMessageBox(CenteredPopupDialog):
    """自定义消息框组件。"""
    def __init__(self, title, message, type="info", parent=None):
        """初始化自定义消息框。"""
        super().__init__(parent, center_reference=parent)
        self.setWindowTitle(title)
        self.result = None
        
        # Load size from config
        config = load_config()
        size_config = config.get('app_settings', {}).get('dialog_sizes', {}).get('CustomMessageBox', None)
        
        if size_config:
            width, height = size_config
        else:
            # Determine window size based on message length
            width = 320
            height = 160
            if len(message) > 50:
                width = 420
                height = 220

        self.resize(width, height)
        self.setMinimumSize(200, 100)
        
        # Apply theme and opacity
        set_dialog_opacity(self)
        
        # Content
        main_layout = QHBoxLayout()
        
        # Icon
        icon_symbol = "ℹ️"
        icon_fg = QColor('#0078d7')
        if type == "error":
            icon_symbol = "❌"
            icon_fg = QColor('#e81123')
        elif type == "warning":
            icon_symbol = "⚠️"
            icon_fg = QColor('#ffb900')
        elif type == "yesno":
            icon_symbol = "❓"
            icon_fg = QColor('#0078d7')
            
        icon_label = QLabel(icon_symbol)
        icon_label.setFont(QFont("Segoe UI Emoji", 24))
        icon_label.setStyleSheet(f'color: {icon_fg.name()};')
        main_layout.addWidget(icon_label, 0, Qt.AlignTop)
        main_layout.addSpacing(15)
        
        msg_label = QLabel(message)
        msg_label.setWordWrap(True)
        msg_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        main_layout.addWidget(msg_label, 1)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setAlignment(Qt.AlignRight)
        
        if type == "yesno":
            yes_btn = QPushButton("是")
            yes_btn.setFixedWidth(80)
            yes_btn.clicked.connect(self.on_yes)
            btn_layout.addWidget(yes_btn)
            
            no_btn = QPushButton("否")
            no_btn.setFixedWidth(80)
            no_btn.clicked.connect(self.on_no)
            btn_layout.addWidget(no_btn)
            btn_layout.insertSpacing(1, 10)
        else:
            ok_btn = QPushButton("确定")
            ok_btn.setFixedWidth(80)
            ok_btn.clicked.connect(self.on_ok)
            btn_layout.addWidget(ok_btn)
            
        # Main layout
        layout = QVBoxLayout()
        layout.addLayout(main_layout)
        layout.addLayout(btn_layout)
        layout.setContentsMargins(20, 20, 20, 20)
        
        self.setLayout(layout)
        
        # Bind enter/escape
        self.setFocusPolicy(Qt.StrongFocus)
        
    def keyPressEvent(self, event):
        """处理键盘事件。"""
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            self.on_ok()
        elif event.key() == Qt.Key_Escape:
            self.on_no()
        super().keyPressEvent(event)

    def on_size(self, event):
        """响应大小相关操作。"""
        if event.type() == event.Resize:
            config = load_config()
            if 'app_settings' not in config: config['app_settings'] = {}
            if 'dialog_sizes' not in config['app_settings']: config['app_settings']['dialog_sizes'] = {}
            
            # Save only if changed
            new_size = [self.width(), self.height()]
            if config['app_settings']['dialog_sizes'].get('CustomMessageBox') != new_size:
                config['app_settings']['dialog_sizes']['CustomMessageBox'] = new_size
                save_config_with_delay(config)

    def on_ok(self):
        """响应确认相关操作。"""
        self.result = True
        self.accept()
        
    def on_yes(self):
        """响应确认相关操作。"""
        self.result = True
        self.accept()
        
    def on_no(self):
        """响应取消相关操作。"""
        self.result = False
        self.reject()


class UIToolkit:
    """
    通用UI工具函数库（静态方法集合）

    【设计目的】：
    将各模块中重复的功能代码提取为通用函数，提高代码复用性和可维护性。
    所有模块（文件生成、CRM订单、对象查询、自定义报表等）都可以直接调用这些方法。

    【包含功能】：
    1. 字段设置对话框（open_field_settings）
    2. 对象查询下拉框刷新（refresh_obj_query_combo）
    3. CRM对象管理即时生效（save_crm_objects_immediately）
    4. 表格列重建（rebuild_table_columns）
    5. 数据分页刷新（refresh_table_page）

    【使用方式】：
    from UIToolkit import UIToolkit

    # 打开字段设置对话框
    visible_headers, header_order = UIToolkit.open_field_settings(
        all_headers=all_headers,
        current_visible=current_visible,
        parent=self
    )

    # 刷新对象查询下拉框
    UIToolkit.refresh_obj_query_combo(main_window)

    【版本历史】：
    - 2026-05-10：初始创建，从各模块提取重复代码
    """

    @staticmethod
    def open_field_settings(all_headers, current_visible=None, current_order=None, parent=None, title="字段设置"):
        """
        通用的字段设置对话框打开方法

        【功能说明】：
        统一封装 ExcelFieldSettingsDialog 的调用逻辑，
        避免每个模块重复编写相同的对话框打开和结果处理代码。

        【适用模块】：
        - 文件生成模块（open_column_settings_dialog）
        - CRM订单模块（open_crm_field_settings_dialog）
        - 对象查询模块（_open_obj_query_column_settings）
        - 自定义报表模块（open_custom_report_filter_dialog）
        - 其他需要字段显示/隐藏功能的场景

        参数：
        ------
        all_headers : list[str]
            所有可能的字段名称列表（完整字段集）
        current_visible : list[str], optional
            当前可见的字段列表（默认为全部可见）
        current_order : list[str], optional
            当前的字段顺序（默认为all_headers的顺序）
        parent : QWidget, optional
            父窗口（用于居中显示和模态控制）
        title : str, optional
            对话框标题（默认为"字段设置"）

        返回值：
        -------
        tuple[list[str], list[str]] | None
            成功时返回 (visible_headers, header_order)
            用户取消或无可见字段时返回 None

        使用示例：
        --------
        # 在文件生成模块中：
        result = UIToolkit.open_field_settings(
            all_headers=self.all_headers,
            current_visible=self.current_visible_headers,
            parent=self
        )
        if result:
            new_vis, new_order = result
            self.current_visible_headers = new_vis
            self._refresh_current_table_page()

        注意事项：
        ----------
        1. 此方法已内置 accept/reject 冲突修复（_dialog_closed 标志位）
        2. 返回值为元组或None，便于判断用户是否确认
        3. 建议在调用后立即处理返回值，避免状态不一致
        """

        if not all_headers:
            print("[UIToolkit] ⚠️ open_field_settings: all_headers 为空")
            return None

        # 创建并显示字段设置对话框
        dialog = ExcelFieldSettingsDialog(
            headers=all_headers,
            visible_headers=list(current_visible or all_headers),
            header_order=list(current_order or all_headers),
            parent=parent
        )

        # 显示对话框并等待用户操作
        result = dialog.exec()

        # 检查返回结果
        if result == QDialog.DialogCode.Accepted:
            # 用户点击了"确定"
            visible_headers, header_order = dialog.get_settings()

            # 验证返回值有效性
            if not visible_headers:
                print("[UIToolkit] ⚠️ open_field_settings: 返回的visible_headers为空")
                return None

            print(f"[UIToolkit] ✅ 字段设置完成 | 可见字段数: {len(visible_headers)} | 前5项: {visible_headers[:5]}")
            return visible_headers, header_order
        else:
            # 用户取消或关闭对话框
            print(f"[UIToolkit] ℹ️ 字段设置已取消")
            return None

    @staticmethod
    def refresh_obj_query_combo(main_window):
        """
        刷新主窗口对象查询模块的对象下拉框

        【功能说明】：
        从最新配置加载CRM对象列表，更新到对象查询模块的下拉框组件。
        用于在添加/删除CRM对象后同步UI显示。

        【调用时机】：
        - 添加新CRM对象后
        - 删除CRM对象后
        - 修改CRM对象配置后
        - 需要手动强制刷新时

        参数：
        ------
        main_window : QMainWindow
            主窗口实例，必须包含 obj_query_object_combo 属性

        返回值：
        -------
        bool
            刷新成功返回True，失败返回False

        工作流程：
        ----------
        1. 检查main_window是否有效
        2. 获取obj_query_object_combo引用
        3. 记录当前选中状态
        4. 从load_config()加载最新配置
        5. 清空并重新填充下拉框
        6. 尝试恢复之前选中项
        7. 返回操作结果

        示例：
        ------
        # 在设置页面的对象管理中：
        UIToolkit.refresh_obj_query_combo(self.parent())
        """

        try:
            # 参数验证
            if not main_window:
                print("[UIToolkit] ⚠️ refresh_obj_query_combo: main_window为空")
                return False

            # 检查是否有对象查询下拉框
            if not hasattr(main_window, 'obj_query_object_combo'):
                print("[UIToolkit] ⚠️ refresh_obj_query_combo: main_window没有obj_query_object_combo属性")
                return False

            combo = main_window.obj_query_object_combo
            if not combo:
                print("[UIToolkit] ⚠️ refresh_obj_query_combo: obj_query_object_combo为空")
                return False

            # 记录当前选中状态
            current_api_name = combo.currentData() or ""
            current_text = combo.currentText() or ""

            # 加载最新配置
            config = load_config()
            crm_objs = config.get('fxiaoke', {}).get('crm_objects', [])

            # 默认对象列表
            if not crm_objs:
                crm_objs = [
                    {'name': '销售订单', 'api_name': 'SalesOrderObj'},
                    {'name': '商机', 'api_name': 'NewOpportunityObj'},
                    {'name': '发货单', 'api_name': 'DeliveryNoteObj'},
                ]

            print(f"[UIToolkit] 🔄 刷新对象查询下拉框 | 对象数: {len(crm_objs)}")

            # 阻止信号触发
            combo.blockSignals(True)
            try:
                # 清空并重新填充
                combo.clear()
                combo.addItem("请选择对象", "")

                for obj in crm_objs:
                    obj_name = obj.get('name', '')
                    api_name = obj.get('api_name', '')
                    if obj_name and api_name:
                        combo.addItem(obj_name, api_name)

                # 尝试恢复选中状态
                restored = False
                if current_api_name:
                    for i in range(combo.count()):
                        if combo.itemData(i) == current_api_name:
                            combo.setCurrentIndex(i)
                            restored = True
                            break

                if not restored and current_text and current_text != "请选择对象":
                    for i in range(combo.count()):
                        if combo.itemText(i) == current_text:
                            combo.setCurrentIndex(i)
                            restored = True
                            break

                if not restored:
                    combo.setCurrentIndex(0)

            finally:
                combo.blockSignals(False)

            print(f"[UIToolkit] ✅ 对象查询下拉框已刷新 | 可用对象数: {combo.count()-1}")
            return True

        except Exception as e:
            print(f"[UIToolkit] ❌ refresh_obj_query_combo 失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    @staticmethod
    def save_crm_objects_immediately(objects_data, settings_dialog=None):
        """
        即时保存CRM对象数据并触发应用（通用方法）

        【功能说明】：
        将CRM对象列表保存到配置文件，并触发相关应用更新。
        整合了保存配置、应用设置、刷新UI等完整流程。

        【参数】：
        ------
        objects_data : list[dict]
            CRM对象数据列表，格式：
            [
                {"name": "销售订单", "api_name": "SalesOrderObj"},
                {"name": "商机", "api_name": "NewOpportunityObj"},
                ...
            ]
        settings_dialog : QWidget, optional
            设置对话框实例（用于获取parent和触发apply_settings_immediately）

        【返回值】：
        -------
        bool
            操作成功返回True，失败返回False

        【工作流程】：
        1. 验证objects_data有效性
        2. 保存到配置文件（immediate=True）
        3. 同步内存中的config引用
        4. 触发settings_dialog.apply_settings_immediately()
        5. 刷新主窗口的对象查询下拉框
        6. 记录日志并返回结果

        【示例】：
        ------
        objects = [
            {"name": "客户", "api_name": "AccountObj"},
            {"name": "联系人", "api_name": "ContactObj"}
        ]

        success = UIToolkit.save_crm_objects_immediately(
            objects_data=objects,
            settings_dialog=self  # self是SettingsDialog实例
        )
        if success:
            print("✅ CRM对象已保存并生效")
        """

        try:
            # 参数验证
            if not objects_data or len(objects_data) == 0:
                print("[UIToolkit] ⚠️ save_crm_objects_immediately: objects_data为空")
                return False

            # 1. 保存到配置文件
            config = load_config()
            if 'fxiaoke' not in config:
                config['fxiaoke'] = {}
            config['fxiaoke']['crm_objects'] = objects_data
            save_config(config, immediate=True)
            print(f"[UIToolkit] ✅ 已保存 {len(objects_data)} 个CRM对象到配置文件")

            # 2. 同步内存中的config引用
            if settings_dialog and hasattr(settings_dialog, 'config'):
                settings_dialog.config = config

            # 3. 触发设置应用（使用专用标签避免触发 CRM 数据缓存加载）
            if settings_dialog and hasattr(settings_dialog, 'apply_settings_immediately'):
                settings_dialog.apply_settings_immediately({'crm_objects_list'})

            # 4. 刷新对象查询下拉框
            if settings_dialog:
                parent = getattr(settings_dialog, 'parent', lambda: None)()
                if parent:
                    UIToolkit.refresh_obj_query_combo(parent)
                    print(f"[UIToolkit] ✅ 对象查询下拉框已同步更新")

            return True

        except Exception as e:
            print(f"[UIToolkit] ❌ save_crm_objects_immediately 失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    @staticmethod
    def rebuild_table_columns(table_widget, visible_labels, data_rows=None, display_to_api_map=None):
        """
        通用表格列重建方法

        【功能说明】：
        根据给定的可见字段标签列表重建QTableWidget的列结构，
        并可选地填充数据。适用于所有需要动态调整表格列的场景。

        【适用场景】：
        - 字段设置后的表格刷新
        - 对象切换时的列重组
        - 视图模式切换（详细/简洁）
        - 动态列显示/隐藏

        【参数】：
        ------
        table_widget : QTableWidget
            目标表格组件
        visible_labels : list[str]
            要显示的列标题列表
        data_rows : list[dict], optional
            要填充的数据行（每行为一个字典）
        display_to_api_map : dict[str, str], optional
            显示名→API字段的映射关系

        【返回值】：
        -------
        int
            实际填充的行数（如果提供了data_rows）

        【示例】：
        ------
        # 场景1：只重建列结构（不填充数据）
        UIToolkit.rebuild_table_columns(
            table_widget=self.my_table,
            visible_labels=["姓名", "电话", "地址"]
        )

        # 场景2：重建列并填充数据
        UIToolkit.rebuild_table_columns(
            table_widget=self.obj_query_table,
            visible_labels=visible_fields,
            data_rows=current_page_data,
            display_to_api_map={"客户名称": "accountName", "手机号": "mobilePhone"}
        )

        【注意事项】：
        ----------
        1. 如果display_to_api_map为空，会尝试直接用label作为key查找数据
        2. 数据填充时会自动处理空值和类型转换
        3. 此方法不会改变table_widget的其他属性（如选择模式、排序等）
        """

        if not table_widget or not visible_labels:
            print("[UIToolkit] ⚠️ rebuild_table_columns: 参数无效")
            return 0

        # 设置列数和表头
        old_col_count = table_widget.columnCount()
        new_col_count = len(visible_labels)
        table_widget.setColumnCount(new_col_count)
        table_widget.setHorizontalHeaderLabels(visible_labels)

        print(f"[UIToolkit] 🔄 重建表格列 | 列数: {old_col_count} → {new_col_count} | 标题: {visible_labels[:5]}...")

        # 如果没有提供数据，只重建列结构
        if not data_rows:
            table_widget.setRowCount(0)
            return 0

        # 填充数据
        table_widget.setRowCount(len(data_rows))
        filled_count = 0

        for row_idx, row_data in enumerate(data_rows):
            for col_idx, label in enumerate(visible_labels):
                # 尝试获取单元格值
                value = ""

                if display_to_api_map and label in display_to_api_map:
                    # 使用映射关系查找API字段
                    api_key = display_to_api_map[label]
                    value = row_data.get(api_key, "")
                elif isinstance(row_data, dict):
                    # 直接用label作为key查找
                    value = row_data.get(label, "")
                elif hasattr(row_data, '__getitem__'):
                    # 支持列表/元组等序列类型
                    try:
                        value = str(row_data[col_idx]) if col_idx < len(row_data) else ""
                    except (IndexError, TypeError):
                        value = ""

                # 设置单元格
                item = QTableWidgetItem(str(value) if value else "")
                table_widget.setItem(row_idx, col_idx, item)
                filled_count += 1

        print(f"[UIToolkit] ✅ 表格已填充 | 行数: {len(data_rows)} | 单元格数: {filled_count}")
        return len(data_rows)


def show_info(title, message, parent=None, **kwargs):
    """显示信息。"""
    if parent is None: parent = kwargs.get('parent')
    dlg = CustomMessageBox(title, message, "info", parent)
    dlg.exec()


def show_error(title, message, parent=None, **kwargs):
    """显示错误。"""
    if parent is None: parent = kwargs.get('parent')
    dlg = CustomMessageBox(title, message, "error", parent)
    dlg.exec()


def show_warning(title, message, parent=None, **kwargs):
    """显示警告。"""
    if parent is None: parent = kwargs.get('parent')
    dlg = CustomMessageBox(title, message, "warning", parent)
    dlg.exec()


def ask_yes_no(title, message, parent=None, **kwargs):
    """显示确认对话框并返回用户选择结果。"""
    if parent is None: parent = kwargs.get('parent')
    dlg = CustomMessageBox(title, message, "yesno", parent)
    result = dlg.exec()
    return dlg.result


# 设置QMessageBox的方法
class MessageBox:
    """消息框调用适配器。"""
    @staticmethod
    def showinfo(title, message, parent=None):
        """显示信息消息框。"""
        show_info(title, message, parent)
    
    @staticmethod
    def showerror(title, message, parent=None):
        """显示错误消息框。"""
        show_error(title, message, parent)
    
    @staticmethod
    def showwarning(title, message, parent=None):
        """显示警告消息框。"""
        show_warning(title, message, parent)
    
    @staticmethod
    def askyesno(title, message, parent=None):
        """显示确认消息框并返回选择结果。"""
        return ask_yes_no(title, message, parent)


# 创建messagebox实例
messagebox = MessageBox()





    # ✅【2026-05-11 删除】已移除年度密码和首次登录密码验证功能
    # 删除的函数列表：
    # - get_custom_field() (3427-3442)
    # - get_correct_password_for_year() (3445-3451)
    # - get_correct_password() (3454-3475)
    # - load_password_config() (3478-3490)
    # - password_verification() (3493-3724)
    #
    # 密码验证功能已完全禁用，所有调用处直接返回 True

def show_login_register_dialog():
    """显示登录或注册界面"""
    # 先检查用户缓存，有缓存则跳过登录界面（此时 __main__ 已完全加载，load_config 可用）
    try:
        cached_user, cached_user_type = load_user_cache()
        if cached_user:
            return cached_user, cached_user_type
    except Exception as _e:
        pass
    # 创建QApplication实例（如果还没有）
    app = ensure_qapplication()
    
    class LoginRegisterDialog(CenteredPopupDialog):
        """登录注册对话框。"""
        def __init__(self, parent=None, title="用户登录/注册"):
            """初始化登录注册对话框。"""
            super().__init__(parent, center_reference=parent)
            self.setWindowTitle(title)
            
            # Load size from config
            config = load_config()
            size_config = config.get('app_settings', {}).get('dialog_sizes', {}).get('LoginRegisterDialog', [400, 300])

            window_width = size_config[0]
            window_height = size_config[1]
            self.resize(window_width, window_height)
            self.setMinimumSize(300, 200)
            
            # 设置主题
            set_dialog_opacity(self)
            
            # 创建布局
            layout = QVBoxLayout()
            
            # 标题
            title_text = QLabel("用户登录/注册")
            title_text.setFont(QFont("Arial", 16, QFont.Weight.Bold))
            title_text.setAlignment(Qt.AlignCenter)
            layout.addWidget(title_text)
            layout.addSpacing(20)
            
            # 用户名
            username_layout = QHBoxLayout()
            username_label = QLabel("用户名:")
            username_label.setFixedWidth(80)
            username_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            username_layout.addWidget(username_label)
            
            self.username_text = QLineEdit()
            username_layout.addWidget(self.username_text)
            layout.addLayout(username_layout)
            layout.addSpacing(10)
            
            # 密码
            password_layout = QHBoxLayout()
            password_label = QLabel("密   码:")
            password_label.setFixedWidth(80)
            password_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            password_layout.addWidget(password_label)
            
            self.password_text = PasswordEntry()
            password_layout.addWidget(self.password_text)
            layout.addLayout(password_layout)
            layout.addSpacing(20)
            
            # 按钮
            btn_layout = QHBoxLayout()
            btn_layout.setAlignment(Qt.AlignCenter)
            
            login_btn = QPushButton("登录")
            login_btn.setFixedWidth(100)
            login_btn.clicked.connect(self.on_login)
            btn_layout.addWidget(login_btn)
            btn_layout.addSpacing(40)
            
            register_btn = QPushButton("注册")
            register_btn.setFixedWidth(100)
            register_btn.clicked.connect(self.on_register)
            btn_layout.addWidget(register_btn)
            
            layout.addLayout(btn_layout)
            self.setLayout(layout)
            
            # 绑定事件
            self.username_text.returnPressed.connect(self.on_login)
            self.password_text.entry.returnPressed.connect(self.on_login)
            
            # 存储结果
            self.result = None
            self.user_type = None
            self._logging_in = False
            
            # 焦点
            self.username_text.setFocus()

        def on_login(self):
            """登录按钮事件"""
            if self._logging_in:
                return
            self._logging_in = True

            try:
                self._do_login()
            finally:
                self._logging_in = False

        def _do_login(self):
            username = self.username_text.text()
            password = self.password_text.get()

            # 检查是否为管理员
            if is_admin(username, password):
                self.result = username
                self.user_type = "admin"
                self.accept()
                return

            # 检查是否为注册用户
            users_data = load_users()
            users = users_data.get('users', [])

            for user in users:
                if (user.get('username') == username or user.get('user_id') == username) and user.get(
                        'password') == password:
                    # 返回用户ID，这样可以确保匹配到正确的用户
                    self.result = user.get('user_id')
                    self.user_type = user.get('role', 'member')
                    self.accept()
                    return

            # 登录失败
            messagebox.showerror("错误", "用户名或密码错误")

        def on_register(self):
            """注册按钮事件"""
            username = self.username_text.text()
            password = self.password_text.get()

            if not username or not password:
                messagebox.showerror("错误", "用户名和密码不能为空")
                return

            # 检查用户名是否已存在
            users_data = load_users()
            users = users_data.get('users', [])

            for user in users:
                if user.get('username') == username:
                    messagebox.showerror("错误", "用户名已存在")
                    return

            # 生成用户ID
            def generate_user_id(users):
                """生成新用户ID"""
                # 提取现有用户ID的数字部分并排序
                user_ids = []
                for user in users:
                    if 'user_id' in user and user['user_id'].isdigit():
                        user_ids.append(int(user['user_id']))
                # 确保管理员ID 0001 不被使用
                if user_ids:
                    max_id = max(user_ids)
                    new_id = str(max_id + 1).zfill(3)
                else:
                    new_id = "002"
                return new_id

            # 添加新用户
            # 从配置文件中读取日期格式
            config = load_config()
            date_formats = config.get('business_rules', {}).get('date_formats', {
                'default': '%Y年%m月%d日',
                'short': '%Y-%m-%d',
                'datetime': '%Y-%m-%d %H:%M:%S'
            })
            
            new_user = {
                "username": username,
                "password": password,
                "user_id": generate_user_id(users),
                "created_at": get_current_time().strftime(date_formats.get('datetime', '%Y-%m-%d %H:%M:%S')),
                "role": "member",
                "permissions": {
                    "path": True,
                    "field_format": True,
                    "field_mapping": True,
                    "template": True,
                    "product_list": True,
                    "feature_selection": True,
                    "database_config": False
                }
            }
            users.append(new_user)
            users_data['users'] = users
            save_users(users_data)

            # 按用户首次验证，无需清空其他用户的密码验证状态

            messagebox.showinfo("成功", "注册成功")
            self.result = username
            self.user_type = "member"
            self.accept()

        def get_result(self):
            """获取登录/注册结果"""
            return self.result, self.user_type

    # 创建并显示对话框
    dialog = LoginRegisterDialog()
    dialog.exec()

    username, user_type = dialog.get_result()

    return username, user_type


# 检查是否需要显示登录/注册界面
try:
    users_data = load_users()
except NameError:
    users_data = {"users": []}
except Exception:
    users_data = {"users": []}
users = users_data.get('users', [])
has_users = len(users) > 0

# 检查用户缓存
try:
    cached_user, cached_user_type = load_user_cache()
except NameError:
    cached_user = None
    cached_user_type = None
except Exception:
    cached_user = None
    cached_user_type = None

current_user = None
user_type = None

if cached_user:
    current_user = cached_user
    user_type = cached_user_type


# 导入线程模块，用于后台运行
import threading


def move_files_to_new_folders(source_dir, target_root, file_extension):
    """移动文件到按文件名命名的新文件夹中"""
    # 遍历源目录下所有指定扩展名的文件
    for file_path in Path(source_dir).glob(f"*{file_extension}"):
        # 检查文件名是否包含 "-"
        if "-" not in file_path.stem:
            # 如果文件名不包含 "-", 跳过当前文件
            continue

        # 获取文件名（不含扩展名）
        base_name = file_path.stem
        # 在目标根目录下创建与文件名相同的新文件夹
        target_folder_path = os.path.join(target_root, base_name)
        os.makedirs(target_folder_path, exist_ok=True)  # 如果文件夹已存在，则不会抛出错误

        # 构建目标文件路径
        dst_file_path = os.path.join(target_folder_path, file_path.name)
        # 将文件从源路径移动到目标路径
        shutil.move(file_path, dst_file_path)
        # 打印移动信息
        # print(f'移动文件到 :{dst_file_path}')


def open_folder(path):
    """打开指定路径的文件夹，若已在资源管理器中打开则置顶"""
    system = platform.system()
    try:
        if system == "Windows":
            path = os.path.abspath(path).replace('/', '\\')
            import subprocess
            # 检查是否已有资源管理器窗口打开该路径，若有则置顶
            try:
                for proc in psutil.process_iter(['name', 'cmdline']):
                    if proc.info['name'] == 'explorer.exe' and len(proc.info['cmdline']) > 1:
                        if proc.info['cmdline'][1].lower() == path.lower():
                            hwnd = proc.info['cmdline'][0] if len(proc.info['cmdline']) > 0 else ''
                            print(f"文件夹 {path} 已打开，尝试置顶")
                            # 通过 PowerShell 将窗口置顶
                            try:
                                ps_script = f'''
                                Add-Type @"
                                using System; using System.Runtime.InteropServices;
                                public class WinAPI {{
                                    [DllImport("user32.dll")] public static extern bool SetForegroundWindow(IntPtr hWnd);
                                    [DllImport("user32.dll")] public static extern bool ShowWindow(IntPtr hWnd, int nCmdShow);
                                }}
"@
                                $shell = New-Object -ComObject Shell.Application
                                $windows = $shell.Windows()
                                foreach ($w in $windows) {{
                                    if ($w.Document.Folder.Self.Path -eq '{path}') {{
                                        $null = [WinAPI]::ShowWindow($w.HWND, 9)
                                        $null = [WinAPI]::SetForegroundWindow($w.HWND)
                                    }}
                                }}
                                '''
                                subprocess.run(['powershell', '-Command', ps_script],
                                               capture_output=True, timeout=5)
                            except Exception:
                                pass
                            return True
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

            os.startfile(path)
        elif system == "Darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
        return True
    except Exception as e:
        print(f"打开失败: {str(e)}")
        return False



def safe_eval(expression):
    """安全计算表达式"""
    from ast import literal_eval
    try:
        return literal_eval(expression)  # 安全地计算表达式
    except (ValueError, SyntaxError) as e:
        logging.error(f"表达式解析失败: {expression} - {str(e)}")
        return None


def process_excel_template(choice_param):
    """处理Excel模板"""
    # 延迟导入pandas，避免启动卡住
    import pandas as pd

    try:
        excel_template_path = AppConfig.get_excel_template_path(choice_param)
        # 检查Excel模板文件是否存在
        if not excel_template_path.exists():
            error_msg = f"Excel文件不存在，请检查路径: {excel_template_path}"
            logging.error(error_msg)
            raise FileNotFoundError(error_msg)

        # 从配置中获取工作表映射
        sheet_mapping = AppConfig._CONFIG.get('sheet_mapping', {})
        
        # 首先尝试从gui_options中获取对应功能的sheet名称
        selected_sheet = "合同明细"  # 默认使用合同明细
        gui_options = AppConfig._CONFIG.get('gui_options', [])
        for option in gui_options:
            if option.get('value') == choice_param:
                if option.get('sheet'):
                    selected_sheet = option.get('sheet')
                    break
        
        # 如果gui_options中没有找到对应的sheet名称，再从sheet_mapping中获取
        if selected_sheet == "合同明细":
            selected_sheet = sheet_mapping.get(str(choice_param), "合同明细")

        # 读取Excel文件中的指定工作表 - 使用dtype=object避免自动日期转换
        df = pd.read_excel(excel_template_path,
                           engine='openpyxl',
                           sheet_name=selected_sheet,
                           keep_default_na=False,
                           dtype=object)  # 以对象类型读取所有数据，避免日期自动转换

        # 计算公式处理 - 遍历所有列和值
        for col in df.columns:
            for idx, value in enumerate(df[col]):
                if isinstance(value, str) and '=' in value:
                    df.at[idx, col] = safe_eval(value)  # 计算表达式值

        # 确保日志Excel文件的父目录存在
        AppConfig.LOG_EXCEL.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(AppConfig.LOG_EXCEL, index=False)  # 保存处理后的数据到Excel
    except Exception as e:
        logging.error(f"Excel处理失败: {str(e)}")
        raise


def generate_contracts(mapping, selected_rows=None):
    """
    生成合同文档并返回生成的合同编号列表

    参数：
        mapping: 字段映射配置
        selected_rows: 选中的行数据列表（从UI复选框获取），如果为None则使用旧逻辑兼容
    """
    # 延迟导入openpyxl，避免启动卡住
    from openpyxl import load_workbook

    generated_contracts = []

    # 记录文件生成操作开始
    operation_log.info("--- 文件生成开始 ---")
    operation_log.info(f"待生成数据行数: {len(selected_rows) if selected_rows else '全部'}")

    try:
        # 如果有传入选中的行数据，直接使用
        if selected_rows is not None and len(selected_rows) > 0:
            data_rows = selected_rows
        else:
            # 兼容旧逻辑：读取全部数据并用validate_row验证
            workbook = load_workbook(AppConfig.LOG_EXCEL)
            sheet = workbook.active
            data_rows = [{header.value: cell.value for header, cell in zip(sheet[1], row)}
                         for row in sheet.iter_rows(min_row=2)]
            workbook.close()

        # 整批生成期间复用同一个 Word 会话，减少反复启动 COM 的耗时和不稳定性。
        with _get_main().pdf_conversion_batch():
            # 遍历每行数据
            for row_data in data_rows:
                # 如果是旧逻辑，需要验证；如果是新逻辑（selected_rows），跳过验证
                if selected_rows is None:
                    if not validate_row(row_data):  # 验证数据行是否有效
                        continue

                template = select_template(row_data["word模板"])  # 选择合适的模板
                context = build_context(row_data, mapping)  # 构建模板上下文

                document_info = generate_documents(template, context, row_data)  # 生成文档

                # 返回列表会用于后续统计和界面提示，因此优先保留最容易识别的业务名称。
                # 收集生成的文件名或合同编号
                if isinstance(document_info, dict) and document_info.get('base_name'):
                    generated_contracts.append(document_info['base_name'])
                # 优先使用Excel中的文件名字段
                elif '文件名' in row_data and row_data['文件名']:
                    generated_contracts.append(row_data['文件名'])
                elif '合同编号' in row_data and row_data['合同编号']:
                    generated_contracts.append(row_data['合同编号'])
                elif '协议编号' in row_data and row_data['协议编号']:
                    generated_contracts.append(row_data['协议编号'])
                elif '序号' in row_data and row_data['序号']:
                    generated_contracts.append(f"序号-{row_data['序号']}")
                else:
                    # 至少添加一个标识，确保generated_contracts不为空
                    generated_contracts.append(f"文件-{len(generated_contracts)+1}")

    except Exception as e:
        logging.error(f"合同生成失败: {str(e)}")
        operation_log.error(f"文件生成失败: {str(e)}")
        raise

    print(f"\n{'='*50}")
    print(f"合同生成统计：")
    print(f"  • 总计生成：{len(generated_contracts)} 个文档")
    print(f"  • Word文档：{len(generated_contracts)} 个")
    print(f"{'='*50}")

    # 记录文件生成操作完成
    operation_log.info(f"文件生成成功: 生成 {len(generated_contracts)} 个文档")
    if generated_contracts:
        operation_log.info(f"生成的文档: {', '.join(generated_contracts[:5])}{'...' if len(generated_contracts) > 5 else ''}")

    return generated_contracts


def validate_row(row_data):
    """验证数据行有效性"""
    required_fields = ['序号', 'word模板','检查']  # 必需字段
    check_status = row_data.get("检查", "") == "Y"  # 检查状态
    contract_generated = row_data.get("是否生成", "") == "Y"  # 是否已生成

    # # 检查日期字段的有效性
    # date_fields = ['日期', '开始日期', '结束日期']
    # for field in date_fields:
    #     date_val = row_data.get(field)
    #     if date_val and date_val != "":
    #         # 尝试格式化日期以验证其有效性
    #         formatted = format_date(date_val, field)
    #         if formatted == str(date_val) and str(date_val).isdigit() and int(date_val) > 2958465:
    #             logging.warning(f"发现无效日期值: {date_val} 在字段 {field}")
    #             return False  # 无效日期则跳过该行

    # 返回验证结果：必需字段存在、检查状态为Y且未生成过
    return (
            all(row_data.get(field) for field in required_fields)
            and check_status
            and not contract_generated
    )


def select_template(software_version):
    """根据软件版本选择合适的Word模板"""
    config = AppConfig()  # 创建配置实例
    template_path = config.WORD_TEMPLATES.get(software_version)  # 获取模板路径

    # 检查模板是否存在
    if not template_path:
        available = list(config.WORD_TEMPLATES.keys())
        raise ValueError(f"无效模板类型: {software_version}，可用模板: {available}")

    if not template_path.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_path}\n"
                                f"请检查配置文件中的 template_dir 和 word_templates 配置")

    return load_docx_template(template_path)  # 返回模板对象


def build_context(row_data, mapping):
    """构建模板上下文"""
    return build_context_from_mapping_values(row_data, mapping, AppConfig.SPECIAL_COLUMN_FORMATS)


def build_context_from_mapping_values(source_values, mapping, special_formats=None):
    """根据字段映射和格式配置构建模板上下文"""
    context = {}
    special_formats = special_formats or {}

    # 构建大小写不敏感的备用索引
    source_lower = {str(k).strip().lower(): v for k, v in source_values.items()}

    for key, mapped_key in mapping.items():
        value = None
        key_lower = str(key).strip().lower()
        if key in source_values:
            value = source_values[key]
        elif key_lower in source_lower:
            value = source_lower[key_lower]
        if value is None:
            continue

        if key in special_formats:
            fmt = special_formats[key]
            try:
                if fmt.startswith('%'):
                    context[mapped_key] = format_date(value, key, fmt=fmt)
                else:
                    if isinstance(value, (int, float)):
                        context[mapped_key] = fmt.format(value)
                    else:
                        context[mapped_key] = fmt.format(str(value))
            except Exception:
                context[mapped_key] = value if value not in ("-", None) else ""
        elif key == '日期' and value:
            context[mapped_key] = format_date(value, key)
        elif key in ('开始日期', '结束日期') and value:
            context[mapped_key] = format_date(value, key)
        else:
            if isinstance(value, float) and value.is_integer():
                context[mapped_key] = str(int(value))
            else:
                context[mapped_key] = str(value) if value not in ("-", None) else ""
    return context


def format_date(value, field_name=None, fmt=None):
    """格式化日期（根据字段名称动态调整格式）"""
    # 延迟导入pandas，避免启动卡住
    import pandas as pd

    # 从配置文件中读取配置
    config = load_config()
    
    # 首先尝试从字段格式设置中获取日期格式
    field_format = ""
    special_columns_config = config.get('business_rules', {}).get('special_columns', [])
    
    # 检查特殊字段配置，获取对应字段的格式
    if field_name:
        if isinstance(special_columns_config, list):
            # 新格式：列表，每个元素是包含name和format的字典
            for column in special_columns_config:
                if column.get('name') == field_name and column.get('format'):
                    field_format = column.get('format')
                    break
        elif isinstance(special_columns_config, dict):
            # 旧格式：字典，键为字段名，值为格式
            field_format = special_columns_config.get(field_name, "")
    
    # 如果提供了自定义格式，使用自定义格式
    if fmt:
        date_format = fmt
    # 如果字段格式设置中存在对应字段的格式，使用该格式
    elif field_format:
        date_format = field_format
    # 如果字段名是 "开始日期" 或 "结束日期"，使用默认格式
    elif field_name in ("开始日期", "结束日期"):
        # 从配置文件中读取默认日期格式
        date_formats = config.get('business_rules', {}).get('date_formats', {
            'default': '%Y年%m月%d日',
            'short': '%Y-%m-%d',
            'datetime': '%Y-%m-%d %H:%M:%S'
        })
        date_format = date_formats.get('default', '%Y年%m月%d日')
    else:
        # 从配置文件中读取默认日期格式
        date_formats = config.get('business_rules', {}).get('date_formats', {
            'default': '%Y年%m月%d日',
            'short': '%Y-%m-%d',
            'datetime': '%Y-%m-%d %H:%M:%S'
        })
        date_format = date_formats.get('default', '%Y年%m月%d日')

    # 检查是否为空值
    if pd.isna(value) or value in ("", "-", None):
        return ""

    # 如果已经是datetime对象
    if isinstance(value, datetime):
        try:
            return value.strftime(date_format)
        except UnicodeEncodeError:
            # 如果中文格式化失败，尝试使用字段格式设置中的格式，或使用默认短日期格式
            return value.strftime('%Y-%m-%d')

    # 尝试解析字符串日期
    try:
        # 检查是否是数值类型（可能是Excel序列号）
        if isinstance(value, (int, float)):
            if value > 1e12:
                converted_date = datetime.fromtimestamp(value / 1000)
                try:
                    return converted_date.strftime(date_format)
                except UnicodeEncodeError:
                    return converted_date.strftime('%Y-%m-%d')
            elif value > 1e9:
                converted_date = datetime.fromtimestamp(value)
                try:
                    return converted_date.strftime(date_format)
                except UnicodeEncodeError:
                    return converted_date.strftime('%Y-%m-%d')
            # 检查是否在合理范围内（Excel日期范围）
            elif 1 <= value <= 2958465:  # Excel日期范围
                # 转换Excel序列号为日期
                excel_start = datetime(1900, 1, 1)
                converted_date = excel_start + timedelta(days=value - 2)  # Excel有闰年bug，所以减2
                try:
                    return converted_date.strftime(date_format)
                except UnicodeEncodeError:
                    # 如果中文格式化失败，尝试使用字段格式设置中的格式，或使用默认短日期格式
                    return converted_date.strftime('%Y-%m-%d')
            else:
                logging.warning(f"Excel日期序列号超出范围: {value}")
                return ""
        else:
            # 处理字符串格式日期
            str_value = str(value).strip()
            if str_value in ("", "-", "None"):
                return ""

            if re.fullmatch(r'\d{13}', str_value):
                converted_date = datetime.fromtimestamp(int(str_value) / 1000)
                try:
                    return converted_date.strftime(date_format)
                except UnicodeEncodeError:
                    return converted_date.strftime('%Y-%m-%d')

            if re.fullmatch(r'\d{10}', str_value):
                converted_date = datetime.fromtimestamp(int(str_value))
                try:
                    return converted_date.strftime(date_format)
                except UnicodeEncodeError:
                    return converted_date.strftime('%Y-%m-%d')

            # 尝试多种日期格式
            possible_formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%Y-%m-%d',
                '%Y/%m/%d',
                '%Y.%m.%d',
                '%Y年%m月%d日',
                '%Y/%m/%d %H:%M:%S',
                '%m/%d/%Y',
                '%d/%m/%Y'
            ]

            for fmt in possible_formats:
                try:
                    dt = datetime.strptime(str_value, fmt)
                    try:
                        return dt.strftime(date_format)
                    except UnicodeEncodeError:
                        # 如果中文格式化失败，尝试使用字段格式设置中的格式，或使用默认短日期格式
                        return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue

            # 如果所有格式都失败，返回原值
            logging.warning(f"无法解析日期格式: {value}")
            return str(value)
    except UnicodeEncodeError:
        # 如果编码出错，返回原值
        return str(value) if value is not None else ""
    except Exception as e:
        logging.error(f"日期格式化错误: {value}, 错误: {str(e)}")
        return str(value) if value is not None else ""


def sanitize_output_name(base_name):
    """清理文件名中的非法字符"""
    text = str(base_name or '').strip()
    if not text:
        return '未命名'
    text = re.sub(r'[<>:"/\\|?*]+', '_', text)
    text = text.rstrip(' .')
    return text or '未命名'


def safe_move_to_dir(source_path, target_dir):
    """将文件安全移动到目标目录，覆盖同名文件。"""
    if not source_path.exists():
        return
    target_dir = Path(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_file = target_dir / source_path.name
    if target_file.exists():
        try:
            target_file.unlink()
        except Exception:
            pass
    try:
        shutil.move(str(source_path), str(target_file))
    except Exception:
        try:
            shutil.copy2(str(source_path), str(target_file))
        except Exception:
            pass


def render_document_files(template, context, base_name, output_dir=None, skip_pdf=False):
    """渲染模板并输出 Word/PDF 文件

    参数：
        skip_pdf: 由调用方根据业务上下文决定是否跳过 PDF 转换。
                  文件生成流：core._keep_word_override == 'word'
                  CRM合同流：core._keep_crm_word_override == 'word'
    """
    try:
        template.render(context)
        clean_template_tables(template)

        safe_base_name = sanitize_output_name(base_name)
        target_output_dir = Path(output_dir) if output_dir else AppConfig.OUTPUT_DIR
        target_output_dir.mkdir(parents=True, exist_ok=True)

        word_path = target_output_dir / f"{safe_base_name}{AppConfig.FILE_EXTENSION}"
        pdf_path = target_output_dir / f"{safe_base_name}.pdf"

        max_attempts = 3
        attempt = 0
        # 先尝试复用原文件名；若旧文件无法删除，则退化为追加时间戳避免覆盖失败。
        while attempt < max_attempts:
            if not word_path.exists():
                break

            try:
                word_path.unlink()
                print(f"已删除已存在的Word文件: {word_path}")
                break
            except Exception as e:
                print(f"删除已存在的Word文件失败: {str(e)}")
                import time
                timestamp = int(time.time())
                safe_base_name = sanitize_output_name(f"{safe_base_name}_{timestamp}")
                word_path = target_output_dir / f"{safe_base_name}{AppConfig.FILE_EXTENSION}"
                pdf_path = target_output_dir / f"{safe_base_name}.pdf"
                print(f"使用唯一文件名: {word_path}")
                attempt += 1

        template.save(word_path)
        print(f"Word文档已保存: {word_path}")

        pdf_success = False
        if skip_pdf:
            print(f"输出格式为Word，跳过PDF转换")
        else:
            try:
                print(f"开始转换为PDF: {word_path} -> {pdf_path}")
                convert_word_to_pdf(word_path, pdf_path)
                pdf_success = True
                print(f"✓ PDF文档已成功生成: {pdf_path}")
            except Exception as pdf_error:
                print(f"✗ PDF转换失败: {str(pdf_error)}")
                logging.error(f"PDF转换失败: {str(pdf_error)}")
                if "comtypes" in str(pdf_error).lower() or "word" in str(pdf_error).lower():
                    print("提示：请确保已安装Microsoft Word或LibreOffice，并安装了comtypes库（pip install comtypes）")

        return {
            'base_name': word_path.stem,
            'word_path': word_path,
            'pdf_path': pdf_path,
            'pdf_success': pdf_success,
        }
    except Exception as e:
        logging.error(f"文档生成失败: {str(e)}")
        raise


def generate_documents(template, context, row_data):
    """生成Word和PDF文档"""
    try:
        # 根据用户选择动态生成文件名
        # 优先使用Excel中的文件名字段
        if '文件名' in row_data and row_data['文件名']:
            base_name = f"{row_data['文件名']}"

        elif choice in (1, 4, 5):
            # 使用合同编号-客户
            base_name = f"{row_data['合同编号']}-{row_data['客户']}"
        else:  # choice=2或3时
            # 使用协议编号-客户
            if '协议编号' in row_data and row_data['协议编号']:
                base_name = f"{row_data['协议编号']}-{row_data['客户']}"
            else:
                # 如果没有协议编号，使用合同编号-客户
                base_name = f"{row_data.get('合同编号', '无编号')}-{row_data['客户']}"
        return render_document_files(template, context, base_name, skip_pdf=(core._keep_word_override == 'word'))
    except Exception as e:
        logging.error(f"文档生成失败: {str(e)}")
        raise


def clean_template_tables(template):
    """清理模板表格 - 删除空行"""
    if len(template.docx.tables) < 2:
        return
    table = template.docx.tables[1]
    # 反向遍历并删除空行
    for row in reversed(list(table.rows[1:])):
        if (row.cells[0].text or "").strip() in ("", "-"):
            table._element.remove(row._element)


def convert_word_to_pdf(word_path, pdf_path):
    """转换Word为PDF"""
    try:
        active_session = getattr(core._WORD_PDF_CONVERTER_LOCAL, 'session', None)
        if active_session is not None:
            active_session.convert(word_path, pdf_path)
            return

        with _get_main().pdf_conversion_batch() as session:
            session.convert(word_path, pdf_path)
    except Exception as e:
        logging.error(f"PDF转换失败: {str(e)}")
        raise


def organize_output_files():
    """整理输出文件 - 将生成的文件移动到目标目录（文件生成专用）"""
    config = AppConfig()

    # ✅ 文件生成使用 keep_output_format 配置

    if core._keep_word_override is not None:
        keep_fmt = core._keep_word_override
    else:
        app_config = load_config()
        keep_fmt = app_config.get('app_settings', {}).get('keep_output_format', 'word_pdf')

    keep_word = keep_fmt in ('word', 'word_pdf')
    keep_pdf = keep_fmt in ('pdf', 'word_pdf')

    print(f"开始整理输出文件 | 格式: {keep_fmt} | 保留Word: {keep_word} | 保留PDF: {keep_pdf}")
    print(f"输出目录: {AppConfig.OUTPUT_DIR}")
    print(f"文件扩展名: {AppConfig.FILE_EXTENSION}")
    print(f"选择的功能: {choice}")
    print(f"目标根目录: {config.get_target_root(choice)}")

    # 根据格式处理PDF文件
    if keep_pdf:
        print("开始处理PDF文件（保留）")
    else:
        print("不保留PDF文件，删除PDF文件")
    pdf_files = list(AppConfig.OUTPUT_DIR.glob('*.pdf'))
    print(f"找到的PDF文件数量: {len(pdf_files)}")
    
    for file in pdf_files:
        # 跳过临时文件
        if file.name.startswith('~$'):
            print(f"跳过临时PDF文件: {file.name}")
            continue
        
        # 检查文件是否为空或损坏
        if file.stat().st_size == 0:
            print(f"跳过空PDF文件: {file.name}")
            try:
                file.unlink()
                print(f"已删除空PDF文件: {file.name}")
            except Exception as e:
                print(f"删除空PDF文件失败: {str(e)}")
            continue
        
        if keep_pdf:
            print(f"处理PDF文件: {file.name}")
            base_stem = file.stem.split('_')[0] if '_' in file.stem else file.stem
            target_dir = config.get_target_root(choice) / base_stem
            target_dir.mkdir(parents=True, exist_ok=True)
            try:
                target_file = target_dir / file.name
                if target_file.exists():
                    try:
                        target_file.unlink()
                    except Exception:
                        pass
                shutil.move(str(file), str(target_file))
                print(f"PDF文件已移动: {file.name} -> {target_file}")
            except Exception as e:
                print(f"移动PDF文件失败: {str(e)}")
                try:
                    target_file = target_dir / file.name
                    shutil.copy2(str(file), str(target_file))
                    try:
                        file.unlink()
                    except Exception:
                        pass
                except Exception:
                    pass
        else:
            try:
                file.unlink()
                print(f"已删除PDF文件: {file.name}")
            except Exception as e:
                print(f"删除PDF文件失败: {str(e)}")
    
    # 根据格式处理Word文件
    if keep_word:
        print("开始处理Word文件（保留）")
        word_files = list(AppConfig.OUTPUT_DIR.glob(f'*{AppConfig.FILE_EXTENSION}'))
        print(f"找到的Word文件数量: {len(word_files)}")
        for file in word_files:
            # 跳过临时文件
            if file.name.startswith('~$'):
                print(f"跳过临时Word文件: {file.name}")
                continue
            
            # 检查文件是否为空
            if file.stat().st_size == 0:
                print(f"跳过空Word文件: {file.name}")
                try:
                    file.unlink()
                    print(f"已删除空Word文件: {file.name}")
                except Exception as e:
                    print(f"删除空Word文件失败: {str(e)}")
                continue
            
            print(f"处理Word文件: {file.name}")
            # 获取目标目录（移除时间戳）
            base_stem = file.stem.split('_')[0] if '_' in file.stem else file.stem
            target_dir = config.get_target_root(choice) / base_stem
            print(f"目标目录: {target_dir}")
            # 创建目标目录（如果不存在）
            target_dir.mkdir(parents=True, exist_ok=True)
            print(f"目标目录已创建")
            # 移动Word文件
            try:
                # 检查目标文件是否存在，如果存在先删除
                target_file = target_dir / file.name
                if target_file.exists():
                    try:
                        target_file.unlink()
                        print(f"已删除已存在的目标文件: {target_file}")
                    except Exception as e:
                        print(f"删除已存在的目标文件失败: {str(e)}")
                # 移动文件
                shutil.move(str(file), str(target_file))
                print(f"Word文件已移动: {file.name} -> {target_file}")
            except Exception as e:
                print(f"移动Word文件失败: {str(e)}")
                # 尝试使用copy2代替move
                try:
                    target_file = target_dir / file.name
                    shutil.copy2(str(file), str(target_file))
                    print(f"Word文件已复制: {file.name} -> {target_file}")
                    # 尝试删除源文件
                    try:
                        file.unlink()
                        print(f"源Word文件已删除: {file.name}")
                    except Exception as e:
                        print(f"删除源Word文件失败: {str(e)}")
                except Exception as e2:
                    print(f"复制Word文件也失败: {str(e2)}")
    else:
        print("不保留Word文件，删除Word文件")
        word_files = list(AppConfig.OUTPUT_DIR.glob(f'*{AppConfig.FILE_EXTENSION}'))
        for file in word_files:
            if file.name.startswith('~$'):
                continue
            try:
                file.unlink()
                print(f"已删除Word文件: {file.name}")
            except Exception as e:
                print(f"删除Word文件失败: {str(e)}")


def move_generated_document_files(generated_files, keep_word=None):
    """将指定的生成文件移动到各自目标目录（CRM订单专用）"""
    # ✅ CRM订单使用 keep_crm_output_format 配置

    if keep_word is None:
        if core._keep_crm_word_override is not None:
            keep_fmt = core._keep_crm_word_override
        else:
            app_config = load_config()
            keep_fmt = app_config.get('app_settings', {}).get('keep_crm_output_format', 'word_pdf')
        keep_word = keep_fmt in ('word', 'word_pdf')
        keep_pdf = keep_fmt in ('pdf', 'word_pdf')
        print(f"[DEBUG-CRM文件整理] 格式: {keep_fmt} | 保留Word: {keep_word} | 保留PDF: {keep_pdf}")
    else:
        # 显式传入 keep_word 时，也需要从配置/全局变量确定 keep_pdf
        if core._keep_crm_word_override is not None:
            keep_fmt = core._keep_crm_word_override
        else:
            app_config = load_config()
            keep_fmt = app_config.get('app_settings', {}).get('keep_crm_output_format', 'word_pdf')
        keep_pdf = keep_fmt in ('pdf', 'word_pdf')
        print(f"[DEBUG-CRM文件整理] ✓ 使用参数传入 | keep_word={keep_word} | keep_pdf={keep_pdf}")

    moved_roots = []

    for file_info in generated_files or []:
        if not isinstance(file_info, dict):
            continue

        base_name = str(file_info.get('base_name', '') or '').strip()
        target_root = file_info.get('target_root')
        word_path = file_info.get('word_path')
        pdf_path = file_info.get('pdf_path')

        if not base_name or not target_root:
            continue

        target_root = Path(target_root)
        target_dir = target_root / base_name
        target_dir.mkdir(parents=True, exist_ok=True)

        if target_root not in moved_roots:
            moved_roots.append(target_root)

        def _move_file(source_path):
            """内部方法：移动文件。"""
            if not source_path:
                return
            source_path = Path(source_path)
            if not source_path.exists() or source_path.name.startswith('~$'):
                return
            if source_path.stat().st_size == 0:
                try:
                    source_path.unlink()
                except Exception:
                    pass
                return

            target_file = target_dir / source_path.name
            if target_file.exists():
                try:
                    target_file.unlink()
                except Exception:
                    pass

            try:
                shutil.move(str(source_path), str(target_file))
            except Exception:
                try:
                    shutil.copy2(str(source_path), str(target_file))
                    try:
                        source_path.unlink()
                    except Exception:
                        pass
                except Exception:
                    pass

        if keep_pdf:
            _move_file(pdf_path)
        elif pdf_path and Path(pdf_path).exists():
            try:
                Path(pdf_path).unlink()
            except Exception:
                pass
        if keep_word:
            _move_file(word_path)
        elif word_path and Path(word_path).exists():
            try:
                Path(word_path).unlink()
                print(f"已删除Word文件: {word_path}")
            except Exception:
                pass

    return moved_roots



def open_output_folder():
    """打开输出目录（受通用设置中'弹出文件夹'选项控制）"""
    try:
        app_config = load_config()
        open_enabled = app_config.get('app_settings', {}).get('open_output_folder', True)
        if not open_enabled:
            print("弹出文件夹功能已关闭，跳过")
            return
        config = AppConfig()
        open_folder(str(config.get_target_root(choice)))
    except Exception as e:
        logging.error(f"打开文件夹失败: {str(e)}")


def commit_to_svn(custom_dir=None):
    """将生成的文件提交到 TortoiseSVN"""
    try:
        if custom_dir:
            target_dir = custom_dir
        else:
            config = AppConfig()  # 创建配置实例
            target_dir = config.get_target_root(choice)

        # 检查目标目录是否存在
        if not target_dir.exists():
            logging.warning(f"目标目录不存在: {target_dir}")
            return

        # 确保路径格式正确，特别是对于包含空格的路径
        # 使用绝对路径
        target_dir_str = str(target_dir.absolute())

        # 使用 TortoiseSVN 命令行工具提交文件
        # 注意：TortoiseProc.exe 需要在系统PATH中，或者使用完整路径
        # 将/path和路径作为两个单独的参数传递
        subprocess.run(["TortoiseProc.exe", "/command:commit", "/path", target_dir_str, "/closeonend:0"])

    except Exception as e:
        logging.error(f"SVN 提交失败: {str(e)}")
        import traceback
        logging.error(f"错误详情: {traceback.format_exc()}")


def main():
    """主函数"""
    # 主函数现在由 MainFrame 中的按钮触发，不再直接显示选择对话框
    pass

# 延迟导入PyQt6模块
from PyQt6.QtWidgets import QMainWindow, QFrame, QLabel, QPushButton, QRadioButton, QComboBox, QSlider, QLineEdit, QTextEdit, QTreeView, QTableView, QTableWidget, QTableWidgetItem, QTabWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout, QScrollArea, QGroupBox, QMessageBox, QFileDialog, QInputDialog, QHeaderView, QMenu, QCheckBox, QStackedWidget, QToolButton, QWidget, QSpinBox, QCalendarWidget, QStyleOptionButton, QStyle, QAbstractSpinBox, QTreeWidget, QTreeWidgetItem, QStyledItemDelegate, QSystemTrayIcon, QDateEdit, QCompleter, QSizePolicy
from PyQt6.QtCore import Qt, QSize, QTimer, QEvent, pyqtSignal, QRect, QDate, QPoint
from PyQt6.QtGui import QFont, QColor, QPalette, QIntValidator, QAction, QIcon, QPainter, QTextCharFormat

class CheckBoxHeader(QHeaderView):
    """首列表头复选框"""

    toggled = pyqtSignal(bool)

    def __init__(self, parent=None):
        """初始化复选框表头。"""
        super().__init__(Qt.Orientation.Horizontal, parent)
        self._check_state = Qt.CheckState.Unchecked
        self._checkbox_press_pending = False
        self.setSectionsClickable(True)

    def set_check_state(self, state):
        """设置检查状态。"""
        self._check_state = state
        self.viewport().update()

    def _event_point(self, event):
        """内部方法：处理事件point逻辑。"""
        if hasattr(event, "position"):
            return event.position().toPoint()
        return event.pos()

    def _checkbox_rect(self):
        """内部方法：处理复选框rect逻辑。"""
        section_width = self.sectionSize(0)
        section_x = self.sectionViewportPosition(0)
        option = QStyleOptionButton()
        size = self.style().sizeFromContents(QStyle.ContentsType.CT_CheckBox, option, QSize(0, 0), self)   # 表头复选框
        return QRect(
            section_x + (section_width - size.width()) // 3,
            (self.height() - size.height()) // 3,
            size.width(),
            size.height()
        )

    def _checkbox_hit_rect(self):
        """内部方法：处理复选框hitrect逻辑。"""
        return self._checkbox_rect().adjusted(-6, -6, 6, 6)

    def paintSection(self, painter, rect, logical_index):
        """绘制表头分区内容。"""
        super().paintSection(painter, rect, logical_index)
        if logical_index != 0:
            return

        option = QStyleOptionButton()
        option.rect = self._checkbox_rect()
        option.state = QStyle.StateFlag.State_Enabled
        if self._check_state == Qt.CheckState.Checked:
            option.state |= QStyle.StateFlag.State_On
        elif self._check_state == Qt.CheckState.PartiallyChecked:
            option.state |= QStyle.StateFlag.State_NoChange
        else:
            option.state |= QStyle.StateFlag.State_Off

        painter.save()
        self.style().drawControl(QStyle.ControlElement.CE_CheckBox, option, painter)
        painter.restore()

    def mousePressEvent(self, event):
        """处理鼠标按下事件。"""
        event_pos = self._event_point(event)
        if self.logicalIndexAt(event_pos) == 0 and self._checkbox_hit_rect().contains(event_pos):
            self._checkbox_press_pending = True
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        """处理鼠标释放事件。"""
        event_pos = self._event_point(event)
        if self._checkbox_press_pending:
            self._checkbox_press_pending = False
            if self.logicalIndexAt(event_pos) == 0 and self._checkbox_hit_rect().contains(event_pos):
                old_state = self._check_state
                self._check_state = (
                    Qt.CheckState.Unchecked
                    if self._check_state == Qt.CheckState.Checked
                    else Qt.CheckState.Checked
                )
                emit_value = self._check_state == Qt.CheckState.Checked
                print(f"[CheckBoxHeader] 状态切换: {old_state} -> {self._check_state}, 发射: {emit_value}")
                self.toggled.emit(emit_value)
                self.viewport().update()
            event.accept()
            return
        super().mouseReleaseEvent(event)


class TableRowCheckBox(QCheckBox):
    """用于表格首列的小尺寸复选框"""

    toggled_with_row_id = pyqtSignal(object, bool)

    def __init__(self, row_id=None, parent=None):
        """初始化表格行复选框。"""
        super().__init__(parent)
        self.row_id = row_id
        self.setTristate(False)
        self.setStyleSheet("""
            QCheckBox {
                spacing: 0px;
                padding: 0px;
                margin: 0px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
        """)
        self.toggled.connect(self._emit_toggled)

    def _emit_toggled(self, checked):
        """内部方法：处理emittoggled逻辑。"""
        self.toggled_with_row_id.emit(self.row_id, checked)


class BrowseLineEdit(QLineEdit):
    """支持双击触发浏览的输入框"""

    doubleClicked = pyqtSignal()

    def mouseDoubleClickEvent(self, event):
        """处理鼠标双击事件。"""
        self.doubleClicked.emit()
        super().mouseDoubleClickEvent(event)


class CheckableOptionPopup(QFrame):
    """支持连续勾选的多选弹层"""

    selection_changed = pyqtSignal()

    def __init__(self, parent=None):
        """初始化多选弹层。"""
        super().__init__(parent, Qt.WindowType.Popup)
        self._updating = False
        self.option_checkboxes = []

        self.setObjectName("checkableOptionPopup")
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setLineWidth(1)
        self.setMinimumWidth(280)
        self.setStyleSheet("""
            QFrame#checkableOptionPopup {
                background-color: rgba(255, 255, 255, 232);
                border: 1px solid rgba(193, 208, 224, 220);
                border-radius: 10px;
            }
            QCheckBox {
                padding: 4px 6px;
                border-radius: 6px;
            }
            QCheckBox:hover {
                background-color: rgba(91, 141, 239, 0.10);
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        toolbar = QHBoxLayout()
        toolbar.setContentsMargins(0, 0, 0, 0)
        toolbar.setSpacing(8)

        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.setTristate(True)
        self.select_all_checkbox.stateChanged.connect(self._on_select_all_state_changed)
        toolbar.addWidget(self.select_all_checkbox)

        self.clear_btn = QPushButton("清空")
        self.clear_btn.setFixedWidth(56)
        self.clear_btn.clicked.connect(self.clear_selection)
        toolbar.addWidget(self.clear_btn)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        layout.addWidget(self.scroll_area)

        self.options_container = QWidget()
        self.options_layout = QVBoxLayout(self.options_container)
        self.options_layout.setContentsMargins(0, 0, 0, 0)
        self.options_layout.setSpacing(4)
        self.scroll_area.setWidget(self.options_container)

    def _clear_option_widgets(self):
        """内部方法：清空选项widgets。"""
        while self.options_layout.count():
            item = self.options_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

    def set_options(self, options, selected_options=None, select_all_when_empty=False):
        """设置选项。"""
        normalized_options = []
        seen = set()
        for option in options or []:
            text = str(option).strip()
            if text and text not in seen:
                seen.add(text)
                normalized_options.append(text)

        selected_set = {
            str(option).strip()
            for option in (selected_options or [])
            if str(option).strip()
        }
        if select_all_when_empty and normalized_options and not selected_set:
            selected_set = set(normalized_options)

        self._updating = True
        self.option_checkboxes = []
        self._clear_option_widgets()

        if normalized_options:
            for option in normalized_options:
                checkbox = QCheckBox(option)
                checkbox.setChecked(option in selected_set)
                checkbox.toggled.connect(self._on_option_toggled)
                self.options_layout.addWidget(checkbox)
                self.option_checkboxes.append(checkbox)
        else:
            empty_label = QLabel("暂无可选业务类型")
            empty_label.setStyleSheet("color: #888; padding: 4px 2px;")
            self.options_layout.addWidget(empty_label)

        self.options_layout.addStretch()
        self._updating = False
        self._update_select_all_state()

    def get_available_options(self):
        """获取可用选项。"""
        return [checkbox.text().strip() for checkbox in self.option_checkboxes if checkbox.text().strip()]

    def get_selected_options(self):
        """获取选中选项。"""
        return [
            checkbox.text().strip()
            for checkbox in self.option_checkboxes
            if checkbox.isChecked() and checkbox.text().strip()
        ]

    def _set_all_checked(self, checked):
        """内部方法：设置allchecked。"""
        if not self.option_checkboxes:
            self._update_select_all_state()
            return

        self._updating = True
        for checkbox in self.option_checkboxes:
            checkbox.setChecked(bool(checked))
        self._updating = False
        self._update_select_all_state()
        self.selection_changed.emit()

    def clear_selection(self):
        """清空选择。"""
        self._set_all_checked(False)

    def _update_select_all_state(self):
        """内部方法：更新选择all状态。"""
        checked_count = sum(1 for checkbox in self.option_checkboxes if checkbox.isChecked())
        total_count = len(self.option_checkboxes)

        self._updating = True
        if total_count and checked_count == total_count:
            self.select_all_checkbox.setCheckState(Qt.CheckState.Checked)
        elif checked_count == 0:
            self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
        else:
            self.select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
        self.select_all_checkbox.setEnabled(bool(total_count))
        self.clear_btn.setEnabled(bool(checked_count))
        self._updating = False

    def _on_select_all_state_changed(self, state):
        """内部方法：响应选择all状态changed相关操作。"""
        if self._updating:
            return

        state_value = state.value if hasattr(state, 'value') else state
        if state_value == Qt.CheckState.Checked.value:
            self._set_all_checked(True)
        elif state_value == Qt.CheckState.Unchecked.value:
            self._set_all_checked(False)

    def _on_option_toggled(self, checked):
        """内部方法：响应选项toggled相关操作。"""
        if self._updating:
            return

        self._update_select_all_state()
        self.selection_changed.emit()

    def show_below(self, anchor_widget):
        """显示below。"""
        if anchor_widget is None:
            return

        option_count = max(1, len(self.option_checkboxes))
        popup_width = max(anchor_widget.width() + 140, 280)
        popup_height = min(320, 56 + min(option_count, 8) * 30)
        self.resize(popup_width, popup_height)
        self.move(anchor_widget.mapToGlobal(anchor_widget.rect().bottomLeft()))
        self.show()
        self.raise_()


class CheckableOptionPopup(QFrame):
    """支持连续勾选的多选弹层"""

    selection_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent, Qt.WindowType.Popup)
        self._updating = False
        self.option_checkboxes = []

        self.setObjectName("checkableOptionPopup")
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setLineWidth(1)
        self.setMinimumWidth(280)
        self.setStyleSheet("""
            QFrame#checkableOptionPopup {
                background-color: rgba(255, 255, 255, 232);
                border: 1px solid rgba(193, 208, 224, 220);
                border-radius: 10px;
            }
            QCheckBox {
                padding: 4px 6px;
                border-radius: 6px;
            }
            QCheckBox:hover {
                background-color: rgba(91, 141, 239, 0.10);
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        toolbar = QHBoxLayout()
        toolbar.setContentsMargins(0, 0, 0, 0)
        toolbar.setSpacing(8)

        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.setTristate(True)
        self.select_all_checkbox.stateChanged.connect(self._on_select_all_state_changed)
        toolbar.addWidget(self.select_all_checkbox)

        self.clear_btn = QPushButton("清空")
        self.clear_btn.setFixedWidth(56)
        self.clear_btn.clicked.connect(self.clear_selection)
        toolbar.addWidget(self.clear_btn)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        layout.addWidget(self.scroll_area)

        self.options_container = QWidget()
        self.options_layout = QVBoxLayout(self.options_container)
        self.options_layout.setContentsMargins(0, 0, 0, 0)
        self.options_layout.setSpacing(4)
        self.scroll_area.setWidget(self.options_container)

    def _clear_option_widgets(self):
        while self.options_layout.count():
            item = self.options_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

    def set_options(self, options, selected_options=None, select_all_when_empty=False):
        normalized_options = []
        seen = set()
        for option in options or []:
            text = str(option).strip()
            if text and text not in seen:
                seen.add(text)
                normalized_options.append(text)

        selected_set = {
            str(option).strip()
            for option in (selected_options or [])
            if str(option).strip()
        }
        if select_all_when_empty and normalized_options and not selected_set:
            selected_set = set(normalized_options)

        self._updating = True
        self.option_checkboxes = []
        self._clear_option_widgets()

        if normalized_options:
            for option in normalized_options:
                checkbox = QCheckBox(option)
                checkbox.setChecked(option in selected_set)
                checkbox.toggled.connect(self._on_option_toggled)
                self.options_layout.addWidget(checkbox)
                self.option_checkboxes.append(checkbox)
        else:
            empty_label = QLabel("暂无可选业务类型")
            empty_label.setStyleSheet("color: #888; padding: 4px 2px;")
            self.options_layout.addWidget(empty_label)

        self.options_layout.addStretch()
        self._updating = False
        self._update_select_all_state()

    def get_available_options(self):
        return [checkbox.text().strip() for checkbox in self.option_checkboxes if checkbox.text().strip()]

    def get_selected_options(self):
        return [
            checkbox.text().strip()
            for checkbox in self.option_checkboxes
            if checkbox.isChecked() and checkbox.text().strip()
        ]

    def _set_all_checked(self, checked):
        if not self.option_checkboxes:
            self._update_select_all_state()
            return

        self._updating = True
        for checkbox in self.option_checkboxes:
            checkbox.setChecked(bool(checked))
        self._updating = False
        self._update_select_all_state()
        self.selection_changed.emit()

    def clear_selection(self):
        self._set_all_checked(False)

    def _update_select_all_state(self):
        checked_count = sum(1 for checkbox in self.option_checkboxes if checkbox.isChecked())
        total_count = len(self.option_checkboxes)

        self._updating = True
        if total_count and checked_count == total_count:
            self.select_all_checkbox.setCheckState(Qt.CheckState.Checked)
        elif checked_count == 0:
            self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
        else:
            self.select_all_checkbox.setCheckState(Qt.CheckState.PartiallyChecked)
        self.select_all_checkbox.setEnabled(bool(total_count))
        self.clear_btn.setEnabled(bool(checked_count))
        self._updating = False

    def _on_select_all_state_changed(self, state):
        if self._updating:
            return

        state_value = state.value if hasattr(state, 'value') else state
        if state_value == Qt.CheckState.Checked.value:
            self._set_all_checked(True)
        elif state_value == Qt.CheckState.Unchecked.value:
            self._set_all_checked(False)

    def _on_option_toggled(self, checked):
        if self._updating:
            return

        self._update_select_all_state()
        self.selection_changed.emit()

    def show_below(self, anchor_widget):
        if anchor_widget is None:
            return

        option_count = max(1, len(self.option_checkboxes))
        popup_width = max(anchor_widget.width() + 140, 280)
        popup_height = min(320, 56 + min(option_count, 8) * 30)
        self.resize(popup_width, popup_height)
        self.move(anchor_widget.mapToGlobal(anchor_widget.rect().bottomLeft()))
        self.show()
        self.raise_()
class CRMInlineComboDelegate(QStyledItemDelegate):
    """CRM表格的单元格内联下拉编辑器"""

    def __init__(self, main_window, column_type, parent=None):
        """初始化CRM行内下拉委托。"""
        super().__init__(parent)
        self.main_window = main_window
        self.column_type = column_type

    def createEditor(self, parent, option, index):
        """创建单元格编辑器。"""
        editor = QComboBox(parent)
        editor.setFrame(False)

        if self.column_type == 'customer_type':
            editor.addItem('经销商', '经销商')
            editor.addItem('终端客户', '终端客户')
            editor.activated.connect(lambda *_: self._commit_and_close(editor))
        else:
            editor.setEditable(True)
            editor.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
            editor_options = []
            if self.main_window and hasattr(self.main_window, 'get_crm_inline_editor_options'):
                editor_options = self.main_window.get_crm_inline_editor_options(index.row(), self.column_type)
            for option_text in editor_options:
                editor.addItem(option_text, option_text)
            line_edit = editor.lineEdit()
            if line_edit:
                line_edit.setPlaceholderText('留空恢复默认')
                line_edit.returnPressed.connect(lambda: self._commit_and_close(editor))
            editor.activated.connect(lambda *_: self._commit_and_close(editor))

        QTimer.singleShot(0, editor.setFocus)
        if editor.count() > 0:
            QTimer.singleShot(0, editor.showPopup)
        return editor

    def setEditorData(self, editor, index):
        """向编辑器写入当前值。"""
        current_text = str(index.data(Qt.ItemDataRole.DisplayRole) or '').strip()
        combo_index = editor.findText(current_text)
        if combo_index >= 0:
            editor.setCurrentIndex(combo_index)
        elif self.column_type in ('matched_price', 'product_remark'):
            editor.setEditText(current_text)
        elif editor.count() > 0:
            editor.setCurrentIndex(0)

    def setModelData(self, editor, model, index):
        """将编辑器内容回写到模型。"""
        value = str(editor.currentText() or '').strip()
        model.setData(index, value, Qt.ItemDataRole.EditRole)
        if self.main_window:
            row = index.row()
            col = index.column()
            QTimer.singleShot(
                0,
                lambda row=row, col=col, value=value: self.main_window.apply_crm_inline_cell_override(row, col, value)
            )

    def updateEditorGeometry(self, editor, option, index):
        """更新编辑器位置和大小。"""
        editor.setGeometry(option.rect)

    def _commit_and_close(self, editor):
        """提交编辑结果并关闭编辑器。"""
        if editor.property('_crm_inline_committed'):
            return
        editor.setProperty('_crm_inline_committed', True)
        self.commitData.emit(editor)
        self.closeEditor.emit(editor, QStyledItemDelegate.EndEditHint.NoHint)


class TableCellEditMenu(QObject):
    """为 QTableWidget 单元格添加右键编辑菜单"""

    def __init__(self, table, parent=None):
        super().__init__(parent)
        self._table = table
        table.installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.ContextMenu:
            # ✅ 转换坐标：event.pos() 是 table 坐标，itemAt 需要 viewport 坐标
            viewport = self._table.viewport()
            pos = viewport.mapFrom(obj, event.pos())
            item = self._table.itemAt(pos)
            if item is not None:
                if not (item.flags() & Qt.ItemFlag.ItemIsEditable):
                    return False
                row, col = item.row(), item.column()
                # ✅ 先选中当前单元格，避免编辑跨行
                self._table.setCurrentCell(row, col)
                self._table.clearSelection()
                self._table.selectRow(row)
                menu = QMenu(self._table)
                edit_action = menu.addAction("✏️ 编辑")
                action = menu.exec(event.globalPos())
                if action == edit_action:
                    self._table.editItem(item)
                    return True
        return False


def _copy_selected_cells(table):
    """复制 QTableWidget 中选中的单元格内容到剪贴板，Tab 分隔列，换行分隔行。"""
    from PyQt6.QtWidgets import QApplication, QComboBox
    ranges = table.selectedRanges()
    if not ranges:
        return
    selected = set()
    min_row, max_row = float('inf'), -1
    min_col, max_col = float('inf'), -1
    for rng in ranges:
        for row in range(rng.topRow(), rng.bottomRow() + 1):
            for col in range(rng.leftColumn(), rng.rightColumn() + 1):
                selected.add((row, col))
                min_row = min(min_row, row)
                max_row = max(max_row, row)
                min_col = min(min_col, col)
                max_col = max(max_col, col)

    lines = []
    for row in range(min_row, max_row + 1):
        parts = []
        for col in range(min_col, max_col + 1):
            if (row, col) in selected:
                item = table.item(row, col)
                if item:
                    parts.append(item.text())
                else:
                    w = table.cellWidget(row, col)
                    if w:
                        cb = w.findChild(QComboBox)
                        if cb:
                            parts.append(cb.currentText())
                        else:
                            parts.append('')
                    else:
                        parts.append('')
            else:
                parts.append('')
        lines.append('\t'.join(parts))
    QApplication.clipboard().setText('\n'.join(lines))


def install_table_copy_handler(table):
    """为 QTableWidget 安装 Ctrl+C 多选复制功能（防止重复安装）"""
    if hasattr(table, '_copy_handler_installed'):
        return
    table._copy_handler_installed = True
    orig_key_press = table.keyPressEvent

    def _handler(event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier and event.key() == Qt.Key.Key_C:
            _copy_selected_cells(table)
            return
        orig_key_press(event)

    table.keyPressEvent = _handler


def install_table_edit_context_menu(table):
    """为指定 QTableWidget 安装右键编辑菜单和 Ctrl+C 多选复制（防止重复安装）"""
    if not hasattr(table, '_edit_menu_installed'):
        table._edit_menu_installed = True
        TableCellEditMenu(table, table)
    install_table_copy_handler(table)


def install_header_alignment_menu(table, config_key, repopulate_callback=None):
    """为 QTableWidget 的水平表头安装右键对齐菜单。"""
    if hasattr(table, '_header_align_menu_installed'):
        return
    table._header_align_menu_installed = True
    header = table.horizontalHeader()

    # 保存原始的 contextMenuEvent
    _orig_context_menu_event = header.contextMenuEvent

    def _new_context_menu_event(event):
        from PyQt6.QtWidgets import QMenu
        pos = event.pos()
        logical_index = header.logicalIndexAt(pos)
        if logical_index < 0:
            _orig_context_menu_event(event)
            return
        header_item = table.horizontalHeaderItem(logical_index)
        if not header_item:
            _orig_context_menu_event(event)
            return
        menu = QMenu()
        left_action = menu.addAction("左对齐")
        center_action = menu.addAction("居中")
        right_action = menu.addAction("右对齐")
        action = menu.exec(event.globalPos())
        if not action:
            return
        align = 'left' if action == left_action else ('center' if action == center_action else 'right')
        header_label = header_item.text().strip()
        if not header_label:
            return
        config = load_config()
        app_set = config.setdefault('app_settings', {})
        table_set = app_set.setdefault(config_key, {})
        alignments = table_set.setdefault('column_alignments', {})
        alignments[header_label] = align
        save_config(config, immediate=True)
        if repopulate_callback:
            repopulate_callback()

    header.contextMenuEvent = _new_context_menu_event


class QuickDatePickerDialog(CenteredPopupDialog):
    """带快捷导航的日期范围选择弹窗"""

    def __init__(self, start_date=None, end_date=None, parent=None, relative_key=None):
        """初始化快捷日期选择对话框。"""
        super().__init__(parent, center_reference=parent)
        self.setObjectName("quickDatePickerDialog")
        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setModal(False)
        self.setWindowTitle("选择日期")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.resize(672, 318)
        self.setMinimumSize(632, 300)
        self._highlighted_dates = []
        self.relative_key = relative_key  # 相对日期锚点：today/this_month等，None=固定日期
        self.setStyleSheet("""
            QDialog#quickDatePickerDialog {
                background-color: #ffffff;
                border: 1px solid #d9e3ef;
                border-radius: 10px;
            }
            QFrame#quickDateShortcutPanel {
                background: transparent;
                border: none;
            }
            QFrame#quickDateCalendarPanel,
            QFrame#quickDateSingleCalendar {
                background-color: #ffffff;
                border: 1px solid #e3ebf5;
                border-radius: 8px;
            }
            QLabel#quickDateSelectionLabel {
                color: #16324f;
                padding: 0;
                font-size: 12px;
                font-weight: 700;
            }
            QLabel#quickDateHintLabel {
                color: #6e7f92;
                padding: 0 0 2px 0;
                font-size: 10px;
            }
            QLabel#quickDateCalendarTitle {
                color: #16324f;
                font-weight: 600;
                padding-left: 0;
                font-size: 11px;
            }
            QPushButton#quickDateShortcutButton {
                text-align: center;
                min-width: 66px;
                max-width: 66px;
                min-height: 27px;
                max-height: 27px;
                padding: 2px 6px;
                border-radius: 5px;
                border: 1px solid #dbe5f0;
                background-color: #fbfdff;
                color: #23405e;
                font-size: 11px;
            }
            QPushButton#quickDateShortcutButton:hover {
                background-color: #eff5fd;
                border-color: #a9bfdb;
            }
            QPushButton#quickDateGhostButton {
                min-width: 50px;
                min-height: 24px;
                padding: 3px 8px;
                border-radius: 5px;
                border: 1px solid #d9e3ef;
                background-color: #ffffff;
                color: #17324d;
                font-size: 11px;
            }
            QPushButton#quickDateGhostButton:hover {
                background-color: #eff5fd;
                border-color: #a9bfdb;
            }
            QPushButton#quickDatePrimaryButton {
                min-width: 56px;
                min-height: 24px;
                padding: 3px 10px;
                border-radius: 5px;
                border: 1px solid #0078d7;
                background-color: #0078d7;
                color: #ffffff;
                font-weight: 600;
                font-size: 11px;
            }
            QPushButton#quickDatePrimaryButton:hover {
                background-color: #0063b1;
                border-color: #0063b1;
            }
            QCalendarWidget {
                background-color: #ffffff;
                border: none;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #f3f7fc;
                border: 1px solid #e1e9f4;
                border-radius: 5px;
                margin-bottom: 6px;
                min-height: 24px;
                max-height: 24px;
            }
            QCalendarWidget QToolButton {
                color: #17324d;
                background: transparent;
                border: none;
                padding: 2px 5px;
                border-radius: 4px;
                font-weight: 600;
                font-size: 10px;
            }
            QCalendarWidget QToolButton:hover {
                background-color: rgba(91, 141, 239, 0.12);
            }
            QCalendarWidget QToolButton#qt_calendar_prevmonth,
            QCalendarWidget QToolButton#qt_calendar_nextmonth {
                min-width: 18px;
                max-width: 18px;
                padding: 1px 2px;
            }
            QCalendarWidget QToolButton#qt_calendar_monthbutton,
            QCalendarWidget QToolButton#qt_calendar_yearbutton {
                min-height: 16px;
                max-height: 16px;
                padding: 0 2px;
                border-radius: 3px;
                font-size: 9px;
            }
            QCalendarWidget QToolButton#qt_calendar_monthbutton {
                min-width: 30px;
                max-width: 34px;
            }
            QCalendarWidget QToolButton#qt_calendar_yearbutton {
                min-width: 34px;
                max-width: 38px;
            }
            QCalendarWidget QToolButton#qt_calendar_monthbutton::menu-indicator,
            QCalendarWidget QToolButton#qt_calendar_yearbutton::menu-indicator {
                subcontrol-position: right center;
                width: 6px;
                height: 6px;
            }
            QCalendarWidget QMenu,
            QCalendarWidget QSpinBox {
                background-color: #ffffff;
                border: 1px solid #d9e3ef;
                border-radius: 5px;
                color: #17324d;
                font-size: 10px;
                min-width: 46px;
                padding: 0 3px;
            }
            QCalendarWidget QAbstractItemView:enabled {
                color: #17324d;
                background-color: #ffffff;
                alternate-background-color: #ffffff;
                selection-background-color: transparent;
                selection-color: #17324d;
                outline: none;
            }
            QCalendarWidget QTableView {
                border: none;
                outline: none;
                background-color: #ffffff;
                alternate-background-color: #ffffff;
                gridline-color: #edf2f7;
            }
            QCalendarWidget QTableView::item {
                background-color: #ffffff;
                border: none;
                padding: 1px 0;
            }
            QCalendarWidget QHeaderView::section {
                background: transparent;
                color: #55657a;
                border: none;
                padding: 2px 0 4px 0;
                font-size: 10px;
                font-weight: 600;
            }
        """)

        self.start_date = start_date if start_date and start_date.isValid() else None
        self.end_date = end_date if end_date and end_date.isValid() else None
        self._selecting_start = self.start_date is None
        focus_date = self.end_date or self.start_date or QDate.currentDate()

        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)

        shortcut_panel = QFrame()
        shortcut_panel.setObjectName("quickDateShortcutPanel")
        shortcut_panel.setFixedWidth(82)
        shortcut_layout = QVBoxLayout(shortcut_panel)
        shortcut_layout.setContentsMargins(0, 2, 0, 2)
        shortcut_layout.setSpacing(4)
        for text, handler in [
            ("上一年度", self.goto_previous_year),
            ("本年度", self.goto_current_year),
            ("本季度", self.goto_current_quarter),
            ("上一季度", self.goto_previous_quarter),
            ("本月", self.goto_current_month),
            ("上月", self.goto_previous_month),
            ("本周", self.goto_current_week),
            ("上周", self.goto_previous_week),
            ("今天", self.goto_today),
            ("昨天", self.goto_yesterday),
        ]:
            button = QPushButton(text)
            button.setObjectName("quickDateShortcutButton")
            button.setCursor(Qt.CursorShape.PointingHandCursor)
            button.clicked.connect(handler)
            shortcut_layout.addWidget(button)
        shortcut_layout.addStretch()

        calendar_panel = QFrame()
        calendar_panel.setObjectName("quickDateCalendarPanel")
        right_panel = QVBoxLayout(calendar_panel)
        right_panel.setContentsMargins(8, 8, 8, 8)
        right_panel.setSpacing(4)

        self.selected_label = QLabel()
        self.selected_label.setObjectName("quickDateSelectionLabel")
        self.selection_hint_label = QLabel()
        self.selection_hint_label.setObjectName("quickDateHintLabel")
        right_panel.addWidget(self.selected_label)
        right_panel.addWidget(self.selection_hint_label)

        calendar_layout = QHBoxLayout()
        calendar_layout.setSpacing(6)
        self.left_calendar = self._create_calendar_widget()
        self.right_calendar = self._create_calendar_widget()
        calendar_layout.addWidget(self._create_calendar_panel("开始日期", self.left_calendar), 1)
        calendar_layout.addWidget(self._create_calendar_panel("结束日期", self.right_calendar), 1)
        right_panel.addLayout(calendar_layout, 1)

        action_layout = QHBoxLayout()
        action_layout.setSpacing(6)
        # 间隔选择
        self.shift_interval_combo = QComboBox()
        self.shift_interval_combo.setFixedWidth(72)
        self.shift_interval_combo.addItem("月", 1)
        self.shift_interval_combo.addItem("季", 3)
        self.shift_interval_combo.addItem("年", 12)
        self.shift_interval_combo.setCurrentIndex(0)
        self.shift_interval_combo.setStyleSheet("QComboBox { font-size: 11px; border: 1px solid #dbe5f0; border-radius: 4px; padding: 1px 3px; background: #fbfdff; color: #23405e; }")
        action_layout.addWidget(self.shift_interval_combo)
        # 平移按钮
        for text, handler in [("←", self.shift_range_prev), ("→", self.shift_range_next)]:
            btn = QPushButton(text)
            btn.setObjectName("quickDateGhostButton")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFixedWidth(28)
            btn.clicked.connect(handler)
            action_layout.addWidget(btn)
        action_layout.addStretch()
        clear_btn = QPushButton("清空")
        confirm_btn = QPushButton("确定")
        clear_btn.setObjectName("quickDateGhostButton")
        confirm_btn.setObjectName("quickDatePrimaryButton")
        clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        confirm_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_btn.clicked.connect(self.clear_selection)
        confirm_btn.clicked.connect(self.accept)
        action_layout.addWidget(clear_btn)
        action_layout.addWidget(confirm_btn)
        right_panel.addLayout(action_layout)

        main_layout.addWidget(shortcut_panel)
        main_layout.addWidget(calendar_panel, 1)

        self.focus_date(focus_date)
        self.update_selected_label()

    def _create_calendar_widget(self):
        """内部方法：创建日历widget。"""
        calendar = QCalendarWidget()
        calendar.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        calendar.setGridVisible(False)
        calendar.setFirstDayOfWeek(Qt.DayOfWeek.Monday)
        self._apply_calendar_theme(calendar)
        calendar.clicked.connect(self.on_calendar_date_selected)
        calendar.activated.connect(self.on_calendar_date_activated)
        return calendar

    def _create_calendar_panel(self, title, calendar):
        """内部方法：创建日历面板。"""
        panel = QFrame()
        panel.setObjectName("quickDateSingleCalendar")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(4)
        title_label = QLabel(title)
        title_label.setObjectName("quickDateCalendarTitle")
        layout.addWidget(title_label)
        layout.addWidget(calendar, 1)
        return panel

    def _apply_calendar_theme(self, calendar):
        """内部方法：应用日历theme。"""
        weekday_format = QTextCharFormat()
        weekday_format.setForeground(QColor("#27405f"))
        weekday_format.setFontWeight(int(QFont.Weight.Medium))

        weekend_format = QTextCharFormat()
        weekend_format.setForeground(QColor("#e24c4b"))
        weekend_format.setFontWeight(int(QFont.Weight.Medium))

        for weekday in (
            Qt.DayOfWeek.Monday,
            Qt.DayOfWeek.Tuesday,
            Qt.DayOfWeek.Wednesday,
            Qt.DayOfWeek.Thursday,
            Qt.DayOfWeek.Friday,
        ):
            calendar.setWeekdayTextFormat(weekday, weekday_format)

        for weekday in (Qt.DayOfWeek.Saturday, Qt.DayOfWeek.Sunday):
            calendar.setWeekdayTextFormat(weekday, weekend_format)

    def _clear_calendar_range_formats(self):
        """内部方法：清空日历范围formats。"""
        if not self._highlighted_dates:
            return
        default_format = QTextCharFormat()
        for date in self._highlighted_dates:
            self.left_calendar.setDateTextFormat(date, default_format)
            self.right_calendar.setDateTextFormat(date, default_format)
        self._highlighted_dates.clear()

    def _refresh_calendar_range_formats(self):
        """内部方法：刷新日历范围formats。"""
        self._clear_calendar_range_formats()
        if not self.start_date or not self.start_date.isValid():
            return

        range_start = self.start_date
        range_end = self.end_date if self.end_date and self.end_date.isValid() else self.start_date

        boundary_format = QTextCharFormat()
        boundary_format.setBackground(QColor("#5b8def"))
        boundary_format.setForeground(QColor("#ffffff"))
        boundary_format.setFontWeight(int(QFont.Weight.DemiBold))

        between_format = QTextCharFormat()
        between_format.setBackground(QColor("#eaf2ff"))
        between_format.setForeground(QColor("#17324d"))

        current_date = QDate(range_start)
        while current_date <= range_end:
            date_format = boundary_format if current_date == range_start or current_date == range_end else between_format
            self.left_calendar.setDateTextFormat(current_date, date_format)
            self.right_calendar.setDateTextFormat(current_date, date_format)
            self._highlighted_dates.append(QDate(current_date))
            current_date = current_date.addDays(1)

    def event(self, event):
        """处理Qt事件。"""
        if event.type() == QEvent.Type.WindowDeactivate and self.isVisible():
            self.reject()
            return True
        return super().event(event)

    def position_below_widget(self, anchor_widget, margin=6):
        """将日期弹窗定位到触发控件下方。"""
        self.set_popup_anchor_widget(anchor_widget, margin)

    def focus_date(self, date):
        """聚焦日期。"""
        if not date or not date.isValid():
            date = QDate.currentDate()
        if self.start_date and self.end_date:
            left_month = QDate(self.start_date.year(), self.start_date.month(), 1)
            right_month = QDate(self.end_date.year(), self.end_date.month(), 1)
            if left_month > right_month:
                left_month, right_month = right_month, left_month
            if left_month == right_month:
                right_month = right_month.addMonths(1)
            left_selected = self.start_date
            right_selected = self.end_date
        else:
            left_month = QDate(date.year(), date.month(), 1)
            right_month = left_month.addMonths(1)
            left_selected = self.start_date or date
            right_selected = self.end_date or left_selected

        self.left_calendar.setCurrentPage(left_month.year(), left_month.month())
        self.right_calendar.setCurrentPage(right_month.year(), right_month.month())
        self.left_calendar.setSelectedDate(left_selected)
        self.right_calendar.setSelectedDate(right_selected)

    def update_selected_label(self):
        """更新选中标签。"""
        if self.start_date and self.end_date:
            days = self.start_date.daysTo(self.end_date) + 1
            self.selected_label.setText(
                f"已选：{self.start_date.toString('yyyy-MM-dd')} ~ {self.end_date.toString('yyyy-MM-dd')}"
            )
            self.selection_hint_label.setText(f"共 {days} 天，点击日期可重选")
        elif self.start_date:
            self.selected_label.setText(f"开始日期：{self.start_date.toString('yyyy-MM-dd')}")
            self.selection_hint_label.setText("再点一个结束日期")
        else:
            self.selected_label.setText("请选择日期范围")
            self.selection_hint_label.setText("支持快捷范围和双击完成")
        self._refresh_calendar_range_formats()

    def on_calendar_date_selected(self, date):
        """响应日历日期选中相关操作。"""
        self.relative_key = None  # 用户手动选择日历日期 → 固定日期
        if self._selecting_start or not self.start_date:
            self.start_date = date
            self.end_date = None
            self._selecting_start = False
        else:
            if date < self.start_date:
                self.end_date = self.start_date
                self.start_date = date
            else:
                self.end_date = date
            self._selecting_start = True
        self.focus_date(date)
        self.update_selected_label()

    def on_calendar_date_activated(self, date):
        """响应日历日期activated相关操作。"""
        self.on_calendar_date_selected(date)
        if self.start_date and self.end_date:
            self.accept()

    def clear_selection(self):
        """清空选择。"""
        self.start_date = None
        self.end_date = None
        self._selecting_start = True
        self.focus_date(QDate.currentDate())
        self.update_selected_label()

    def goto_today(self):
        """跳转到今天。"""
        self.relative_key = "today"
        today = QDate.currentDate()
        self.start_date = today
        self.end_date = today
        self._selecting_start = True
        self.focus_date(today)
        self.update_selected_label()

    def goto_yesterday(self):
        """跳转到yesterday。"""
        self.relative_key = "yesterday"
        yesterday = QDate.currentDate().addDays(-1)
        self.start_date = yesterday
        self.end_date = yesterday
        self._selecting_start = True
        self.focus_date(yesterday)
        self.update_selected_label()

    def goto_current_week(self):
        """跳转到当前周。"""
        self.relative_key = "this_week"
        today = QDate.currentDate()
        week_start = today.addDays(1 - today.dayOfWeek())
        week_end = week_start.addDays(6)
        self.start_date = week_start
        self.end_date = week_end
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_previous_week(self):
        """跳转到上一页周。"""
        self.relative_key = "last_week"
        today = QDate.currentDate()
        current_week_start = today.addDays(1 - today.dayOfWeek())
        self.start_date = current_week_start.addDays(-7)
        self.end_date = self.start_date.addDays(6)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_next_week(self):
        """跳转到下一页周。"""
        start_date = QDate.currentDate().addDays(1)
        self.start_date = start_date
        self.end_date = start_date.addDays(6)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_current_month(self):
        """跳转到当前月份。"""
        self.relative_key = "this_month"
        current = QDate.currentDate()
        self.start_date = QDate(current.year(), current.month(), 1)
        self.end_date = self.start_date.addMonths(1).addDays(-1)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_previous_month(self):
        """跳转到上一页月份。"""
        self.relative_key = "last_month"
        current = QDate.currentDate().addMonths(-1)
        self.start_date = QDate(current.year(), current.month(), 1)
        self.end_date = self.start_date.addMonths(1).addDays(-1)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    @staticmethod
    def _shift_month_preserve_eom(d, months):
        """平移月份，保留月末语义：若原日期是当月最后一天，平移后也是目标月最后一天"""
        is_eom = d.day() == d.daysInMonth()
        new_d = d.addMonths(months)
        if is_eom:
            new_d = QDate(new_d.year(), new_d.month(), new_d.daysInMonth())
        return new_d

    def _get_shift_months(self):
        """读取间隔下拉框的月数"""
        return self.shift_interval_combo.currentData() or 1

    def shift_range_prev(self):
        """将已选日期范围向前平移指定间隔（保留月末语义）。平移后变为固定日期。"""
        self.relative_key = None
        if not self.start_date or not self.start_date.isValid():
            return
        if not self.end_date or not self.end_date.isValid():
            return
        months = self._get_shift_months()
        self.start_date = self._shift_month_preserve_eom(self.start_date, -months)
        self.end_date = self._shift_month_preserve_eom(self.end_date, -months)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def shift_range_next(self):
        """将已选日期范围向后平移指定间隔（保留月末语义）。平移后变为固定日期。"""
        self.relative_key = None
        if not self.start_date or not self.start_date.isValid():
            return
        if not self.end_date or not self.end_date.isValid():
            return
        months = self._get_shift_months()
        self.start_date = self._shift_month_preserve_eom(self.start_date, months)
        self.end_date = self._shift_month_preserve_eom(self.end_date, months)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_next_month(self):
        """跳转到下一页月份。"""
        current = QDate.currentDate().addMonths(1)
        self.start_date = QDate(current.year(), current.month(), 1)
        self.end_date = self.start_date.addMonths(1).addDays(-1)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_current_year(self):
        """跳转到当前年份。"""
        self.relative_key = "this_year"
        current = QDate.currentDate()
        self.start_date = QDate(current.year(), 1, 1)
        self.end_date = QDate(current.year(), 12, 31)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_previous_year(self):
        """跳转到上一页年份。"""
        self.relative_key = "last_year"
        year = QDate.currentDate().year() - 1
        self.start_date = QDate(year, 1, 1)
        self.end_date = QDate(year, 12, 31)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_next_year(self):
        """跳转到下一页年份。"""
        year = QDate.currentDate().year() + 1
        self.start_date = QDate(year, 1, 1)
        self.end_date = QDate(year, 12, 31)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_current_quarter(self):
        """跳转到当前quarter。"""
        self.relative_key = "this_quarter"
        current = QDate.currentDate()
        quarter_month = ((current.month() - 1) // 3) * 3 + 1
        self.start_date = QDate(current.year(), quarter_month, 1)
        self.end_date = self.start_date.addMonths(3).addDays(-1)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_previous_quarter(self):
        """跳转到上一页quarter。"""
        self.relative_key = "last_quarter"
        current = QDate.currentDate()
        quarter_month = ((current.month() - 1) // 3) * 3 + 1
        self.start_date = QDate(current.year(), quarter_month, 1).addMonths(-3)
        self.end_date = self.start_date.addMonths(3).addDays(-1)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()

    def goto_next_quarter(self):
        """跳转到下一页quarter。"""
        current = QDate.currentDate()
        quarter_month = ((current.month() - 1) // 3) * 3 + 1
        self.start_date = QDate(current.year(), quarter_month, 1).addMonths(3)
        self.end_date = self.start_date.addMonths(3).addDays(-1)
        self._selecting_start = True
        self.focus_date(self.start_date)
        self.update_selected_label()


class DatePartPickerDialog(CenteredPopupDialog):
    """提取日期对话框 - 批量生成 =YEAR()/=MONTH()/=WEEKNUM()/=QUARTER() 公式。"""

    TIME_UNITS = [
        ('year', '年'),
        ('month', '月'),
        ('week', '周'),
        ('quarter', '季度'),
    ]

    UNIT_LABELS = {
        'year': '年',
        'month': '月',
        'week': '周',
        'quarter': '季度',
    }

    DEFAULT_NAMES = {
        'year': '年',
        'month': '月',
        'week': '周',
        'quarter': '季度',
    }

    def __init__(self, available_fields=None, parent=None):
        super().__init__(parent, center_reference=parent, close_on_outside=True)
        self.setWindowTitle("提取日期")
        self.resize(460, 340)
        self.setMinimumSize(420, 300)

        self.available_fields = available_fields or []
        self.source_field_name = ''
        self.selected_units = []

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 16, 20, 16)
        main_layout.setSpacing(14)

        # 标题
        title = QLabel("提取日期")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #333;")
        main_layout.addWidget(title)

        # ---- 提取源字段 ----
        src_label = QLabel("提取源字段:")
        src_label.setStyleSheet("font-size: 13px; color: #333; font-weight: 500;")
        main_layout.addWidget(src_label)

        self.source_combo = QComboBox()
        self.source_combo.setEditable(True)
        self.source_combo.setFixedHeight(30)
        self.source_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #D9D9D9; border-radius: 4px;
                padding: 4px 10px; font-size: 13px; background: #FFF;
            }
            QComboBox:hover { border-color: #FF8C00; }
            QComboBox QAbstractItemView { font-size: 13px; }
        """)
        for display_name, field_name in self.available_fields:
            self.source_combo.addItem(f"{display_name} ({field_name})", field_name)
        main_layout.addWidget(self.source_combo)

        # ---- 显示名称前缀 ----
        name_label = QLabel("显示名称:")
        name_label.setStyleSheet("font-size: 13px; color: #333; font-weight: 500;")
        main_layout.addWidget(name_label)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText('输入前缀，如「创建」→ 预览「创建年」')
        self.name_input.setFixedHeight(30)
        self.name_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #D9D9D9; border-radius: 4px;
                padding: 4px 10px; font-size: 13px; background: #FFF;
            }
            QLineEdit:focus { border-color: #FF8C00; }
        """)
        self.name_input.textChanged.connect(self._update_previews)
        main_layout.addWidget(self.name_input)

        # ---- 选择时间单位 ----
        unit_label = QLabel("选择时间单位:")
        unit_label.setStyleSheet("font-size: 13px; color: #333; font-weight: 500;")
        main_layout.addWidget(unit_label)

        units_frame = QFrame()
        units_frame.setStyleSheet("""
            QFrame {
                border: 1px solid #E0E0E0; border-radius: 6px;
                background: #FAFAFA; padding: 8px;
            }
        """)
        units_layout = QVBoxLayout(units_frame)
        units_layout.setContentsMargins(12, 8, 12, 8)
        units_layout.setSpacing(6)

        self.unit_checkboxes = {}
        self.preview_labels = {}

        for unit_key, unit_name in self.TIME_UNITS:
            row_layout = QHBoxLayout()
            row_layout.setSpacing(8)

            cb = QCheckBox(unit_name)
            cb.setStyleSheet("""
                QCheckBox { font-size: 13px; font-weight: 500; }
                QCheckBox::indicator { width: 16px; height: 16px; }
            """)
            cb.toggled.connect(lambda checked, uk=unit_key: self._on_unit_toggled(uk, checked))
            row_layout.addWidget(cb)
            self.unit_checkboxes[unit_key] = cb

            row_layout.addStretch()

            preview_label = QLabel(f"预览: {self.DEFAULT_NAMES[unit_key]}")
            preview_label.setStyleSheet("font-size: 12px; color: #999;")
            row_layout.addWidget(preview_label)
            self.preview_labels[unit_key] = preview_label

            units_layout.addLayout(row_layout)

        main_layout.addWidget(units_frame)

        main_layout.addStretch()

        # ---- 按钮行 ----
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedSize(80, 32)
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setStyleSheet("""
            QPushButton {
                border: 1px solid #D9D9D9; border-radius: 4px;
                background: #FFF; color: #333; font-size: 13px;
            }
            QPushButton:hover { border-color: #FF8C00; color: #FF8C00; }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)

        confirm_btn = QPushButton("确定")
        confirm_btn.setFixedSize(80, 32)
        confirm_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        confirm_btn.setStyleSheet("""
            QPushButton {
                border: none; border-radius: 4px;
                background: #FF8C00; color: #FFF; font-size: 13px; font-weight: 500;
            }
            QPushButton:hover { background: #E67A00; }
        """)
        confirm_btn.clicked.connect(self._on_confirm)
        btn_row.addWidget(confirm_btn)

        main_layout.addLayout(btn_row)

        self._update_previews()

    def _update_previews(self):
        """根据输入框内容实时更新预览标签。"""
        prefix = self.name_input.text().strip()
        for unit_key, unit_name in self.UNIT_LABELS.items():
            if unit_key in self.preview_labels:
                display = f"{prefix}{unit_name}" if prefix else unit_name
                cb = self.unit_checkboxes.get(unit_key)
                if cb and cb.isChecked():
                    self.preview_labels[unit_key].setText(f"预览: {display}")
                    self.preview_labels[unit_key].setStyleSheet("font-size: 12px; color: #FF8C00; font-weight: 500;")
                else:
                    self.preview_labels[unit_key].setText(f"预览: {display}")
                    self.preview_labels[unit_key].setStyleSheet("font-size: 12px; color: #999;")

    def _on_unit_toggled(self, unit_key, checked):
        """勾选状态变化时刷新预览样式。"""
        self._update_previews()

    def _on_confirm(self):
        """确定：收集勾选的单位，构建结果。"""
        prefix = self.name_input.text().strip()

        # 获取源字段
        idx = self.source_combo.currentIndex()
        if idx >= 0:
            self.source_field_name = self.source_combo.itemData(idx) or self.source_combo.currentText().strip()
        else:
            self.source_field_name = self.source_combo.currentText().strip()

        if not self.source_field_name:
            show_warning("提示", "请选择提取源字段。", self)
            return

        self.selected_units = []
        for unit_key, unit_name in self.UNIT_LABELS.items():
            cb = self.unit_checkboxes.get(unit_key)
            if cb and cb.isChecked():
                display = f"{prefix}{unit_name}" if prefix else unit_name
                self.selected_units.append({
                    'unit': unit_key,
                    'display_name': display,
                })

        if not self.selected_units:
            show_warning("提示", "请至少勾选一个时间单位。", self)
            return

        self._dialog_closed = True
        self.accept()


class ExcelFieldSettingsDialog(CenteredPopupDialog):
    """
    Excel表格字段显示和排序设置对话框。

    【重要说明】关于字段设置无法生效的Bug修复（2026-05-10）：
    =====================================================
    本类继承自 CenteredPopupDialog，该基类提供了居中显示和点击外部关闭功能。
    但在 3.3.9 版本中发现了一个严重Bug：字段设置对话框的"确定"按钮点击后，
    设置不会生效，表格不会更新。

    根本原因：
    --------
    CenteredPopupDialog 的 event() 方法在检测到 WindowDeactivate（窗口失活）事件时，
    会自动调用 reject() 关闭对话框。当用户点击"确定"按钮时：

    时序问题：
    1. 用户点击"确定"
    2. → 调用 accept() 方法
    3. → accept() 内部调用 super.accept()
    4. → 对话框开始关闭，窗口失活
    5. → 触发 WindowDeactivate 事件
    6. → CenteredPopupDialog.event() 捕获事件
    7. → event() 自动调用 self.reject()
    8. → reject() **覆盖**了 accept() 的结果
    9. → dialog.exec() 返回 Rejected（错误）
    10. → 字段设置代码不执行

    解决方案：
    --------
    在本类的 accept() 方法中，调用 super.accept() **之前**设置标志位：
        self._dialog_closed = True

    同时修改 CenteredPopupDialog 的 event() 方法，检查此标志位：
        if self._dialog_closed:
            return super().event(event)  # 跳过外部关闭逻辑

    这样确保 accept()/reject() 只执行一次，避免结果被覆盖。

    使用位置：
    --------
    本对话框类被以下模块使用：
    1. 文件生成模块 - open_column_settings_dialog() 方法
    2. CRM订单模块 - open_crm_field_settings_dialog() 方法
    3. 对象查询模块 - _open_obj_query_column_settings() 方法

    测试状态：
    --------
    ✅ 文件生成模块 - 字段设置正常生效
    ✅ CRM订单模块 - 字段设置正常生效
    ✅ 对象查询模块 - 字段设置正常生效

    版本历史：
    --------
    - 3.3.3版本：直接继承 QDialog，无此问题（但缺少居中和外部关闭功能）
    - 3.3.9版本：改继承 CenteredPopupDialog，引入此Bug
    - 2026-05-10：添加 _dialog_closed 标志位修复此Bug

    维护提醒：
    --------
    如果未来需要重写 accept() 或 reject() 方法，请务必保持 _dialog_closed
    标志位的设置逻辑，否则会导致同样的字段设置失效问题复发。
    """

    def __init__(self, headers, visible_headers=None, header_order=None, parent=None):
        super().__init__(parent, center_reference=parent, close_on_outside=True)
        self.setWindowTitle("字段设置")

        # 从配置加载上次保存的对话框尺寸
        config = load_config()
        size_config = config.get('app_settings', {}).get('dialog_sizes', {}).get('ExcelFieldSettingsDialog', None)
        self._saved_size = size_config  # None 或 [width, height]
        if size_config:
            self.resize(size_config[0], size_config[1])
        else:
            self.resize(900, 620)
        self.setMinimumSize(820, 560)

        # 延迟启用尺寸保存，避免 showEvent → center_dialog → adjustSize() 产生的
        # 程序化 resize 覆盖用户上次保存的尺寸
        self._save_size_armed = False

        self.all_headers = [str(header).strip() for header in headers if str(header).strip()]
        ordered_headers = [header for header in (header_order or []) if header in self.all_headers]
        ordered_headers.extend(header for header in self.all_headers if header not in ordered_headers)
        self.ordered_headers = ordered_headers or self.all_headers.copy()

        visible_set = {header for header in (visible_headers or self.ordered_headers) if header in self.all_headers}
        if not visible_set:
            visible_set = set(self.ordered_headers)

        self._checkboxes = {}
        self._updating_items = False
        self._updating_select_all = False

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        title_label = QLabel("勾选需要显示的字段，并在右侧拖动调整显示顺序")
        title_font = title_label.font()
        title_font.setPointSize(10)
        title_font.setBold(True)
        title_label.setFont(title_font)
        main_layout.addWidget(title_label)

        content_layout = QHBoxLayout()
        content_layout.setSpacing(16)

        left_layout = QVBoxLayout()
        left_layout.setSpacing(8)
        self.available_column_count = 4
        left_header_layout = QHBoxLayout()
        left_header_layout.setContentsMargins(0, 0, 0, 0)
        left_header_layout.addWidget(QLabel(f"全部字段（{len(self.all_headers)}）"))
        left_header_layout.addStretch()
        self.select_all_checkbox = QCheckBox("全选/取消全选")
        self.select_all_checkbox.setTristate(False)
        self.select_all_checkbox.setStyleSheet("""
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
        """)
        self.select_all_checkbox.clicked.connect(self.on_select_all_clicked)
        left_header_layout.addWidget(self.select_all_checkbox)
        left_layout.addLayout(left_header_layout)

        self.available_search_edit = QLineEdit()
        self.available_search_edit.setPlaceholderText("搜索字段")
        self.available_search_edit.textChanged.connect(self.refresh_available_grid)
        left_layout.addWidget(self.available_search_edit)

        self.available_scroll = QScrollArea()
        self.available_scroll.setWidgetResizable(True)
        self.available_container = QWidget()
        self.available_grid = QGridLayout(self.available_container)
        self.available_grid.setContentsMargins(8, 8, 8, 8)
        self.available_grid.setHorizontalSpacing(18)
        self.available_grid.setVerticalSpacing(10)
        self.available_scroll.setWidget(self.available_container)
        left_layout.addWidget(self.available_scroll, 1)

        right_layout = QVBoxLayout()
        right_layout.setSpacing(8)
        right_layout.addWidget(QLabel("显示字段（可拖动排序）"))
        self.visible_search_edit = QLineEdit()
        self.visible_search_edit.setPlaceholderText("搜索已选字段...")
        self.visible_search_edit.textChanged.connect(self._filter_visible_list)
        right_layout.addWidget(self.visible_search_edit)
        self.visible_list = QListWidget()
        self.visible_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.visible_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        right_layout.addWidget(self.visible_list, 1)

        content_layout.addLayout(left_layout, 7)
        content_layout.addLayout(right_layout, 3)
        main_layout.addLayout(content_layout, 1)

        button_layout = QHBoxLayout()
        button_layout.addStretch()
        reset_btn = QPushButton("重置")
        cancel_btn = QPushButton("取消")
        confirm_btn = QPushButton("确定")
        reset_btn.clicked.connect(self.reset_to_default)
        cancel_btn.clicked.connect(self.reject)
        confirm_btn.clicked.connect(self.accept)
        button_layout.addWidget(reset_btn)
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(confirm_btn)
        main_layout.addLayout(button_layout)

        self._build_available_checkboxes(visible_set)
        self._populate_visible_list([header for header in self.ordered_headers if header in visible_set])

    def _build_available_checkboxes(self, visible_set):
        for header in self.all_headers:
            checkbox = QCheckBox(header)
            checkbox.setStyleSheet("""
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                }
            """)
            checkbox.setChecked(header in visible_set)
            checkbox.toggled.connect(lambda checked, name=header: self.on_available_checkbox_toggled(name, checked))
            self._checkboxes[header] = checkbox
        self.refresh_available_grid()
        self.update_select_all_checkbox_state()

    def refresh_available_grid(self):
        while self.available_grid.count():
            item = self.available_grid.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)

        keyword = self.available_search_edit.text().strip().lower() if hasattr(self, 'available_search_edit') else ""
        filtered_headers = [header for header in self.all_headers if keyword in header.lower()]
        for index, header in enumerate(filtered_headers):
            row = index // self.available_column_count
            column = index % self.available_column_count
            self.available_grid.addWidget(self._checkboxes[header], row, column)
        self.available_grid.setRowStretch((len(filtered_headers) + self.available_column_count - 1) // self.available_column_count, 1)

    def _populate_visible_list(self, visible_headers):
        self._updating_items = True
        self.visible_list.clear()
        for header in visible_headers:
            item = QListWidgetItem()
            item.setData(Qt.UserRole, header)
            item_widget = self._create_visible_item_widget(header)
            item.setSizeHint(item_widget.sizeHint())
            self.visible_list.addItem(item)
            self.visible_list.setItemWidget(item, item_widget)
        self._updating_items = False
        self.update_select_all_checkbox_state()
        # 重新应用搜索过滤
        if hasattr(self, 'visible_search_edit'):
            self._filter_visible_list(self.visible_search_edit.text())

    def _create_visible_item_widget(self, header):
        item_widget = QWidget()
        item_layout = QHBoxLayout(item_widget)
        item_layout.setContentsMargins(8, 2, 8, 2)
        item_layout.setSpacing(4)

        text_label = QLabel(header)
        item_layout.addWidget(text_label, 1)

        # 上移按钮
        up_btn = QPushButton("▲")
        up_btn.setFixedSize(26, 26)
        up_btn.setToolTip("上移")
        up_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        up_btn.setStyleSheet("""
            QPushButton { border: 1px solid #D9D9D9; border-radius: 4px; font-size: 12px; color: #555; background: #FFFFFF; }
            QPushButton:hover { color: #1890FF; border-color: #1890FF; background: #E6F7FF; }
        """)
        up_btn.clicked.connect(lambda _, name=header: self._move_visible_item_up(name))
        item_layout.addWidget(up_btn)

        # 下移按钮
        down_btn = QPushButton("▼")
        down_btn.setFixedSize(26, 26)
        down_btn.setToolTip("下移")
        down_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        down_btn.setStyleSheet("""
            QPushButton { border: 1px solid #D9D9D9; border-radius: 4px; font-size: 12px; color: #555; background: #FFFFFF; }
            QPushButton:hover { color: #1890FF; border-color: #1890FF; background: #E6F7FF; }
        """)
        down_btn.clicked.connect(lambda _, name=header: self._move_visible_item_down(name))
        item_layout.addWidget(down_btn)

        # 置顶按钮
        top_button = QPushButton("置顶")
        top_button.setFixedSize(50, 26)
        top_button.setToolTip("置顶")
        top_button.setCursor(Qt.CursorShape.PointingHandCursor)
        top_button.setStyleSheet("""
            QPushButton { border: 1px solid #D9D9D9; border-radius: 4px; font-size: 11px; padding: 2px 4px; color: #555; background: #FFFFFF; }
            QPushButton:hover { color: #1890FF; border-color: #1890FF; background: #E6F7FF; }
        """)
        top_button.clicked.connect(lambda _, name=header: self.move_visible_item_to_top(name))
        item_layout.addWidget(top_button)

        return item_widget

    def _visible_contains(self, header):
        for index in range(self.visible_list.count()):
            item = self.visible_list.item(index)
            if item and item.data(Qt.UserRole) == header:
                return True
        return False

    def _get_visible_headers(self):
        visible_headers = []
        for index in range(self.visible_list.count()):
            item = self.visible_list.item(index)
            header = item.data(Qt.UserRole) if item else ""
            if header:
                visible_headers.append(header)
        return visible_headers

    def on_available_checkbox_toggled(self, header, checked):
        if self._updating_items:
            return
        visible_headers = self._get_visible_headers()
        if checked:
            if header not in visible_headers:
                visible_headers.append(header)
        else:
            visible_headers = [name for name in visible_headers if name != header]
        self._populate_visible_list(visible_headers)

    def on_select_all_clicked(self, checked):
        """全选框点击事件处理"""
        if self._updating_select_all:
            return

        if checked:
            target_checked = True
        else:
            target_checked = False

        from PyQt6.QtCore import QTimer
        QTimer.singleShot(0, lambda: self._execute_select_all(target_checked))

    def on_select_all_state_changed(self, state):
        pass

    def _execute_select_all(self, checked):
        """实际执行全选/取消全选操作"""
        try:
            self._updating_items = True
            self._updating_select_all = True

            for header, checkbox in self._checkboxes.items():
                try:
                    checkbox.toggled.disconnect()
                except Exception:
                    pass

            for checkbox in self._checkboxes.values():
                checkbox.setChecked(checked)

            for header, checkbox in self._checkboxes.items():
                checkbox.toggled.connect(lambda c, name=header: self.on_available_checkbox_toggled(name, c))

            visible_headers = self.ordered_headers.copy() if checked else []
            self._populate_visible_list_silent(visible_headers)

            final_state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
            self.select_all_checkbox.setCheckState(final_state)
        finally:
            self._updating_items = False
            self._updating_select_all = False

    def _populate_visible_list_silent(self, visible_headers):
        """静默更新显示列表"""
        self.visible_list.clear()
        for header in visible_headers:
            item = QListWidgetItem()
            item.setData(Qt.UserRole, header)
            item_widget = self._create_visible_item_widget(header)
            item.setSizeHint(item_widget.sizeHint())
            self.visible_list.addItem(item)
            self.visible_list.setItemWidget(item, item_widget)
        if hasattr(self, 'visible_search_edit'):
            self._filter_visible_list(self.visible_search_edit.text())

    def update_select_all_checkbox_state(self):
        checked_count = sum(1 for checkbox in self._checkboxes.values() if checkbox.isChecked())
        total_count = len(self._checkboxes)

        self._updating_select_all = True
        if checked_count == 0:
            self.select_all_checkbox.setCheckState(Qt.Unchecked)
        elif checked_count == total_count:
            self.select_all_checkbox.setCheckState(Qt.Checked)
        else:
            self.select_all_checkbox.setCheckState(Qt.PartiallyChecked)
        self._updating_select_all = False

    def move_visible_item_to_top(self, header):
        visible_headers = self._get_visible_headers()
        if header not in visible_headers or visible_headers[0] == header:
            return

        visible_headers.remove(header)
        visible_headers.insert(0, header)
        self._populate_visible_list(visible_headers)

    def _move_visible_item_up(self, header):
        """将指定字段上移一位"""
        visible_headers = self._get_visible_headers()
        if header not in visible_headers:
            return
        idx = visible_headers.index(header)
        if idx <= 0:
            return
        visible_headers.pop(idx)
        visible_headers.insert(idx - 1, header)
        self._populate_visible_list(visible_headers)

    def _move_visible_item_down(self, header):
        """将指定字段下移一位"""
        visible_headers = self._get_visible_headers()
        if header not in visible_headers:
            return
        idx = visible_headers.index(header)
        if idx >= len(visible_headers) - 1:
            return
        visible_headers.pop(idx)
        visible_headers.insert(idx + 1, header)
        self._populate_visible_list(visible_headers)

    def _filter_visible_list(self, text):
        """根据搜索文本过滤显示字段列表"""
        keyword = text.strip().lower()
        for index in range(self.visible_list.count()):
            item = self.visible_list.item(index)
            if item:
                header = str(item.data(Qt.UserRole) or '').lower()
                self.visible_list.setRowHidden(index, keyword not in header if keyword else False)

    def showEvent(self, event):
        """对话框显示后重新应用保存的尺寸，并延迟启用尺寸保存"""
        super().showEvent(event)
        # super().showEvent() → center_dialog() → adjustSize() 会覆盖 __init__ 中
        # 恢复的保存尺寸，因此在下一个事件循环中重新应用保存的尺寸
        if self._saved_size:
            QTimer.singleShot(0, lambda: self.resize(self._saved_size[0], self._saved_size[1]))
        # 再延迟 300ms 启用保存，确保所有程序化布局调整均已结束
        QTimer.singleShot(300, lambda: setattr(self, '_save_size_armed', True))

    def resizeEvent(self, event):
        """保存对话框调整后的尺寸（仅限用户手动调整）"""
        super().resizeEvent(event)
        if not getattr(self, '_save_size_armed', False):
            return
        try:
            config = load_config()
            if 'app_settings' not in config:
                config['app_settings'] = {}
            if 'dialog_sizes' not in config['app_settings']:
                config['app_settings']['dialog_sizes'] = {}
            new_size = [self.width(), self.height()]
            if config['app_settings']['dialog_sizes'].get('ExcelFieldSettingsDialog') != new_size:
                config['app_settings']['dialog_sizes']['ExcelFieldSettingsDialog'] = new_size
                save_config_with_delay(config)
        except Exception:
            pass  # 保存尺寸失败时不影响主要功能

    def reset_to_default(self):
        self._updating_items = True
        for checkbox in self._checkboxes.values():
            checkbox.setChecked(True)
        self._updating_items = False
        self._populate_visible_list(self.all_headers)

    def get_settings(self):
        visible_headers = self._get_visible_headers()
        header_order = visible_headers + [header for header in self.all_headers if header not in visible_headers]
        return visible_headers, header_order

    def accept(self):
        """
        确认字段设置对话框（用户点击"确定"按钮时调用）。

        【重要】此方法包含修复"字段设置无法生效"Bug的关键代码：
        -----------------------------------------------------------
        问题背景：
        本类继承自 CenteredPopupDialog，该基类在 event() 方法中实现了
        "点击外部自动关闭"功能。当对话框窗口失活时（WindowDeactivate事件），
        会自动调用 reject() 关闭对话框。

        Bug触发流程：
        1. 用户点击"确定"按钮 → 调用本 accept() 方法
        2. 本方法调用 super().accept() → 对话框开始关闭
        3. 窗口失活 → 触发 WindowDeactivate 事件
        4. CenteredPopupDialog.event() 捕获该事件
        5. event() 自动调用 self.reject()
        6. reject() 将结果设为 Rejected，**覆盖** accept() 的 Accepted
        7. dialog.exec() 返回 Rejected（错误状态）
        8. 调用方的 if dialog.exec() == Accepted: 条件不成立
        9. 字段设置代码全部跳过，表格不更新

        修复代码：
        在调用 super.accept() **之前**设置 self._dialog_closed = True
        这样当后续 WindowDeactivate 事件触发 event() 时，
        event() 检测到 _dialog_closed == True 就会直接返回，
        不再调用 reject()，从而保持 accept() 的结果不被覆盖。

        验证方式：
        - 设置几个字段后点击确定
        - 检查 dialog.exec() 返回值应为 1 (Accepted)
        - 表格应立即更新为只显示选中的字段

        影响模块：
        - 文件生成模块的字段设置
        - CRM订单模块的字段设置
        - 对象查询模块的字段设置
        """

        # 获取用户选择的可见字段列表
        visible_headers, _ = self.get_settings()

        # 验证：至少保留一个显示字段
        if not visible_headers:
            QMessageBox.warning(self, "提示", "请至少保留一个显示字段。")
            return  # 直接返回，不关闭对话框（让用户重新选择）

        # ✅【关键修复】在调用 super.accept() 之前设置标志位
        #
        # 为什么必须在这里设置？
        # ---------------------
        # 1. 时序要求：必须在 super.accept() 之前设置
        #    因为 super.accept() 会触发对话框关闭 → 窗口失活 → WindowDeactivate 事件
        #    如果不提前设置标志，event() 就会在我们调用完 super.accept() 后立刻调用 reject()
        #
        # 2. 标志作用：告诉 CenteredPopupDialog.event() "我已经处理过关闭了"
        #    event() 检测到此标志为 True 后会跳过外部关闭逻辑
        #
        # 3. 结果保证：确保 dialog.exec() 返回 QDialog.DialogCode.Accepted (值为1)
        #    而不是被覆盖为 QDialog.DialogCode.Rejected (值为0)
        #
        # 4. 业务影响：只有 exec() 返回 Accepted，调用方才会执行字段设置应用代码
        #    否则所有字段选择都白费，表现为"字段设置无法生效"
        #
        self._dialog_closed = True

        # 调用父类的 accept() 方法，正式接受对话框并关闭
        # 此时会设置对话框结果为 Accepted，并触发 closed 信号
        super().accept()

# ═══════════════════════════════════════════════════════════════
# 筛选器公共组件（FilterConditionRow / FilterPanel / ExposedTagsBar）
# ═══════════════════════════════════════════════════════════════
# 这三组件将 CRM 订单、商机、对象查询、自定义报表等页面的筛选 UI
# 共用逻辑抽取为可配置单元，统一管理条件行的创建/编辑/删除。
#
# 配置维度（通过构造参数注入，无需子类化）：
#   - field_options       : 可选字段列表
#   - text_operators      : 文本字段的操作符集合
#   - date_operators      : 日期字段的操作符集合
#   - value_pages         : 值输入区域启用的页面类型
#   - show_expose / show_picker / show_remove : 可选 UI 控件开关
#   - debounce_ms         : 值变化防抖间隔
#   - is_date_field_cb    : 判断字段是否为日期类型的回调
#   - picker_cb           : 多选按钮点击回调
#   - style               : 样式预设名（如 'crm' / 'compact'）
#
# 对现有功能的影响：零（所有类为纯新增，暂不被引用）。
# ═══════════════════════════════════════════════════════════════

# ── 需数字输入的操作符集合（过去/未来 N 天/周/月/季度等） ──
_N_TYPE_OPERATORS = frozenset({
    'past_n_days', 'future_n_days',
    'past_n_days_exclusive', 'future_n_days_exclusive',
    'past_n_months_exclusive', 'future_n_months_exclusive',
    'past_n_weeks_exclusive', 'future_n_weeks_exclusive',
    'past_n_days_inclusive', 'future_n_days_inclusive',
    'past_n_weeks_inclusive', 'future_n_weeks_inclusive',
    'past_n_months_inclusive', 'future_n_months_inclusive',
    'n_days_ago', 'n_days_later',
    'n_weeks_ago', 'n_weeks_later',
    'past_n_quarters_inclusive',
})

# ── 日期相关操作符（用于判断是否展示日期输入） ──
_DATE_OPS = frozenset({
    'eq', 'ne', 'date_before', 'date_after',
    'date_before_eq', 'date_after_eq', 'date_range',
    'gt', 'lt', 'gte', 'lte',
    'past_n_days', 'future_n_days',
    'past_n_days_exclusive', 'future_n_days_exclusive',
    'past_n_months_exclusive', 'future_n_months_exclusive',
    'past_n_weeks_exclusive', 'future_n_weeks_exclusive',
    'past_n_days_inclusive', 'future_n_days_inclusive',
    'past_n_weeks_inclusive', 'future_n_weeks_inclusive',
    'past_n_months_inclusive', 'future_n_months_inclusive',
    'n_days_ago', 'n_days_later',
    'n_weeks_ago', 'n_weeks_later',
    'past_n_quarters_inclusive',
})

# ── 隐藏值输入的操作符（empty / not_empty 不需要输入值） ──
_BLANK_OPS = frozenset({'empty', 'not_empty'})


# ═══════════════════════════════════════════════════════════════
# 预设操作符工厂（各页面可用不同子集）
# ═══════════════════════════════════════════════════════════════

def default_text_operators():
    """默认文本字段操作符（14 项）。"""
    return [
        ("等于", "eq"), ("不等于", "ne"),
        ("包含", "contains"), ("不包含", "not_contains"),
        ("属于", "in"), ("不属于", "not_in"),
        ("为空（未填写）", "empty"), ("不为空", "not_empty"),
        ("开头是", "starts_with"), ("结尾是", "ends_with"),
        ("大于", "gt"), ("小于", "lt"),
        ("大于等于", "gte"), ("小于等于", "lte"),
    ]


def default_date_operators():
    """默认日期字段操作符（25 项）。"""
    return [
        ("等于", "eq"), ("不等于", "ne"),
        ("早于", "date_before"), ("晚于", "date_after"),
        ("早于等于", "date_before_eq"), ("晚于等于", "date_after_eq"),
        ("为空（未填写）", "empty"), ("不为空", "not_empty"),
        ("时间段", "date_range"),
        ("过去N天内(不含当天)", "past_n_days_exclusive"),
        ("未来N天内(不含当天)", "future_n_days_exclusive"),
        ("过去N月内(不含当月)", "past_n_months_exclusive"),
        ("未来N月内(不含当月)", "future_n_months_exclusive"),
        ("过去N周内(不含当周)", "past_n_weeks_exclusive"),
        ("未来N周内(不含当周)", "future_n_weeks_exclusive"),
        ("过去N天内(含当天)", "past_n_days_inclusive"),
        ("未来N天内(含当天)", "future_n_days_inclusive"),
        ("过去N周内(含当周)", "past_n_weeks_inclusive"),
        ("未来N周内(含当周)", "future_n_weeks_inclusive"),
        ("过去N月内(含当月)", "past_n_months_inclusive"),
        ("未来N月内(含当月)", "future_n_months_inclusive"),
        ("N天前", "n_days_ago"), ("N天后", "n_days_later"),
        ("N周前", "n_weeks_ago"), ("N周后", "n_weeks_later"),
        ("过去N季度内(含当季度)", "past_n_quarters_inclusive"),
    ]


def crm_date_operators():
    """CRM 订单使用的日期操作符（10 项，含 past_n_days 简版）。"""
    return [
        ("等于", "eq"), ("不等于", "ne"),
        ("早于", "lt"), ("晚于", "gt"),
        ("早于等于", "lte"), ("晚于等于", "gte"),
        ("为空", "empty"), ("不为空", "not_empty"),
        ("时间段", "date_range"), ("过去N天", "past_n_days"),
    ]


def simple_date_operators():
    """简化日期操作符（7 项，商机/对象查询使用）。"""
    return [
        ("等于", "eq"), ("不等于", "ne"),
        ("早于", "lt"), ("晚于", "gt"),
        ("区间", "date_range"),
        ("为空", "empty"), ("不为空", "not_empty"),
    ]


# ═══════════════════════════════════════════════════════════════
# FilterConditionRow — 单行筛选条件
# ═══════════════════════════════════════════════════════════════

class FilterConditionRow(QFrame):
    """单行筛选条件 — 字段 + 操作符 + 值 + [外露] + [多选]

    所有可配置维度通过构造参数注入，消除子类化需求。

    信号：
        filtersChanged()      值变化时发射（含 debounce）
        rowRemoveRequested()  用户点击删除按钮

    公共方法：
        get_condition()       返回条件字典
        set_condition(dict)   从字典恢复 UI 状态
        set_field_options(list)  更新可选字段列表
        clear_value()         清空当前值
        update_input_mode()   根据字段类型切换输入控件
    """

    filtersChanged = pyqtSignal()
    rowRemoveRequested = pyqtSignal()

    def __init__(self, parent=None, *,
                 # ── 字段 ──
                 field_options: list = None,
                 field_placeholder: str = "选择字段",
                 field_width: int = 140,
                 # ── 操作符 ──
                 text_operators: list = None,
                 date_operators: list = None,
                 op_width: int = 95,
                 default_operator: str = "contains",
                 # ── 值输入页面配置 ──
                 value_pages: tuple = ("text", "date", "date_range", "spin"),
                 # ── 可选 UI 控件 ──
                 show_expose: bool = True,
                 show_picker: bool = False,
                 show_remove: bool = True,
                 expose_label: str = "外露",
                 # ── 行为 ──
                 debounce_ms: int = 250,
                 is_date_field_cb: callable = None,
                 picker_cb: callable = None,
                 # ── 样式 ──
                 row_height: int = 32,
                 row_bg: str = "#F8F8F8",
                 remove_text: str = "✕",
                 ):
        super().__init__(parent)
        self.setFixedHeight(row_height)
        self.setStyleSheet(
            f"QFrame {{ background: {row_bg}; border: none; border-radius: 4px; }}")

        # 保存配置
        self._field_options = field_options or []
        self._text_operators = text_operators or default_text_operators()
        self._date_operators = date_operators or default_date_operators()
        self._default_operator = default_operator
        self._value_pages = value_pages
        self._show_picker = show_picker
        self._debounce_ms = debounce_ms
        self._is_date_field_cb = is_date_field_cb
        self._picker_cb = picker_cb
        self._current_field_is_date = False

        # 防抖定时器
        if debounce_ms > 0:
            self._debounce_timer = QTimer(singleShot=True,
                                          interval=debounce_ms)
            self._debounce_timer.timeout.connect(self.filtersChanged)
        else:
            self._debounce_timer = None

        # ── 构建 UI ──
        layout = QHBoxLayout(self)
        layout.setContentsMargins(6, 0, 6, 0)
        layout.setSpacing(5)

        # 删除按钮
        if show_remove:
            self._remove_btn = QPushButton(remove_text)
            self._remove_btn.setFixedSize(20, 20)
            self._remove_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            self._remove_btn.setStyleSheet(
                "QPushButton { font-weight: bold; color: #FF4D4F; "
                "border: 1px solid #FFCCC7; border-radius: 10px; "
                "font-size: 11px; background: #FFF2F0; } "
                "QPushButton:hover { background: #FF4D4F; color: #FFF; }")
            self._remove_btn.clicked.connect(self.rowRemoveRequested)
            layout.addWidget(self._remove_btn,
                             alignment=Qt.AlignmentFlag.AlignVCenter)
        else:
            self._remove_btn = None

        # 字段下拉
        self._field_combo = QComboBox()
        self._field_combo.setEditable(True)
        self._field_combo.setFixedWidth(field_width)
        self._field_combo.setFixedHeight(26)
        self._field_combo.setPlaceholderText(field_placeholder)
        self._field_combo.setStyleSheet(
            "QComboBox { font-size: 12px; border: 1px solid #D9D9D9; "
            "border-radius: 3px; padding: 2px 4px; background: #FFF; }")
        self._populate_field_combo()
        layout.addWidget(self._field_combo,
                         alignment=Qt.AlignmentFlag.AlignVCenter)

        # 操作符下拉
        self._op_combo = QComboBox()
        self._op_combo.setFixedWidth(op_width)
        self._op_combo.setFixedHeight(26)
        self._op_combo.setStyleSheet(
            "QComboBox { font-size: 12px; border: 1px solid #D9D9D9; "
            "border-radius: 3px; padding: 2px 2px; background: #FFF; }")
        self._populate_op_combo(use_text_ops=True)
        layout.addWidget(self._op_combo,
                         alignment=Qt.AlignmentFlag.AlignVCenter)

        # 值输入区域（QStackedWidget）
        self._value_stack = QStackedWidget()
        self._value_stack.setFixedHeight(26)
        self._value_stack.setStyleSheet(
            "QStackedWidget { background: transparent; }")

        # Page 0: 文本输入
        self._text_input = QLineEdit()
        self._text_input.setFixedHeight(26)
        self._text_input.setPlaceholderText("值")
        self._text_input.setStyleSheet(
            "font-size: 12px; border: 1px solid #D9D9D9; "
            "border-radius: 3px; padding: 2px 4px;")
        self._text_input.textChanged.connect(self._on_value_changed)
        self._value_stack.addWidget(self._text_input)

        # Page 1: 单日期选择器
        if "date" in value_pages:
            self._date_input = QDateEdit()
            self._date_input.setFixedHeight(26)
            self._date_input.setCalendarPopup(True)
            self._date_input.setDisplayFormat("yyyy-MM-dd")
            self._date_input.setDate(QDate.currentDate())
            self._date_input.setStyleSheet(
                "font-size: 12px; border: 1px solid #D9D9D9; "
                "border-radius: 3px; padding: 2px 4px;")
            self._date_input.dateChanged.connect(self._on_value_changed)
            self._value_stack.addWidget(self._date_input)
        else:
            self._date_input = None

        # Page 2: 日期范围按钮
        if "date_range" in value_pages:
            self._date_range_btn = QPushButton("选择日期范围")
            self._date_range_btn.setFixedHeight(26)
            self._date_range_btn.setCursor(
                Qt.CursorShape.PointingHandCursor)
            self._date_range_btn.setStyleSheet(
                "QPushButton { font-size: 12px; border: 1px solid #D9D9D9; "
                "border-radius: 3px; padding: 2px 6px; background: #FFF; } "
                "QPushButton:hover { border-color: #1890FF; }")
            self._date_range_btn.setProperty('date_range',
                                              {'start': None, 'end': None})
            self._date_range_btn.clicked.connect(
                lambda: self._on_date_range_clicked())
            self._value_stack.addWidget(self._date_range_btn)
        else:
            self._date_range_btn = None

        # Page 3: 数字微调（N天/月/周）
        if "spin" in value_pages:
            self._spin_input = QSpinBox()
            self._spin_input.setFixedHeight(26)
            self._spin_input.setRange(1, 3650)
            self._spin_input.setValue(7)
            self._spin_input.setSuffix("")
            self._spin_input.setStyleSheet(
                "font-size: 12px; border: 1px solid #D9D9D9; "
                "border-radius: 3px; padding: 2px 4px;")
            self._spin_input.valueChanged.connect(self._on_value_changed)
            self._value_stack.addWidget(self._spin_input)
        else:
            self._spin_input = None

        self._value_stack.setCurrentIndex(0)
        layout.addWidget(self._value_stack, stretch=1)

        # 多选按钮
        if show_picker:
            self._picker_btn = QPushButton("📋")
            self._picker_btn.setFixedSize(24, 26)
            self._picker_btn.setToolTip("从选项中选择多个值")
            self._picker_btn.setCursor(
                Qt.CursorShape.PointingHandCursor)
            self._picker_btn.setStyleSheet(
                "QPushButton { font-size: 12px; "
                "border: 1px solid #D9D9D9; border-radius: 3px; "
                "background: #FFF; } "
                "QPushButton:hover { border-color: #1890FF; "
                "background: #E6F7FF; }")
            self._picker_btn.setVisible(False)
            self._picker_btn.clicked.connect(self._on_picker_clicked)
            layout.addWidget(self._picker_btn,
                             alignment=Qt.AlignmentFlag.AlignVCenter)
        else:
            self._picker_btn = None

        # 外露复选框
        if show_expose:
            self._expose_check = QCheckBox(expose_label)
            self._expose_check.setChecked(False)
            self._expose_check.setToolTip("勾选后在表格上方显示")
            self._expose_check.setStyleSheet(
                "QCheckBox { font-size: 11px; color: #666; }")
            layout.addWidget(self._expose_check,
                             alignment=Qt.AlignmentFlag.AlignVCenter)
        else:
            self._expose_check = None

        # ── 信号连接 ──
        self._field_combo.currentIndexChanged.connect(
            lambda idx: self._on_field_changed())
        self._field_combo.editTextChanged.connect(
            lambda text: self._on_field_edit_changed(text))
        self._op_combo.currentIndexChanged.connect(
            lambda idx: self._on_op_changed(idx))

    # ──────────────────────────────────────────────────────────
    # 内部 UI 填充
    # ──────────────────────────────────────────────────────────

    def _populate_field_combo(self):
        """填充字段下拉列表。"""
        self._field_combo.blockSignals(True)
        self._field_combo.clear()
        for item in self._field_options:
            if isinstance(item, (tuple, list)):
                label = str(item[0])
                data = item[1] if len(item) > 1 else label
            else:
                label = str(item)
                data = label
            self._field_combo.addItem(label, data)
        self._field_combo.blockSignals(False)

    def _populate_op_combo(self, use_text_ops: bool = True):
        """填充操作符下拉列表。"""
        ops = self._text_operators if use_text_ops else self._date_operators
        self._op_combo.blockSignals(True)
        self._op_combo.clear()
        for label, key in ops:
            self._op_combo.addItem(label, key)
        self._op_combo.blockSignals(False)

    # ──────────────────────────────────────────────────────────
    # 信号处理
    # ──────────────────────────────────────────────────────────

    def _on_value_changed(self, *_):
        """值变化时启动防抖定时器。"""
        if self._debounce_timer:
            self._debounce_timer.start()
        else:
            self.filtersChanged.emit()

    def _on_field_changed(self):
        """字段变化时更新输入模式。"""
        self.update_input_mode()

    def _on_field_edit_changed(self, text):
        """字段编辑文本变化（用于 IME 兼容延迟处理）。"""
        # 保留延迟更新逻辑入口，子类或外部可覆盖
        if self._debounce_timer:
            self._debounce_timer.start()

    def _on_op_changed(self, idx):
        """操作符变化时切换值输入页面。"""
        op = self._op_combo.currentData()
        if op is None:
            return
        op_str = str(op)

        if op_str in _BLANK_OPS:
            self._value_stack.hide()
            return

        self._value_stack.show()

        if op_str == 'date_range':
            if self._date_range_btn:
                self._value_stack.setCurrentWidget(self._date_range_btn)
            return

        if op_str in _N_TYPE_OPERATORS:
            if self._spin_input:
                self._spin_input.setSuffix(self._spin_suffix_for_op(op_str))
                self._value_stack.setCurrentWidget(self._spin_input)
            return

        # 判断是否为日期类操作符（需要日期输入）
        if self._current_field_is_date or op_str in _DATE_OPS:
            if self._date_input:
                self._value_stack.setCurrentWidget(self._date_input)
            else:
                self._value_stack.setCurrentIndex(0)
            return

        self._value_stack.setCurrentIndex(0)

    def _on_date_range_clicked(self):
        """日期范围按钮点击 — 打开 QuickDatePickerDialog 弹窗。"""
        from PyQt6.QtCore import QDate
        from PyQt6.QtWidgets import QDialog
        dr = self._date_range_btn.property('date_range') or {}
        parent = self.window()

        dlg = QuickDatePickerDialog(
            start_date=dr.get('start'),
            end_date=dr.get('end'),
            parent=parent)
        dlg.set_popup_anchor_widget(self._date_range_btn)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.start_date and dlg.end_date:
            self._date_range_btn.setProperty(
                'date_range', {'start': dlg.start_date, 'end': dlg.end_date})
            self._date_range_btn.setText(
                f"{dlg.start_date.toString('yyyy-MM-dd')} ~ "
                f"{dlg.end_date.toString('yyyy-MM-dd')}")
            self._on_value_changed()

    def _on_picker_clicked(self):
        """多选按钮点击 — 委托给外部回调。"""
        if self._picker_cb:
            self._picker_cb(self)
        else:
            print("[FilterConditionRow] 未配置多选回调 (picker_cb)")

    # ──────────────────────────────────────────────────────────
    # 辅助
    # ──────────────────────────────────────────────────────────

    @staticmethod
    def _spin_suffix_for_op(op: str) -> str:
        """根据操作符类型返回合适的后缀。"""
        if 'month' in op or 'quarter' in op:
            return " 月"
        if 'week' in op:
            return " 周"
        return " 天"

    # ──────────────────────────────────────────────────────────
    # 公共 API
    # ──────────────────────────────────────────────────────────

    def set_field_options(self, options: list):
        """重新设置可选字段列表并刷新下拉。"""
        self._field_options = options or []
        self._populate_field_combo()

    def set_field(self, field_label: str):
        """设置当前选中字段（按 label 匹配）。"""
        idx = self._field_combo.findText(field_label)
        if idx >= 0:
            self._field_combo.setCurrentIndex(idx)
        else:
            self._field_combo.setEditText(field_label)

    def set_operator(self, op_key: str):
        """设置当前操作符（按 key 匹配）。"""
        idx = self._op_combo.findData(op_key)
        if idx >= 0:
            self._op_combo.setCurrentIndex(idx)

    def set_value(self, value):
        """设置当前值（根据当前输入模式写入对应控件）。"""
        text = str(value) if value is not None else ''
        self._text_input.setText(text)
        if self._date_input:
            if text:
                parsed = QDate.fromString(text[:10], "yyyy-MM-dd")
                if parsed.isValid():
                    self._date_input.setDate(parsed)
        if self._spin_input and text.isdigit():
            self._spin_input.setValue(int(text))

    def clear_value(self):
        """清空当前值。"""
        self._text_input.clear()
        if self._date_input:
            self._date_input.setDate(QDate.currentDate())
        if self._date_range_btn:
            self._date_range_btn.setText("选择日期范围")
            self._date_range_btn.setProperty(
                'date_range', {'start': None, 'end': None})
        if self._spin_input:
            self._spin_input.setValue(7)

    def get_condition(self) -> dict:
        """获取当前行的条件字典。

        返回：
            {'field': str, 'operator': str, 'value': str,
             'expose': bool, 'is_date': bool}
        """
        field = self._field_combo.currentText().strip()
        operator = self._op_combo.currentData()
        if operator is None:
            operator = ''
        else:
            operator = str(operator)

        expose = (self._expose_check.isChecked()
                  if self._expose_check else False)

        if operator in _BLANK_OPS:
            value = ''
        elif operator == 'date_range' and self._date_range_btn:
            r = self._date_range_btn.property('date_range') or {}
            s = r.get('start')
            e = r.get('end')
            if s is not None and e is not None:
                value = (f"{s.toString('yyyy-MM-dd')} ~ "
                         f"{e.toString('yyyy-MM-dd')}")
            else:
                value = self._date_range_btn.text() or ''
        elif operator in _N_TYPE_OPERATORS and self._spin_input:
            value = str(self._spin_input.value())
        elif self._current_field_is_date and self._date_input:
            value = self._date_input.date().toString('yyyy-MM-dd')
        elif self._value_stack.currentWidget() is getattr(
                self, '_date_input', None) and self._date_input:
            value = self._date_input.date().toString('yyyy-MM-dd')
        else:
            value = self._text_input.text().strip()

        return {
            'field': field,
            'operator': operator,
            'value': value,
            'expose': expose,
            'is_date': self._current_field_is_date,
        }

    def set_condition(self, condition: dict):
        """从条件字典恢复 UI 状态。"""
        c = condition or {}
        self.set_field(c.get('field', ''))
        self.set_operator(c.get('operator', self._default_operator))

        # 根据字段类型调整操作符列表
        if self._is_date_field_cb and c.get('field'):
            is_date = self._is_date_field_cb(c['field'])
        else:
            is_date = c.get('is_date', False)
        self._current_field_is_date = is_date
        if is_date:
            self._populate_op_combo(use_text_ops=False)
        else:
            self._populate_op_combo(use_text_ops=True)

        # 恢复操作符
        self.set_operator(c.get('operator', self._default_operator))

        # 恢复值
        val = c.get('value', '')
        self._text_input.setText(str(val) if val else '')
        if self._date_input and val:
            parsed = QDate.fromString(str(val)[:10], "yyyy-MM-dd")
            if parsed.isValid():
                self._date_input.setDate(parsed)
        if self._date_range_btn and val and '~' in str(val):
            parts = str(val).split('~')
            if len(parts) == 2:
                s = QDate.fromString(parts[0].strip(), 'yyyy-MM-dd')
                e = QDate.fromString(parts[1].strip(), 'yyyy-MM-dd')
                if s.isValid() and e.isValid():
                    self._date_range_btn.setProperty('date_range',
                                                      {'start': s, 'end': e})
                    self._date_range_btn.setText(
                        f"{s.toString('yyyy-MM-dd')} ~ "
                        f"{e.toString('yyyy-MM-dd')}")

        # 恢复外露
        if self._expose_check:
            self._expose_check.setChecked(bool(c.get('expose', False)))

        # 触发输入模式更新
        self.update_input_mode()

    def update_input_mode(self):
        """根据当前字段类型自动切换操作符列表和输入控件。"""
        field_text = self._field_combo.currentText().strip()
        if not field_text:
            return

        if self._is_date_field_cb:
            self._current_field_is_date = self._is_date_field_cb(field_text)
        else:
            self._current_field_is_date = self._is_date_by_keyword(field_text)

        current_op = self._op_combo.currentData()
        current_op_str = str(current_op) if current_op else ''

        self._op_combo.blockSignals(True)
        self._op_combo.clear()

        if self._current_field_is_date:
            for label, key in self._date_operators:
                self._op_combo.addItem(label, key)
            if current_op_str:
                idx = self._op_combo.findData(current_op_str)
            else:
                idx = self._op_combo.findData('date_range')
        else:
            for label, key in self._text_operators:
                self._op_combo.addItem(label, key)
            if current_op_str:
                idx = self._op_combo.findData(current_op_str)
            else:
                idx = self._op_combo.findData(self._default_operator)

        if idx >= 0:
            self._op_combo.setCurrentIndex(idx)
        self._op_combo.blockSignals(False)

        # 多选按钮可见性
        new_op = self._op_combo.currentData()
        if self._picker_btn:
            self._picker_btn.setVisible(
                str(new_op) in ('in', 'not_in'))

        self._on_op_changed(self._op_combo.currentIndex())

    @staticmethod
    def _is_date_by_keyword(field_label: str) -> bool:
        """通过关键字判断字段是否为日期类型（fallback）。"""
        label_lower = (field_label or '').lower()
        date_keywords = (
            '日期', '时间', 'date', 'time', '创建', '修改',
            'birth', '生日', '成立', '到期', '结束', '审核',
            '批准', '签约', '签订', '交付',
        )
        return any(k in label_lower for k in date_keywords)

    def set_is_date_field_callback(self, cb: callable):
        """设置判断日期字段的回调。"""
        self._is_date_field_cb = cb

    def set_picker_callback(self, cb: callable):
        """设置多选按钮点击回调。"""
        self._picker_cb = cb
        if self._picker_btn:
            self._picker_btn.clicked.disconnect()
            self._picker_btn.clicked.connect(
                lambda: cb(self) if cb else None)

    # ──────────────────────────────────────────────────────────
    # 属性访问器（方便外部读取控件）
    # ──────────────────────────────────────────────────────────

    @property
    def field_combo(self):
        return self._field_combo

    @property
    def op_combo(self):
        return self._op_combo

    @property
    def value_stack(self):
        return self._value_stack

    @property
    def text_input(self):
        return self._text_input

    @property
    def date_input(self):
        return self._date_input

    @property
    def date_range_btn(self):
        return self._date_range_btn

    @property
    def spin_input(self):
        return self._spin_input

    @property
    def picker_btn(self):
        return self._picker_btn

    @property
    def expose_check(self):
        return self._expose_check

    @property
    def is_date_field(self):
        return self._current_field_is_date


# ═══════════════════════════════════════════════════════════════
# ExposedTagsBar — 外露筛选条件标签栏
# ═══════════════════════════════════════════════════════════════

class ExposedTagsBar(QFrame):
    """外露筛选条件标签栏 — 在搜索框下方以标签形式展示勾选"外露"的条件。

    信号：
        tagClicked(dict)   点击标签，传递条件字典
        tagRemoved(dict)   移除标签，传递条件字典
        tagsChanged()      标签集合发生变化
    """

    tagClicked = pyqtSignal(dict)
    tagRemoved = pyqtSignal(dict)
    tagsChanged = pyqtSignal()

    def __init__(self, parent=None, *,
                 max_tag_text_length: int = 10,
                 max_tag_value_length: int = 8,
                 tag_style: str = "default",
                 ):
        super().__init__(parent)
        self.setVisible(False)
        self.setStyleSheet(
            "QFrame { background-color: transparent; border: none; }")
        self._max_field_len = max_tag_text_length
        self._max_value_len = max_tag_value_length
        self._tag_style = tag_style
        self._tags: list[dict] = []    # 当前标签对应的条件列表
        self._tag_widgets: list = []   # (frame, label, close_btn)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 2, 0, 2)
        layout.setSpacing(6)
        layout.addStretch()
        self._tags_layout = layout

    def refresh(self, conditions: list[dict]):
        """根据条件列表重建标签（仅保留 expose=True 的条目）。

        conditions: 由 FilterConditionRow.get_condition() 返回的列表。
        """
        # 清理旧标签
        self._clear_widgets()

        exposed = [c for c in conditions if c.get('expose', False)
                   and c.get('field', '').strip()]
        self._tags = exposed

        for i, cond in enumerate(exposed):
            tag_frame = self._create_tag(cond, i)
            # 插入到 stretch 之前
            self._tags_layout.insertWidget(
                self._tags_layout.count() - 1, tag_frame)

        self.setVisible(len(exposed) > 0)
        self.tagsChanged.emit()

    def get_exposed_conditions(self) -> list[dict]:
        """获取当前所有外露标签对应的条件。"""
        return list(self._tags)

    def _clear_widgets(self):
        """移除所有标签控件。"""
        for tag_frame in self._tag_widgets:
            tag_frame.setParent(None)
            tag_frame.deleteLater()
        self._tag_widgets.clear()
        self._tags.clear()

    def _create_tag(self, cond: dict, index: int) -> QFrame:
        """创建一个标签控件。"""
        field = cond.get('field', '')
        operator = cond.get('operator', '')
        value = cond.get('value', '')

        # 截断显示
        field_short = (field[:self._max_field_len] + '…'
                       if len(field) > self._max_field_len else field)
        val_short = (value[:self._max_value_len] + '…'
                     if len(value) > self._max_value_len else value)

        if value and operator not in _BLANK_OPS:
            label_text = f"{field_short}: {val_short}"
        else:
            label_text = field_short

        tag_frame = QFrame()
        tag_frame.setStyleSheet(
            "QFrame { background: #E6F7FF; border: 1px solid #91D5FF; "
            "border-radius: 10px; padding: 2px 6px; }")
        tag_layout = QHBoxLayout(tag_frame)
        tag_layout.setContentsMargins(6, 2, 2, 2)
        tag_layout.setSpacing(4)

        lbl = QLabel(label_text)
        lbl.setStyleSheet(
            "QLabel { font-size: 11px; color: #1890FF; "
            "border: none; background: transparent; }")
        lbl.setCursor(Qt.CursorShape.PointingHandCursor)
        lbl.mousePressEvent = (
            lambda ev, c=cond: self.tagClicked.emit(c))
        tag_layout.addWidget(lbl)

        close_btn = QPushButton("×")
        close_btn.setFixedSize(14, 14)
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setStyleSheet(
            "QPushButton { font-size: 10px; color: #91D5FF; "
            "border: none; background: transparent; } "
            "QPushButton:hover { color: #1890FF; }")
        close_btn.clicked.connect(lambda: self._on_tag_close(cond))
        tag_layout.addWidget(close_btn)

        self._tag_widgets.append(tag_frame)
        return tag_frame

    def _on_tag_close(self, cond: dict):
        self.tagRemoved.emit(cond)


# ═══════════════════════════════════════════════════════════════
# FilterPanel — 多行筛选面板
# ═══════════════════════════════════════════════════════════════

class FilterPanel(QWidget):
    """多行筛选条件面板 — 管理多个 FilterConditionRow + 可选外露标签栏。

    面板模式（mode）：
      - "inline"   : 内嵌面板，初始隐藏，通过 toggle() 显示/隐藏。
                     适用于 CRM 订单/商机/对象查询。
      - "embedded" : 始终可见的面板，作为编辑器的一部分嵌入。
                     适用于自定义报表编辑器（参考 filter_bar.py）。
      - "popup"    : 独立 Dialog 弹窗。由 FilterPanel 内部创建
                     CenteredPopupDialog 包装。适用于自定义报表筛选弹窗。

    信号：
        filtersChanged()  任何行的值变化时发射
        panelToggled(bool)  inline 模式下面板显示/隐藏

    公共方法：
        add_row(condition)           添加一行，返回 FilterConditionRow
        remove_row(row)              移除指定行
        clear_all()                  清除所有行
        get_all_conditions()         获取所有行的条件列表
        load_conditions(list)        从条件列表恢复到面板
        toggle()                     切换面板可见性（inline 模式）
        refresh_exposed_tags()       刷新外露标签栏
        set_available_fields(list)   批量更新所有行的字段选项
    """

    filtersChanged = pyqtSignal()
    panelToggled = pyqtSignal(bool)

    def __init__(self, parent=None, *,
                 # ── 面板模式 ──
                 mode: str = "inline",
                 # ── 条件行默认配置 ──
                 row_defaults: dict = None,
                 # ── 面板 UI 文案 ──
                 title: str = "设置筛选",
                 add_btn_text: str = "+ 添加条件",
                 apply_btn_text: str = "筛选",
                 clear_btn_text: str = "清除筛选值",
                 save_btn_text: str = "另存为",
                 # ── 面板 UI 显示 ──
                 show_title: bool = True,
                 show_add_btn: bool = True,
                 show_apply_btn: bool = True,
                 show_clear_btn: bool = True,
                 show_save_btn: bool = False,
                 show_exposed_tags: bool = False,
                 # ── 面板 UI 开关 ──
                 toggle_btn: QPushButton = None,
                 toggle_badge: bool = True,
                 # ── 约束 ──
                 max_condition_rows: int = 30,
                 min_rows_to_allow_clear_all: int = 1,
                 # ── 弹窗模式 ──
                 popup_parent=None,
                 # ── 扩展回调 ──
                 on_save_preset: callable = None,
                 on_load_preset: callable = None,
                 on_add_row: callable = None,
                 on_apply: callable = None,
                 on_clear: callable = None,
                 # ── 内部保留旧 row_info 兼容（详见 2.4.2 节） ──
                 _legacy_row_info_list: list = None,  # 兼容旧代码
                 ):
        super().__init__(parent)
        self._mode = mode
        self._row_defaults = row_defaults or {}
        self._title = title
        self._add_btn_text = add_btn_text
        self._show_add_btn = show_add_btn
        self._show_clear_btn = show_clear_btn
        self._show_apply_btn = show_apply_btn
        self._show_save_btn = show_save_btn
        self._show_exposed_tags = show_exposed_tags
        self._show_title = show_title
        self._toggle_btn = toggle_btn
        self._toggle_badge = toggle_badge
        self._max_rows = max_condition_rows
        self._on_save_preset = on_save_preset
        self._on_apply = on_apply
        self._on_clear = on_clear
        self._on_add_row_cb = on_add_row

        self._rows: list[FilterConditionRow] = []
        # 兼容层：如果外部传入遗留的 row_info list，同步维护
        self._legacy_rows = (_legacy_row_info_list
                             if _legacy_row_info_list is not None else [])

        self._rows_layout = QVBoxLayout()
        self._rows_layout.setSpacing(6)

        if mode == "popup":
            # 弹窗模式：构建完整 UI（标题 + 条件区 + 按钮栏）
            self._build_popup_ui()
        elif mode == "inline":
            self._build_inline_ui()
        else:
            # embedded 模式：仅条件区（外部自行嵌入）
            self._build_embedded_ui()

    # ──────────────────────────────────────────────────────────
    # UI 构建（三种模式）
    # ──────────────────────────────────────────────────────────

    def _build_inline_ui(self):
        """内嵌模式：可折叠面板。"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(6)

        # 面板容器（初始隐藏）
        self._panel_frame = QFrame()
        self._panel_frame.setVisible(False)
        self._panel_frame.setStyleSheet(
            "QFrame#filterPanelFrame { "
            "background: #FAFAFA; border: 1px solid #E8E8E8; "
            "border-radius: 6px; }")
        self._panel_frame.setObjectName("filterPanelFrame")

        panel_layout = QVBoxLayout(self._panel_frame)
        panel_layout.setContentsMargins(12, 10, 12, 10)
        panel_layout.setSpacing(8)

        if self._show_title:
            title_row = QHBoxLayout()
            title_lbl = QLabel(self._title)
            title_lbl.setStyleSheet(
                "font-size: 14px; font-weight: bold; color: #333; "
                "border: none; background: transparent;")
            title_row.addWidget(title_lbl)
            title_row.addStretch()
            panel_layout.addLayout(title_row)

        # 条件行区域（可滚动）
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(
            "QScrollArea { border: none; background: transparent; }")
        scroll.setMaximumHeight(300)
        rows_widget = QWidget()
        rows_widget.setStyleSheet("background: transparent;")
        rows_widget.setLayout(self._rows_layout)
        scroll.setWidget(rows_widget)
        panel_layout.addWidget(scroll, stretch=1)

        # 底部按钮栏
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        if self._show_save_btn:
            save_btn = QPushButton(self.save_btn_text if hasattr(self, 'save_btn_text') else "另存为")
            save_btn.setFixedHeight(28)
            save_btn.setStyleSheet(
                "QPushButton { border: 1px solid #D9D9D9; "
                "border-radius: 4px; padding: 4px 12px; font-size: 12px; "
                "background: #FFF; } "
                "QPushButton:hover { border-color: #1890FF; color: #1890FF; }")
            save_btn.clicked.connect(
                lambda: self._on_save_preset and self._on_save_preset(self))
            btn_row.addWidget(save_btn)

        btn_row.addStretch()

        if self._show_add_btn:
            add_btn = QPushButton(self._add_btn_text)
            add_btn.setFixedHeight(28)
            add_btn.setStyleSheet(
                "QPushButton { border: 1px dashed #1890FF; "
                "border-radius: 4px; padding: 4px 12px; font-size: 12px; "
                "color: #1890FF; background: #FFF; } "
                "QPushButton:hover { background: #E6F7FF; }")
            add_btn.clicked.connect(lambda: self.add_row())
            btn_row.addWidget(add_btn)

        if self._show_clear_btn:
            clear_btn = QPushButton(self.clear_btn_text if hasattr(self, 'clear_btn_text') else "清除")
            clear_btn.setFixedHeight(28)
            clear_btn.setStyleSheet(
                "QPushButton { border: 1px solid #FFCCC7; "
                "border-radius: 4px; padding: 4px 12px; font-size: 12px; "
                "color: #FF4D4F; background: #FFF; } "
                "QPushButton:hover { background: #FFF2F0; }")
            clear_btn.clicked.connect(self.clear_all)
            btn_row.addWidget(clear_btn)

        if self._show_apply_btn:
            apply_btn = QPushButton(self.apply_btn_text if hasattr(self, 'apply_btn_text') else "筛选")
            apply_btn.setFixedHeight(28)
            apply_btn.setStyleSheet(
                "QPushButton { background: #1890FF; color: #FFF; "
                "border: none; border-radius: 4px; padding: 4px 16px; "
                "font-size: 13px; font-weight: bold; } "
                "QPushButton:hover { background: #40A9FF; }")
            apply_btn.clicked.connect(self._on_apply_clicked)
            btn_row.addWidget(apply_btn)

        panel_layout.addLayout(btn_row)
        main_layout.addWidget(self._panel_frame)

    def _build_embedded_ui(self):
        """嵌入模式：仅有条件行区域（无折叠/标题/按钮）。"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(8)

        # 条件行区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(
            "QScrollArea { border: none; background: transparent; }")
        rows_widget = QWidget()
        rows_widget.setStyleSheet("background: transparent;")
        rows_widget.setLayout(self._rows_layout)
        scroll.setWidget(rows_widget)
        main_layout.addWidget(scroll, stretch=1)

        # 底部按钮栏（精简版）
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_row.addStretch()

        if self._show_add_btn:
            add_btn = QPushButton(self._add_btn_text)
            add_btn.setFixedHeight(28)
            add_btn.setStyleSheet(
                "QPushButton { border: 1px dashed #1890FF; "
                "border-radius: 4px; padding: 4px 12px; font-size: 12px; "
                "color: #1890FF; background: #FFF; } "
                "QPushButton:hover { background: #E6F7FF; }")
            add_btn.clicked.connect(lambda: self.add_row())
            btn_row.addWidget(add_btn)

        if self._show_clear_btn:
            clear_btn = QPushButton(self.clear_btn_text if hasattr(self, 'clear_btn_text') else "清除")
            clear_btn.setFixedHeight(28)
            clear_btn.clicked.connect(self.clear_all)
            btn_row.addWidget(clear_btn)

        main_layout.addLayout(btn_row)

        # 外露标签栏
        if self._show_exposed_tags:
            self._exposed_tags_bar = ExposedTagsBar(self)
            self._exposed_tags_bar.tagRemoved.connect(
                self._on_exposed_tag_removed)
            main_layout.addWidget(self._exposed_tags_bar)
        else:
            self._exposed_tags_bar = None

    def _build_popup_ui(self):
        """弹窗模式：构建完整的筛选弹窗 UI。"""
        # 不在此处创建 Dialog，而是由外部调用 show_popup() 时创建
        # 此处构建的是嵌入到 Dialog 中的内容
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(18, 14, 18, 14)
        main_layout.setSpacing(10)

        if self._show_title:
            title_row = QHBoxLayout()
            title_lbl = QLabel(self._title)
            title_lbl.setStyleSheet(
                "font-size: 16px; font-weight: bold;")
            title_row.addWidget(title_lbl)
            title_row.addStretch()
            main_layout.addLayout(title_row)

        # 条件行容器
        self._rows_layout.setSpacing(8)
        main_layout.addLayout(self._rows_layout)

        if self._show_add_btn:
            add_btn = QPushButton(self._add_btn_text)
            add_btn.setFixedWidth(110)
            add_btn.clicked.connect(lambda: self.add_row())
            main_layout.addWidget(add_btn)

        main_layout.addStretch()

        # 按钮栏
        btn_row = QHBoxLayout()
        if self._show_save_btn:
            save_btn = QPushButton(self.save_btn_text if hasattr(self, 'save_btn_text') else "另存为")
            save_btn.setFixedWidth(90)
            save_btn.clicked.connect(
                lambda: self._on_save_preset and self._on_save_preset(self))
            btn_row.addWidget(save_btn)
        btn_row.addStretch()

        if self._show_apply_btn:
            apply_btn = QPushButton(self.apply_btn_text if hasattr(self, 'apply_btn_text') else "筛选")
            apply_btn.setFixedWidth(90)
            apply_btn.clicked.connect(self._on_apply_clicked)
            btn_row.addWidget(apply_btn)

        if self._show_clear_btn:
            clear_btn = QPushButton(self.clear_btn_text if hasattr(self, 'clear_btn_text') else "清除筛选值")
            clear_btn.setFixedWidth(110)
            clear_btn.clicked.connect(self.clear_all)
            btn_row.addWidget(clear_btn)

        main_layout.addLayout(btn_row)

    # ──────────────────────────────────────────────────────────
    # 行管理
    # ──────────────────────────────────────────────────────────

    def add_row(self, condition: dict = None) -> FilterConditionRow:
        """添加一行筛选条件。

        Args:
            condition: 可选的条件字典，用于初始化该行。

        Returns:
            FilterConditionRow 实例。
        """
        # 合并配置：共用默认 + 调用时传入的 condition
        row = FilterConditionRow(
            self,
            field_options=self._row_defaults.get('field_options', []),
            field_placeholder=self._row_defaults.get(
                'field_placeholder', '选择字段'),
            field_width=self._row_defaults.get('field_width', 140),
            text_operators=self._row_defaults.get('text_operators'),
            date_operators=self._row_defaults.get('date_operators'),
            op_width=self._row_defaults.get('op_width', 95),
            default_operator=self._row_defaults.get(
                'default_operator', 'contains'),
            value_pages=self._row_defaults.get(
                'value_pages', ('text', 'date', 'date_range', 'spin')),
            show_expose=self._row_defaults.get('show_expose', True),
            show_picker=self._row_defaults.get('show_picker', False),
            show_remove=self._row_defaults.get('show_remove', True),
            expose_label=self._row_defaults.get('expose_label', '外露'),
            debounce_ms=self._row_defaults.get('debounce_ms', 250),
            is_date_field_cb=self._row_defaults.get('is_date_field_cb'),
            picker_cb=self._row_defaults.get('picker_cb'),
            row_height=self._row_defaults.get('row_height', 32),
        )

        # 连接信号
        row.filtersChanged.connect(self._on_row_changed)
        row.rowRemoveRequested.connect(lambda: self.remove_row(row))

        # 初始化条件
        if condition:
            row.set_condition(condition)
        else:
            row.update_input_mode()

        # 添加到布局和列表
        self._rows_layout.addWidget(row)
        self._rows.append(row)
        self._legacy_rows.append(row)

        # 回调
        if self._on_add_row_cb:
            self._on_add_row_cb(row)

        self._update_ui_state()
        self._adjust_panel_height()
        return row

    def remove_row(self, row: FilterConditionRow):
        """移除指定行。"""
        if row not in self._rows:
            return
        self._rows.remove(row)
        if row in self._legacy_rows:
            self._legacy_rows.remove(row)
        self._rows_layout.removeWidget(row)
        row.hide()
        row.setParent(None)
        row.deleteLater()

        # 确保至少保留一行
        if not self._rows:
            self.add_row()

        self._update_ui_state()
        self._adjust_panel_height()
        self.filtersChanged.emit()

    def clear_all(self):
        """清除所有条件行，恢复为一行空行。"""
        for row in list(self._rows):
            self._rows_layout.removeWidget(row)
            row.hide()
            row.setParent(None)
            row.deleteLater()
        self._rows.clear()
        self._legacy_rows.clear()

        # 添加一行空行
        self.add_row()
        self._update_ui_state()
        self._adjust_panel_height()
        self.filtersChanged.emit()

    # ──────────────────────────────────────────────────────────
    # 值存取
    # ──────────────────────────────────────────────────────────

    def get_all_conditions(self) -> list[dict]:
        """获取所有有效行的条件字典列表（过滤空字段）。"""
        conditions = []
        for row in self._rows:
            cond = row.get_condition()
            if not cond.get('field', '').strip():
                continue
            # 非 empty/not_empty 操作符需要非空值
            if (cond['operator'] not in _BLANK_OPS
                    and not cond.get('value', '')):
                continue
            conditions.append(cond)
        return conditions

    def load_conditions(self, conditions: list[dict]):
        """从条件列表恢复到面板。

        清除现有行，然后根据 conditions 逐行创建。
        如果 conditions 为空或 None，则仅保留一行空行。
        """
        self.clear_all()
        # clear_all 已经添加了一行空行，先移除
        if self._rows:
            self.remove_row(self._rows[0])

        if conditions:
            for c in conditions:
                self.add_row(c)
        else:
            self.add_row()

    # ──────────────────────────────────────────────────────────
    # UI 状态管理
    # ──────────────────────────────────────────────────────────

    def toggle(self):
        """切换面板可见性（inline 模式）。"""
        if self._mode != "inline" or not hasattr(self, '_panel_frame'):
            return
        visible = not self._panel_frame.isVisible()
        self._panel_frame.setVisible(visible)
        self.panelToggled.emit(visible)

    def is_panel_visible(self) -> bool:
        """面板当前是否可见。"""
        if hasattr(self, '_panel_frame'):
            return self._panel_frame.isVisible()
        return True

    def set_panel_visible(self, visible: bool):
        """设置面板可见性。"""
        if hasattr(self, '_panel_frame'):
            self._panel_frame.setVisible(visible)
            self.panelToggled.emit(visible)

    def _update_ui_state(self):
        """更新 UI 状态（按钮徽标等）。"""
        self._update_toggle_badge()

    def _adjust_panel_height(self):
        """根据条件行数量动态调整面板高度。"""
        if not hasattr(self, '_panel_frame') or not self._rows:
            return
        row_h = 32 + 6  # 行高 + 间距
        header_h = 40 if self._show_title else 0
        footer_h = 56  # 底部按钮栏
        ideal = header_h + max(len(self._rows), 1) * row_h + footer_h + 20
        ideal = min(ideal, 500)  # 上限
        ideal = max(ideal, 140)  # 下限
        self._panel_frame.setFixedHeight(ideal)

    def _update_toggle_badge(self):
        """更新切换按钮上的条件计数徽标。"""
        if not self._toggle_btn or not self._toggle_badge:
            return
        count = len([r for r in self._rows
                     if r.get_condition().get('field', '').strip()])
        if count > 0:
            self._toggle_btn.setText(f"筛选({count})")
        else:
            self._toggle_btn.setText("筛选")

    def _on_row_changed(self):
        """任意行值变化时转发信号。"""
        self._update_ui_state()
        if self._show_exposed_tags and hasattr(self, '_exposed_tags_bar'):
            self._exposed_tags_bar.refresh(self.get_all_conditions())
        self.filtersChanged.emit()

    def _on_apply_clicked(self):
        """点击「筛选」按钮。"""
        if self._on_apply:
            self._on_apply(self)
        else:
            self.filtersChanged.emit()
            if self._mode == "inline":
                self.set_panel_visible(False)

    def _on_exposed_tag_removed(self, cond: dict):
        """外露标签被移除时，取消对应行的外露勾选。"""
        for row in self._rows:
            c = row.get_condition()
            if (c.get('field') == cond.get('field')
                    and c.get('operator') == cond.get('operator')
                    and c.get('value') == cond.get('value')):
                if row.expose_check:
                    row.expose_check.setChecked(False)
                break
        self.refresh_exposed_tags()

    # ──────────────────────────────────────────────────────────
    # 外露标签
    # ──────────────────────────────────────────────────────────

    def refresh_exposed_tags(self):
        """刷新外露标签栏。"""
        if not self._show_exposed_tags or not hasattr(
                self, '_exposed_tags_bar'):
            return
        self._exposed_tags_bar.refresh(self.get_all_conditions())

    # ──────────────────────────────────────────────────────────
    # 批量操作
    # ──────────────────────────────────────────────────────────

    def set_available_fields(self, options: list):
        """批量更新所有行的可选字段列表。"""
        self._row_defaults['field_options'] = options
        for row in self._rows:
            row.set_field_options(options)

    def set_is_date_field_callback(self, cb: callable):
        """批量设置所有行的日期字段判断回调。"""
        self._row_defaults['is_date_field_cb'] = cb
        for row in self._rows:
            row.set_is_date_field_callback(cb)

    def set_picker_callback(self, cb: callable):
        """批量设置所有行的多选回调。"""
        self._row_defaults['picker_cb'] = cb
        for row in self._rows:
            row.set_picker_callback(cb)

    def for_each_row(self, callback: callable):
        """遍历所有行，对每行调用 callback(row)。"""
        for row in self._rows:
            callback(row)

    # ──────────────────────────────────────────────────────────
    # 弹窗模式辅助
    # ──────────────────────────────────────────────────────────

    def show_popup(self, parent=None, conditions: list[dict] = None,
                   on_accepted: callable = None):
        """以弹窗形式显示筛选面板。

        Args:
            parent: 父窗口
            conditions: 初始条件列表
            on_accepted: 用户确认后的回调，接收条件列表
        """
        from PyQt6.QtWidgets import QDialog

        dlg = CenteredPopupDialog(parent, close_on_outside=True)
        dlg.setWindowTitle(self._title)
        dlg.resize(780, 420)

        # 将当前面板的内容嵌入弹窗
        dlg_layout = QVBoxLayout(dlg)
        dlg_layout.setContentsMargins(0, 0, 0, 0)
        dlg_layout.addWidget(self)

        if conditions is not None:
            self.load_conditions(conditions)

        def _on_dlg_accepted():
            conditions = self.get_all_conditions()
            if on_accepted:
                on_accepted(conditions)
            dlg._dialog_closed = True
            dlg.accept()

        dlg.accepted.connect(_on_dlg_accepted)

        self._popup_dlg = dlg
        dlg.exec()

    def for_each_row_info(self):
        """兼容层：生成与旧 row_info dict 兼容的迭代器。"""
        for row in self._rows:
            yield {
                'frame': row,
                'field_combo': row.field_combo,
                'op_combo': row.op_combo,
                'value_stack': row.value_stack,
                'text_input': row.text_input,
                'value_input': row.text_input,
                'date_input': row.date_input,
                'date_range_btn': row.date_range_btn,
                'past_days_spin': row.spin_input,
                'n_input': row.spin_input,
                'spin_input': row.spin_input,
                'picker_btn': row.picker_btn,
                'expose_check': row.expose_check,
                'row': row,
            }


class CustomReportFilterDialog(CenteredPopupDialog):
    """自定义报表多条件筛选弹窗 — 基于 FilterPanel 重构。

    保持与旧版完全兼容的公开 API：
        __init__(fields, conditions=None, parent=None, date_fields=None)
        get_conditions() -> list[dict]
        exec() -> DialogCode (继承自 QDialog)

    内部使用 FilterPanel 管理条件行，不再手动构建 UI。
    """

    OPERATORS = [
        ("包含", "contains"),
        ("不包含", "not_contains"),
        ("属于", "in"),
        ("不属于", "not_in"),
        ("等于", "eq"),
        ("不等于", "ne"),
        ("为空（未填写）", "empty"),
        ("不为空", "not_empty"),
        ("开始于", "starts_with"),
        ("结束于", "ends_with"),
        ("大于", "gt"),
        ("小于", "lt"),
    ]

    def __init__(self, fields, conditions=None, parent=None, date_fields=None):
        """初始化筛选弹窗。

        Args:
            fields: 可选字段列表（字符串列表）
            conditions: 初始筛选条件列表
            parent: 父窗口
            date_fields: 日期字段集合（字符串列表或 set）
        """
        super().__init__(parent, close_on_outside=True)
        self.setWindowTitle("设置筛选")
        self.resize(780, 420)
        self._fields = fields or []
        self._date_fields = {str(f).strip() for f in (date_fields or [])
                             if str(f).strip()}

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(18, 14, 18, 14)
        root_layout.setSpacing(10)

        # 标题行
        title_row = QHBoxLayout()
        title_lbl = QLabel("设置筛选")
        title_lbl.setStyleSheet("font-size: 16px; font-weight: bold;")
        title_row.addWidget(title_lbl)
        title_row.addStretch()
        close_btn = QPushButton("×")
        close_btn.setFixedSize(28, 28)
        close_btn.clicked.connect(self.reject)
        title_row.addWidget(close_btn)
        root_layout.addLayout(title_row)

        # 使用 FilterPanel 管理条件行
        field_options = [(f, f) for f in self._fields]
        self._panel = FilterPanel(
            self,
            mode="popup",
            title="",
            show_title=False,
            show_add_btn=True,
            show_apply_btn=True,
            show_clear_btn=True,
            show_save_btn=True,
            show_exposed_tags=False,
            add_btn_text="+ 添加条件",
            apply_btn_text="筛选",
            clear_btn_text="清除筛选值",
            save_btn_text="另存为",
            row_defaults={
                'field_options': field_options,
                'field_width': 180,
                'text_operators': self.OPERATORS,
                'date_operators': self.OPERATORS,
                'op_width': 120,
                'default_operator': 'contains',
                'value_pages': ('text', 'date'),
                'show_expose': False,
                'show_picker': True,
                'show_remove': True,
                'debounce_ms': 0,
                'is_date_field_cb': self._is_date_field,
                'picker_cb': self._on_picker_clicked,
            },
            on_apply=lambda panel: self.accept(),
        )

        # 将 FilterPanel 的 rows_layout 注入到弹窗布局中
        # FilterPanel 在 popup 模式下会创建完整的内部布局，
        # 我们需要的是它的条件行区域 + 按钮栏
        # 直接把它添加到 root_layout（它会重复标题行，但我们已隐藏了）
        root_layout.addWidget(self._panel)

        # 覆盖 FilterPanel 的 on_apply 使其调用弹窗的 accept()
        self._panel._on_apply = lambda panel: self.accept()

        # 加载初始条件
        display_conditions = conditions or []
        if display_conditions:
            self._panel.load_conditions(display_conditions)
        else:
            self._panel.add_row()

    # ── 辅助方法 ──

    def _is_date_field(self, field_name):
        """判断字段是否应使用日期输入。"""
        normalized = str(field_name or '').strip()
        if not normalized:
            return False
        if normalized in self._date_fields:
            return True
        lower_name = normalized.lower()
        return any(k in lower_name
                   for k in ('日期', '时间', 'date', 'time'))

    def _on_picker_clicked(self, row):
        """多选按钮点击 — 委托给父窗口获取候选值。"""
        field_display = row.field_combo.currentText().strip()
        if not field_display:
            return

        parent = self.parent()
        if not parent or not hasattr(parent, '_get_custom_report_field_value_options'):
            return

        field_key = parent._get_custom_report_field_key(field_display)
        mappings = parent._get_custom_report_field_value_options(
            field_key, field_display)

        if not mappings:
            show_info('提示',
                      f'字段「{field_display}」暂无候选值。\n'
                      f'请先刷新数据，或手动输入筛选值。',
                      self)
            return

        current_value = row.text_input.text().strip()
        show_multi_select_dropdown(self, mappings, current_value,
                                   anchor=row.text_input,
                                   target_input=row.text_input)

    # ── 公开 API（与旧版完全兼容） ──

    def get_conditions(self):
        """获取筛选条件列表。

        Returns:
            list[dict]: 每个元素包含 field, operator, value 三个键。
        """
        conditions = []
        for row in self._panel._rows:
            cond = row.get_condition()
            if not cond.get('field', '').strip():
                continue
            operator = cond.get('operator', '')
            value = cond.get('value', '')
            if operator not in ('empty', 'not_empty') and not value:
                continue
            conditions.append({
                'field': cond['field'],
                'operator': operator,
                'value': value,
            })
        return conditions

    # 保持向后兼容的属性访问
    @property
    def condition_rows(self):
        """兼容旧代码：返回 row_info 风格的 list。"""
        return list(self._panel.for_each_row_info())

class PasswordEntry(QFrame):
    """密码输入框组件"""
    returnPressed = pyqtSignal()

    def __init__(self, parent=None):
        """初始化密码输入框组件。"""
        super().__init__(parent)
        target_height = get_standard_line_edit_height()
        self.setFixedHeight(target_height)
        self.setStyleSheet("""
            QFrame {
            border: 1px solid #7a7a7a;

            border-radius: 4px;

            background-color: #ffffff;

            }
        """)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(6, 0, 3, 0)
        layout.setSpacing(2)

        self.entry = QLineEdit()
        self.entry.setEchoMode(QLineEdit.Password)
        self.entry.setFixedHeight(max(18, target_height - 4))
        self.entry.setStyleSheet("""
            QLineEdit {
            border: none;

            background: transparent;

            padding: 0px 2px;

            }
        """)
        self.entry.returnPressed.connect(self.returnPressed.emit)
        layout.addWidget(self.entry)

        self.show_btn = QToolButton()
        self.show_btn.setCheckable(True)
        self.show_btn.setChecked(False)
        self.show_btn.setText("👁️")
        self.show_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.show_btn.setFixedSize(max(18, target_height - 6), max(18, target_height - 6))
        self.show_btn.setStyleSheet("""
            QToolButton {
            border: none;

            background: transparent;

            padding: 0px;

            min-width: 20px;

            }
        """)
        self.show_btn.clicked.connect(self.toggle_password)
        layout.addWidget(self.show_btn)

        self.is_shown = False

    def toggle_password(self):
        """切换密码显示/隐藏"""
        self.is_shown = not self.is_shown
        if self.is_shown:
            self.entry.setEchoMode(QLineEdit.Normal)
            self.show_btn.setText("🙈")
        else:
            self.entry.setEchoMode(QLineEdit.Password)
            self.show_btn.setText("👁️")

    def get(self):
        """获取密码"""
        return self.entry.text()

    def setText(self, text):
        """设置密码"""
        self.entry.setText(text)

    def clear(self):
        """清空密码"""
        self.entry.clear()

from file_mover import *
from file_generation import *
from pdf_watermark import *
# Mixin imports are handled by the main file
from bi_dashboard import *
from department import *

