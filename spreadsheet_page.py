# -*- coding: utf-8 -*-
from core import *
from common import *
from common import frameless_input_text, frameless_message_box

"""
spreadsheet_page.py — 电子表格 Mixin
──────────────────────────────────────
负责：MainFrame 中电子表格相关方法
  - create_spreadsheet_page()    创建电子表格页面（页面9，Luckysheet）
  - spreadsheet_load_data()      加载工作簿数据（Python → JS）
  - spreadsheet_load_dataframe() 从 DataFrame 加载（自动转 cellData）
  - spreadsheet_read_data()      读回所有数据（JS → Python）
  - 导入/导出 XLSX
依赖：core.py / common.py / openpyxl
被导入：主程序（作为 MainFrame 的 Mixin 父类）
"""

import os
import json
import logging
import socketserver
import threading
from functools import partial
from http.server import SimpleHTTPRequestHandler
from pathlib import Path

from PyQt6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QFrame, QPushButton,
    QLabel, QFileDialog, QMessageBox, QDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QTextEdit,
    QStackedWidget, QWidget,
)
from PyQt6.QtCore import Qt, QUrl, QTimer, QThread, pyqtSignal, pyqtSlot, QObject
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWebChannel import QWebChannel
from PyQt6.QtGui import QColor

logger = logging.getLogger(__name__)


# ============================================================
# _QuietHandler / _start_local_http — 本地 HTTP 服务器
# 解决 setHtml() 下 ES Module (type="module") 无法加载外部 CDN 的问题
# ============================================================

class _QuietHandler(SimpleHTTPRequestHandler):
    """静默 HTTP 请求处理器（不输出日志到控制台）"""
    def log_message(self, format, *args):
        pass


_server_port = None


def _start_local_http(directory: str) -> int:
    """
    在 directory 上启动一个本地 HTTP 服务器（仅绑定 127.0.0.1）。
    返回端口号。服务器在 daemon 线程中运行，主进程退出时自动关闭。
    """
    global _server_port
    if _server_port is not None:
        return _server_port

    handler = partial(_QuietHandler, directory=directory)

    class _TCPServer(socketserver.TCPServer):
        allow_reuse_address = True

    for port in range(18900, 19000):
        try:
            httpd = _TCPServer(("127.0.0.1", port), handler)
            _server_port = port
            t = threading.Thread(target=httpd.serve_forever, daemon=True)
            t.start()
            logger.info(f"[Spreadsheet] Local HTTP server started: http://127.0.0.1:{port}")
            return port
        except OSError:
            continue

    raise RuntimeError("无法启动本地 HTTP 服务器（18900-18999 端口均被占用）")


# ============================================================
# SpreadsheetBridge — Python ↔ JS 通信桥
# 复用 DashboardBridge 模式 (custom_report.py:6251)
# ============================================================

# ============================================================
# _ImportWorker — 后台导入 XLSX 工作线程
# ============================================================

class _ImportWorker(QThread):
    """后台线程：加载 XLSX 并转换为 Univer cellData，避免 UI 卡死"""
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, file_path: str, parent=None):
        super().__init__(parent)
        self._file_path = file_path

    def run(self):
        try:
            import openpyxl
            self.progress.emit(f"正在读取: {os.path.basename(self._file_path)} ...")

            # read_only=True 流式读取，避免 MemoryError（不会为空单元格创建对象）
            wb_formulas = openpyxl.load_workbook(self._file_path, data_only=False, read_only=True)
            try:
                wb_values = openpyxl.load_workbook(self._file_path, data_only=True, read_only=True)
            except Exception:
                wb_values = None

            sheets_data = {}
            total_sheets = len(wb_formulas.sheetnames)
            for s_idx, sheet_name in enumerate(wb_formulas.sheetnames):
                self.progress.emit(f"解析 Sheet {s_idx + 1}/{total_sheets}: {sheet_name}")
                ws_f = wb_formulas[sheet_name]
                ws_v = wb_values[sheet_name] if wb_values else None
                cell_data = {}

                # 构建缓存值字典（read_only 模式不支持随机访问 cell()）
                cached_values = {}
                if ws_v:
                    try:
                        for v_row in ws_v.iter_rows():
                            for v_cell in v_row:
                                if v_cell.value is not None:
                                    cached_values[(v_cell.row, v_cell.column)] = v_cell.value
                    except Exception:
                        pass

                for row_idx, row in enumerate(ws_f.iter_rows()):
                    row_key = str(row_idx)
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        raw = cell.value
                        if raw is None:
                            continue
                        cell_info = {}
                        if isinstance(raw, str) and raw.startswith('='):
                            cell_info["f"] = raw
                            cached = cached_values.get((cell.row, cell.column))
                            if cached is not None:
                                cell_info["v"] = cached
                        else:
                            cell_info["v"] = raw
                        if cell_info:
                            row_data[str(col_idx)] = cell_info
                    if row_data:
                        cell_data[row_key] = row_data
                sheets_data[sheet_name] = {
                    "name": sheet_name,
                    "cellData": cell_data,
                }

                # 释放该 sheet 的缓存值
                cached_values.clear()
            first_sheet = wb_formulas.sheetnames[0]
            wb_formulas.close()
            if wb_values:
                wb_values.close()
            self.finished.emit({
                "sheets": sheets_data,
                "activeSheet": first_sheet,
            })
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.error(f"[ImportWorker] XLSX import failed: {e}\n{tb}")
            self.error.emit(str(e) or repr(e) or "未知错误（详见日志）")


# ============================================================
# _MysqlQueryWorker — 后台执行 MySQL 查询并转为 cellData
# ============================================================

class _MysqlQueryWorker(QThread):
    """后台线程：执行 MySQL 查询 → DataFrame → cellData"""
    finished = pyqtSignal(dict, str)   # (workbook_data, sheet_name)
    error = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, mysql_config: dict, sql: str, sheet_name: str, parent=None):
        super().__init__(parent)
        self._cfg = mysql_config
        self._sql = sql
        self._sheet_name = sheet_name

    def run(self):
        try:
            import pymysql
            import pandas as pd
            self.progress.emit("正在连接 MySQL ...")
            conn = pymysql.connect(
                host=self._cfg.get('host', '127.0.0.1'),
                port=int(self._cfg.get('port', 3306)),
                user=self._cfg.get('user', 'root'),
                password=self._cfg.get('password', ''),
                database=self._cfg.get('database', ''),
                charset='utf8mb4',
                connect_timeout=10,
                cursorclass=pymysql.cursors.DictCursor,
            )
            try:
                self.progress.emit("正在查询数据 ...")
                with conn.cursor() as cur:
                    cur.execute(self._sql)
                    rows = cur.fetchall()
                df = pd.DataFrame(rows)
            finally:
                conn.close()

            self.progress.emit(f"正在转换 {len(df)} 行数据 ...")
            cell_data = _dataframe_to_cell_data_static(df)
            workbook_data = {
                "sheets": {
                    self._sheet_name: {
                        "name": self._sheet_name,
                        "cellData": cell_data,
                    }
                },
                "activeSheet": self._sheet_name,
            }
            self.finished.emit(workbook_data, self._sheet_name)
        except Exception as e:
            logger.error(f"[MysqlQueryWorker] Query failed: {e}")
            self.error.emit(str(e))


# ============================================================
# _MysqlListTablesWorker — 后台获取 MySQL 表列表
# ============================================================

class _MysqlListTablesWorker(QThread):
    """后台线程：获取 MySQL 表列表及行数"""
    finished = pyqtSignal(list)  # [(table_name, row_count), ...]
    error = pyqtSignal(str)

    def __init__(self, mysql_config: dict, parent=None):
        super().__init__(parent)
        self._cfg = mysql_config

    def run(self):
        try:
            import pymysql
            conn = pymysql.connect(
                host=self._cfg.get('host', '127.0.0.1'),
                port=int(self._cfg.get('port', 3306)),
                user=self._cfg.get('user', 'root'),
                password=self._cfg.get('password', ''),
                database=self._cfg.get('database', ''),
                charset='utf8mb4',
                connect_timeout=10,
                cursorclass=pymysql.cursors.DictCursor,
            )
            try:
                with conn.cursor() as cur:
                    cur.execute("SHOW TABLES")
                    tables = [list(row.values())[0] for row in cur.fetchall()]
                result = []
                for t in tables:
                    try:
                        with conn.cursor() as cur:
                            cur.execute(f"SELECT COUNT(*) AS cnt FROM `{t}`")
                            cnt = cur.fetchone()['cnt']
                        result.append((t, cnt))
                    except Exception:
                        result.append((t, -1))
            finally:
                conn.close()
            self.finished.emit(result)
        except Exception as e:
            logger.error(f"[MysqlListTablesWorker] Failed: {e}")
            self.error.emit(str(e))


# ============================================================
# _dataframe_to_cell_data_static — 线程安全的 DataFrame → cellData
# ============================================================

def _dataframe_to_cell_data_static(df) -> dict:
    """将 DataFrame 转为 Univer cellData（静态函数，可在子线程调用）"""
    import numpy as np
    from datetime import datetime, date, time
    from decimal import Decimal

    def _to_json_safe(val):
        """将任意 Python/MySQL/numpy 类型转为 JSON 可序列化的原生类型"""
        if val is None:
            return None
        if isinstance(val, bool):
            return val
        if isinstance(val, (int, float, str)):
            return val
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            f = float(val)
            return None if np.isnan(f) or np.isinf(f) else f
        if isinstance(val, np.bool_):
            return bool(val)
        if isinstance(val, np.ndarray):
            return val.tolist()
        # datetime / date / time → 字符串
        if isinstance(val, (datetime, date, time)):
            return str(val)
        # Decimal（MySQL 精确数值）→ float
        if isinstance(val, Decimal):
            try:
                return float(val)
            except (ValueError, OverflowError):
                return str(val)
        # bytes → 字符串
        if isinstance(val, (bytes, bytearray)):
            try:
                return val.decode('utf-8')
            except UnicodeDecodeError:
                return val.hex()
        # numpy datetime64 / timedelta64
        if hasattr(val, 'item'):
            try:
                return _to_json_safe(val.item())
            except Exception:
                return str(val)
        # 其它类型 → 字符串兜底
        return str(val)

    cell_data = {}
    # 表头行 (row 0)
    header = {}
    for col_idx, col_name in enumerate(df.columns):
        header[str(col_idx)] = {"v": str(col_name)}
    cell_data["0"] = header

    # 数据行
    try:
        values = df.values
        for row_idx in range(len(values)):
            row_key = str(row_idx + 1)
            row_data = {}
            for col_idx in range(values.shape[1]):
                val = values[row_idx, col_idx]
                safe_val = _to_json_safe(val)
                if safe_val is not None:
                    row_data[str(col_idx)] = {"v": safe_val}
            if row_data:
                cell_data[row_key] = row_data
    except Exception:
        import pandas as pd
        for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
            row_key = str(row_idx)
            row_data = {}
            for col_idx, value in enumerate(row):
                if pd.notna(value):
                    safe_val = _to_json_safe(value)
                    if safe_val is not None:
                        row_data[str(col_idx)] = {"v": safe_val}
            cell_data[row_key] = row_data

    return cell_data


# ============================================================
# Univer ↔ Luckysheet 数据格式转换
# ============================================================

def _univer_to_luckysheet(workbook_data: dict) -> list:
    """将 Univer 格式的 workbookData 转为 Luckysheet 格式的 sheets 数组。

    Univer:  {"sheets": {"id": {"name":, "cellData": {"0": {"0": {"v":, "f":}}}}}}
    Luckysheet: [{"name":, "celldata": [{"r":, "c":, "v": {"v":, "f":, "m":}}], ...}]
    """
    sheets = []
    for sheet_id, info in workbook_data.get("sheets", {}).items():
        celldata = []
        max_row = 0
        max_col = 0
        for rk, row in info.get("cellData", {}).items():
            ri = int(rk)
            if ri > max_row:
                max_row = ri
            for ck, cell in row.items():
                ci = int(ck)
                if ci > max_col:
                    max_col = ci
                v = cell.get("v")
                v_obj = {"v": v, "m": str(v) if v is not None else ""}
                if "f" in cell:
                    v_obj["f"] = cell["f"]
                    v_obj["m"] = str(v) if v is not None else ""
                celldata.append({"r": ri, "c": ci, "v": v_obj})
        sheets.append({
            "name": info.get("name", str(sheet_id)),
            "celldata": celldata,
            "row": max(max_row + 2, 100),
            "column": max(max_col + 2, 20),
            "status": 1 if len(sheets) == 0 else 0,
        })
    return sheets


def _luckysheet_to_univer(sheets: list) -> dict:
    """将 Luckysheet 格式的 sheets 数组转回 Univer 格式的 workbookData。

    用于 cell_data_response 回调，使 _do_auto_save / _do_export_xlsx
    等依赖 Univer 格式的方法无需修改。
    """
    result = {}
    first_name = ""
    for sheet in sheets:
        name = sheet.get("name", "Sheet")
        if not first_name:
            first_name = name
        cell_data = {}
        for cell in sheet.get("celldata", []):
            rk = str(cell.get("r", 0))
            ck = str(cell.get("c", 0))
            cell_data.setdefault(rk, {})
            v = cell.get("v", {})
            cell_info = {}
            if v.get("v") is not None:
                cell_info["v"] = v["v"]
            if v.get("f"):
                cell_info["f"] = v["f"]
            if cell_info:
                cell_data[rk][ck] = cell_info
        result[name] = {"name": name, "cellData": cell_data}
    return {"sheets": result, "activeSheet": first_name}


# ============================================================
# MysqlImportDialog — MySQL 数据导入对话框
# ============================================================

class MysqlImportDialog(QDialog):
    """从 MySQL 选择表/自定义 SQL 导入数据到电子表格"""

    def __init__(self, mysql_config: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🗄️ 从 MySQL 导入数据")
        self.setMinimumSize(720, 520)
        self._cfg = mysql_config
        self._tables = []
        self._selected_sql = None
        self._selected_sheet = None

        self._init_ui()
        self._load_tables()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        info = QLabel(
            f"连接: {self._cfg.get('host', '?')}:{self._cfg.get('port', 3306)} "
            f"/ {self._cfg.get('database', '?')}"
        )
        info.setStyleSheet("color: #666; font-size: 12px;")
        layout.addWidget(info)

        top_label = QLabel("📋 选择数据表（双击快速导入）")
        top_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        layout.addWidget(top_label)

        self._table_widget = QTableWidget()
        self._table_widget.setColumnCount(2)
        self._table_widget.setHorizontalHeaderLabels(["表名", "行数"])
        install_autofilter_header(self._table_widget)
        self._table_widget.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        self._table_widget.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents
        )
        self._table_widget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table_widget.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table_widget.doubleClicked.connect(self._on_table_double_click)
        self._table_widget.itemSelectionChanged.connect(self._on_table_selected)
        layout.addWidget(self._table_widget, 1)

        sql_label = QLabel("✏️ 或输入自定义 SQL 查询")
        sql_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        layout.addWidget(sql_label)

        self._sql_edit = QTextEdit()
        self._sql_edit.setPlaceholderText("SELECT * FROM table_name LIMIT 10000")
        self._sql_edit.setMaximumHeight(80)
        self._sql_edit.setStyleSheet("font-family: Consolas, monospace; font-size: 12px;")
        layout.addWidget(self._sql_edit)

        self._status_label = QLabel("")
        self._status_label.setStyleSheet("color: #999; font-size: 11px;")
        layout.addWidget(self._status_label)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        self._btn_refresh = QPushButton("🔄 刷新表列表")
        self._btn_refresh.clicked.connect(self._load_tables)
        btn_layout.addWidget(self._btn_refresh)

        self._btn_import_sql = QPushButton("📥 执行SQL并导入")
        self._btn_import_sql.clicked.connect(self._on_import_sql)
        btn_layout.addWidget(self._btn_import_sql)

        self._btn_cancel = QPushButton("取消")
        self._btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(self._btn_cancel)

        layout.addLayout(btn_layout)

    def _load_tables(self):
        self._status_label.setText("正在加载表列表 ...")
        self._btn_refresh.setEnabled(False)
        self._worker = _MysqlListTablesWorker(self._cfg, self)
        self._worker.finished.connect(self._on_tables_loaded)
        self._worker.error.connect(self._on_tables_error)
        self._worker.start()

    def _on_tables_loaded(self, tables: list):
        self._btn_refresh.setEnabled(True)
        self._tables = tables
        self._table_widget.setRowCount(len(tables))
        for i, (name, cnt) in enumerate(tables):
            self._table_widget.setItem(i, 0, QTableWidgetItem(name))
            cnt_text = f"{cnt:,}" if cnt >= 0 else "N/A"
            self._table_widget.setItem(i, 1, QTableWidgetItem(cnt_text))
        self._status_label.setText(f"共 {len(tables)} 张表")

    def _on_tables_error(self, msg: str):
        self._btn_refresh.setEnabled(True)
        self._status_label.setText(f"❌ 加载失败: {msg}")

    def _on_table_selected(self):
        rows = self._table_widget.selectionModel().selectedRows()
        if rows:
            idx = rows[0].row()
            table_name = self._tables[idx][0]
            self._sql_edit.setPlainText(f"SELECT * FROM `{table_name}` LIMIT 10000")

    def _on_table_double_click(self, index):
        idx = index.row()
        table_name = self._tables[idx][0]
        self._selected_sql = f"SELECT * FROM `{table_name}` LIMIT 10000"
        self._selected_sheet = table_name
        self.accept()

    def _on_import_sql(self):
        sql = self._sql_edit.toPlainText().strip()
        if not sql:
            frameless_message_box(self, "提示", "请输入 SQL 查询语句")
            return
        if not sql.upper().lstrip().startswith('SELECT'):
            frameless_message_box(self, "安全提示", "仅支持 SELECT 查询")
            return
        self._selected_sql = sql
        import re
        m = re.search(r'FROM\s+`?(\w+)`?', sql, re.IGNORECASE)
        self._selected_sheet = m.group(1) if m else "MySQL数据"
        self.accept()

    def get_import_info(self) -> tuple:
        """返回 (sql, sheet_name) 或 (None, None)"""
        return self._selected_sql, self._selected_sheet


class SpreadsheetBridge(QObject):
    """电子表格 Python <-> JS 通信桥"""

    # Python → JS 信号
    actionReceived = pyqtSignal(str, str)  # (action, payload_json)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._page = None

    def set_page(self, page):
        """保存 QWebEnginePage 引用，用于 runJavaScript"""
        self._page = page

    # ===== Python → JS（通过 runJavaScript 直接注入，绕过 Channel）=====

    def load_workbook_data(self, data: dict):
        """加载完整工作簿数据（自动从内部格式转为 Luckysheet 格式）"""
        try:
            luckysheet_data = _univer_to_luckysheet(data)
            payload = json.dumps(luckysheet_data, ensure_ascii=False, default=str)
        except (TypeError, ValueError) as e:
            logger.error(f"[SpreadsheetBridge] JSON serialization failed: {e}")
            return
        js = f"window.loadWorkbookData({payload});"
        self._run_js(js)

    def set_cell_value(self, row: int, col: int, value, sheet_id: str = None):
        """设置单个单元格值（sheet_id 在 Luckysheet 中为 sheet index）"""
        sheet_ref = f"'{sheet_id}'" if sheet_id else "null"
        js = f"window.setCellValue({row}, {col}, {json.dumps(value, ensure_ascii=False)}, {sheet_ref});"
        self._run_js(js)

    def request_cell_data(self):
        """请求 JS 侧返回所有单元格数据（异步，结果通过 handle_js_action('cell_data_response', ...) 回传）"""
        self._run_js(
            "(function(){"
            "  try {"
            "    if(typeof window.getAllCellData !== 'function'){"
            "      if(window.bridge) window.bridge.handle_js_action('cell_data_response', JSON.stringify({error:'getAllCellData 函数不存在'}));"
            "      return;"
            "    }"
            "    var d = window.getAllCellData();"
            "    if(!d || d === 'null' || d === 'undefined'){"
            "      if(window.bridge) window.bridge.handle_js_action('cell_data_response', JSON.stringify({error:'返回数据为空'}));"
            "      return;"
            "    }"
            "    if(window.bridge) window.bridge.handle_js_action('cell_data_response', d);"
            "  } catch(e) {"
            "    if(window.bridge) window.bridge.handle_js_action('cell_data_response', JSON.stringify({error: e.message || String(e)}));"
            "  }"
            "})();"
        )

    def set_editable(self, editable: bool):
        """启用/禁用编辑"""
        js = f"window.setEditable({str(editable).lower()});"
        self._run_js(js)

    def run_js(self, js: str):
        """执行原始 JS"""
        self._run_js(js)

    def _run_js(self, js: str):
        """安全执行 JS"""
        if self._page:
            self._page.runJavaScript(js)

    # ===== JS → Python（通过 QWebChannel）=====

    @pyqtSlot(str, str)
    def handle_js_action(self, action: str, payload_json: str):
        """
        JS 统一回调入口。

        action 类型:
          - "page_ready"           页面加载完成，电子表格初始化完毕
          - "cell_changed"         用户编辑了单元格
          - "cell_data_response"   完整数据回传（来自 request_cell_data）
          - "selection_changed"    选区变化
          - "data_loaded"          数据加载完成通知
          - "formula_error"        公式/执行错误
          - "xlsx_export_ready"    xlsx 导出完成
        """
        logger.debug(f"[SpreadsheetBridge] JS → Python: {action}")
        self.actionReceived.emit(action, payload_json)

    @pyqtSlot(str, result=str)
    def get_initial_data(self) -> str:
        """JS 懒加载请求初始数据（可由外部重写）"""
        return "{}"


# ============================================================
# SpreadsheetPageMixin — 电子表格页面
# 复用 bi_dashboardMixin 模式 (bi_dashboard.py:66)
# ============================================================

class SpreadsheetListPage(QWidget):
    """电子表格列表页 — 显示已保存的表格文件，双击进入编辑器"""

    spreadsheetSelected = pyqtSignal(str)  # 文件路径
    newSpreadsheet = pyqtSignal()

    def __init__(self, storage_dir: str, parent=None):
        super().__init__(parent)
        self._storage_dir = storage_dir
        os.makedirs(storage_dir, exist_ok=True)
        self._files = []
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(8)

        # 工具栏
        toolbar = QHBoxLayout()
        title = QLabel("🧮 电子表格")
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #333;")
        toolbar.addWidget(title)
        toolbar.addStretch()

        btn_new = QPushButton("+ 新建表格")
        btn_new.setFixedHeight(28)
        btn_new.setStyleSheet("background: #1890FF; color: white; border: none; border-radius: 4px; padding: 0 12px;")
        btn_new.clicked.connect(self.newSpreadsheet.emit)
        toolbar.addWidget(btn_new)

        btn_import = QPushButton("📥 导入本地表格")
        btn_import.setFixedHeight(28)
        btn_import.clicked.connect(self._on_import_local)
        toolbar.addWidget(btn_import)

        btn_refresh = QPushButton("🔄 刷新")
        btn_refresh.setFixedHeight(28)
        btn_refresh.clicked.connect(self.populate)
        toolbar.addWidget(btn_refresh)

        btn_delete = QPushButton("🗑️ 删除选中")
        btn_delete.setFixedHeight(28)
        btn_delete.clicked.connect(self._on_delete_selected)
        toolbar.addWidget(btn_delete)

        layout.addLayout(toolbar)

        # 表格
        self._table = QTableWidget()
        self._table.setColumnCount(3)
        self._table.setHorizontalHeaderLabels(["文件名", "修改时间", "大小"])

        # 列宽：Interactive 模式，支持手动拖拽调整
        header = install_autofilter_header(self._table)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(False)
        self._table.verticalHeader().setVisible(False)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.cellDoubleClicked.connect(self._on_double_click)
        layout.addWidget(self._table, 1)

        # 加载已保存的列宽
        self._loading_widths = True
        saved_widths = self._load_column_widths()
        if saved_widths:
            for i, w in enumerate(saved_widths):
                if i < 3:
                    self._table.setColumnWidth(i, w)
        else:
            self._table.setColumnWidth(0, 260)
            self._table.setColumnWidth(1, 150)
            self._table.setColumnWidth(2, 90)
        self._loading_widths = False

        # 列宽变化时自动保存
        header.sectionResized.connect(self._on_column_resized)

        # 底部状态
        self._status = QLabel("")
        self._status.setStyleSheet("color: #999; font-size: 11px;")
        layout.addWidget(self._status)

    def populate(self):
        """扫描存储目录，填充列表"""
        self._files = []
        if os.path.isdir(self._storage_dir):
            for f in sorted(os.listdir(self._storage_dir)):
                if f.lower().endswith('.xlsx') and not f.startswith('~'):
                    full = os.path.join(self._storage_dir, f)
                    self._files.append(full)

        self._table.setRowCount(len(self._files))
        for i, path in enumerate(self._files):
            name = os.path.splitext(os.path.basename(path))[0]
            mtime = os.path.getmtime(path)
            size = os.path.getsize(path)
            from datetime import datetime
            time_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
            if size < 1024:
                size_str = f"{size} B"
            elif size < 1024 * 1024:
                size_str = f"{size / 1024:.1f} KB"
            else:
                size_str = f"{size / 1024 / 1024:.1f} MB"
            self._table.setItem(i, 0, QTableWidgetItem(name))
            self._table.setItem(i, 1, QTableWidgetItem(time_str))
            self._table.setItem(i, 2, QTableWidgetItem(size_str))
        self._status.setText(f"共 {len(self._files)} 个表格")

    def _on_double_click(self, row, col):
        if 0 <= row < len(self._files):
            self.spreadsheetSelected.emit(self._files[row])

    def _on_import_local(self):
        """导入本地 xlsx 文件到表格目录"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if not file_path:
            return
        import shutil
        name = os.path.basename(file_path)
        dest = os.path.join(self._storage_dir, name)
        if os.path.exists(dest):
            reply = frameless_message_box(
                self, "文件已存在",
                f"{name} 已存在，是否覆盖？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        try:
            shutil.copy2(file_path, dest)
            self.populate()
            logger.info(f"[SpreadsheetList] Imported: {name}")
        except Exception as e:
            logger.error(f"[SpreadsheetList] Import failed: {e}")

    def _on_delete_selected(self):
        rows = set(idx.row() for idx in self._table.selectedIndexes())
        if not rows:
            return
        names = [os.path.basename(self._files[r]) for r in rows if r < len(self._files)]
        reply = frameless_message_box(
            self, "确认删除",
            f"确定删除 {len(names)} 个表格？\n" + "\n".join(names[:5]),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            for r in sorted(rows, reverse=True):
                if r < len(self._files):
                    try:
                        os.remove(self._files[r])
                    except Exception as e:
                        logger.error(f"[SpreadsheetList] Delete failed: {e}")
            self.populate()

    def _load_column_widths(self) -> list:
        """从个人配置文件读取列宽"""
        try:
            cfg = load_config()
            widths = cfg.get('spreadsheet_list_column_widths')
            if isinstance(widths, list) and len(widths) == 3:
                return widths
        except Exception:
            pass
        return []

    def _save_column_widths(self):
        """将当前列宽保存到个人配置文件"""
        try:
            cfg = load_config()
            cfg['spreadsheet_list_column_widths'] = [
                self._table.columnWidth(0),
                self._table.columnWidth(1),
                self._table.columnWidth(2),
            ]
            save_config(cfg, immediate=True)
        except Exception as e:
            logger.error(f"[SpreadsheetList] Save column widths failed: {e}")

    def _on_column_resized(self, logicalIndex, oldSize, newSize):
        """列宽被拖拽改变时自动保存"""
        if not self._loading_widths:
            self._save_column_widths()

    def get_storage_dir(self) -> str:
        return self._storage_dir


class SpreadsheetPageMixin:
    """电子表格页面功能（MainFrame Mixin）"""

    def create_spreadsheet_page(self):
        """创建电子表格页面（列表 + 编辑器二级 stacked）"""
        page = QFrame()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        # 电子表格资源目录 & 存储目录
        self._spreadsheet_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'luckysheet'
        )
        self._spreadsheet_storage_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'user', 'spreadsheets'
        )

        # ===== 二级 stacked: 列表(0) + 编辑器(1) =====
        self._spreadsheet_page_stack = QStackedWidget()
        page_layout.addWidget(self._spreadsheet_page_stack)

        # --- index 0: 列表页 ---
        self._spreadsheet_list_page = SpreadsheetListPage(
            self._spreadsheet_storage_dir
        )
        self._spreadsheet_list_page.spreadsheetSelected.connect(
            self._on_spreadsheet_file_selected
        )
        self._spreadsheet_list_page.newSpreadsheet.connect(
            self._on_spreadsheet_new
        )
        self._spreadsheet_page_stack.addWidget(self._spreadsheet_list_page)

        # 首次创建时扫描已有文件
        self._spreadsheet_list_page.populate()

        # --- index 1: 编辑器页（懒加载占位） ---
        self._spreadsheet_editor_frame = None
        self._spreadsheet_editor_initialized = False

        self.content_stack.addWidget(page)
        self._spreadsheet_page = page

        # 状态标记
        self._spreadsheet_ready = False
        self._spreadsheet_pending_data = None
        self._spreadsheet_last_data = None
        self._spreadsheet_current_file = None  # 当前打开的文件路径

    def _ensure_spreadsheet_editor(self):
        """懒加载编辑器页面"""
        if self._spreadsheet_editor_initialized:
            return
        self._spreadsheet_editor_initialized = True

        editor = QFrame()
        editor_layout = QVBoxLayout(editor)
        editor_layout.setContentsMargins(0, 0, 0, 0)
        editor_layout.setSpacing(0)

        # ===== 编辑器工具栏 =====
        toolbar = QFrame()
        toolbar.setFixedHeight(44)
        toolbar.setStyleSheet(
            "QFrame { background: #FAFAFA; border-bottom: 1px solid #E8E8E8; }"
        )
        tb_layout = QHBoxLayout(toolbar)
        tb_layout.setContentsMargins(12, 6, 12, 6)

        btn_back = QPushButton("← 返回列表")
        btn_back.setFixedHeight(28)
        btn_back.clicked.connect(self.switch_to_spreadsheet_list)
        tb_layout.addWidget(btn_back)

        self._spreadsheet_title_label = QLabel("🧮 电子表格")
        self._spreadsheet_title_label.setStyleSheet(
            "font-size: 14px; font-weight: bold; color: #333; margin-left: 8px;"
        )
        tb_layout.addWidget(self._spreadsheet_title_label)

        tb_layout.addStretch()

        btn_import = QPushButton("📥 导入 XLSX")
        btn_import.setFixedHeight(28)
        btn_import.clicked.connect(self._on_spreadsheet_import_xlsx)
        tb_layout.addWidget(btn_import)

        btn_mysql = QPushButton("🗄️ MySQL导入")
        btn_mysql.setFixedHeight(28)
        btn_mysql.setToolTip("从已配置的 MySQL 数据库导入数据表")
        btn_mysql.clicked.connect(self._on_spreadsheet_import_mysql)
        tb_layout.addWidget(btn_mysql)

        btn_export = QPushButton("📤 导出 XLSX")
        btn_export.setFixedHeight(28)
        btn_export.clicked.connect(self._on_spreadsheet_export_xlsx)
        tb_layout.addWidget(btn_export)

        btn_read = QPushButton("📋 读取数据")
        btn_read.setFixedHeight(28)
        btn_read.setToolTip("将表格数据读回 Python 内存")
        btn_read.clicked.connect(self._on_spreadsheet_read_data)
        tb_layout.addWidget(btn_read)

        self._spreadsheet_status_label = QLabel("")
        self._spreadsheet_status_label.setStyleSheet(
            "color: #1890FF; font-size: 12px; margin-left: 8px;"
        )
        tb_layout.addWidget(self._spreadsheet_status_label)

        editor_layout.addWidget(toolbar)

        # ===== QWebEngineView =====
        self._spreadsheet_webview = QWebEngineView()
        self._spreadsheet_webview.setStyleSheet("background: #FFFFFF;")
        self._spreadsheet_webview.page().setBackgroundColor(QColor("#FFFFFF"))
        self._spreadsheet_webview.setContextMenuPolicy(
            Qt.ContextMenuPolicy.NoContextMenu
        )
        # 确保键盘事件（Ctrl+C/V/F/H 等）传递到电子表格
        self._spreadsheet_webview.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        # 启用 JavaScript 剪贴板访问（Ctrl+C/V 需要）
        from PyQt6.QtWebEngineCore import QWebEngineSettings
        self._spreadsheet_webview.settings().setAttribute(
            QWebEngineSettings.WebAttribute.JavascriptCanAccessClipboard, True
        )

        # ===== QWebChannel 接线 =====
        self._spreadsheet_bridge = SpreadsheetBridge(self)
        self._spreadsheet_channel = QWebChannel(self._spreadsheet_webview.page())
        self._spreadsheet_channel.registerObject("bridge", self._spreadsheet_bridge)
        self._spreadsheet_webview.page().setWebChannel(self._spreadsheet_channel)
        self._spreadsheet_bridge.set_page(self._spreadsheet_webview.page())
        self._spreadsheet_bridge.actionReceived.connect(self._on_spreadsheet_js_action)
        self._spreadsheet_webview.page().javaScriptConsoleMessage = self._on_spreadsheet_js_console

        editor_layout.addWidget(self._spreadsheet_webview, 1)

        self._spreadsheet_editor_frame = editor
        self._spreadsheet_page_stack.addWidget(editor)

        # 加载 HTML
        html_path = Path(self._spreadsheet_dir, 'luckysheet.html')
        if html_path.exists():
            try:
                port = _start_local_http(self._spreadsheet_dir)
                url = QUrl(f"http://127.0.0.1:{port}/luckysheet.html")
                self._spreadsheet_webview.load(url)
                logger.info(f"[Spreadsheet] Loading via local HTTP: {url.toString()}")
            except Exception as e:
                logger.error(f"[Spreadsheet] Local HTTP server failed: {e}")
                base_url = QUrl.fromLocalFile(self._spreadsheet_dir)
                self._spreadsheet_webview.setHtml(
                    html_path.read_text(encoding='utf-8'), base_url
                )
        else:
            logger.warning(f"[Spreadsheet] HTML template not found: {html_path}")

    # ===== 列表 ↔ 编辑器 导航 =====

    def switch_to_spreadsheet_list(self):
        """从编辑器返回列表（自动保存当前文件）"""
        # 自动保存
        if self._spreadsheet_current_file and self._spreadsheet_ready:
            self._auto_save_current_file()
        self._spreadsheet_list_page.populate()
        self._spreadsheet_page_stack.setCurrentIndex(0)

    def _auto_save_current_file(self):
        """自动保存当前编辑器数据到文件"""
        if not self._spreadsheet_current_file or not self._spreadsheet_last_data:
            # 请求最新数据并保存
            self._spreadsheet_bridge.request_cell_data()
            # 延迟保存，等待数据回传
            QTimer.singleShot(1000, self._do_auto_save)
            return
        self._do_auto_save()

    def _do_auto_save(self):
        """执行自动保存"""
        data = self._spreadsheet_last_data
        if not data or not isinstance(data, dict) or not self._spreadsheet_current_file:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            sheets = data.get('sheets', data)
            for sheet_id, sheet_info in sheets.items():
                if not isinstance(sheet_info, dict):
                    continue
                sheet_name = sheet_info.get('name', str(sheet_id))[:31]
                ws = wb.create_sheet(title=sheet_name)
                cell_data = sheet_info.get('cellData', {})
                for row_key, row_data in cell_data.items():
                    if not isinstance(row_data, dict):
                        continue
                    row_idx = int(row_key) + 1
                    for col_key, cell_info in row_data.items():
                        if not isinstance(cell_info, dict):
                            continue
                        col_idx = int(col_key) + 1
                        formula = cell_info.get('f')
                        if formula and isinstance(formula, str):
                            ws.cell(row=row_idx, column=col_idx, value=formula)
                        else:
                            val = cell_info.get('v')
                            if val is not None:
                                ws.cell(row=row_idx, column=col_idx, value=val)
            wb.save(self._spreadsheet_current_file)
            wb.close()
            logger.info(f"[Spreadsheet] Auto-saved: {self._spreadsheet_current_file}")
        except Exception as e:
            logger.error(f"[Spreadsheet] Auto-save failed: {e}")

    def _on_spreadsheet_file_selected(self, file_path: str):
        """双击列表中的文件 → 打开编辑器"""
        self._open_spreadsheet_editor(file_path)

    def _on_spreadsheet_new(self):
        """新建表格 → 创建空 xlsx → 打开编辑器"""
        import openpyxl
        name = f"表格_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path = os.path.join(self._spreadsheet_storage_dir, f"{name}.xlsx")
        os.makedirs(self._spreadsheet_storage_dir, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()
        self._open_spreadsheet_editor(file_path)

    def _open_spreadsheet_editor(self, file_path: str):
        """打开文件进入编辑器"""
        self._ensure_spreadsheet_editor()
        self._spreadsheet_current_file = file_path
        name = os.path.splitext(os.path.basename(file_path))[0]
        if hasattr(self, '_spreadsheet_title_label'):
            self._spreadsheet_title_label.setText(f"🧮 {name}")
        # 切到编辑器
        self._spreadsheet_page_stack.setCurrentIndex(1)
        # 聚焦 webview（确保键盘快捷键生效）
        QTimer.singleShot(200, lambda: self._spreadsheet_webview.setFocus() if hasattr(self, '_spreadsheet_webview') else None)
        # 导入数据
        self._import_file_to_editor(file_path)

    def _import_file_to_editor(self, file_path: str):
        """将 xlsx 文件导入到编辑器"""
        if not self._spreadsheet_ready:
            # 编辑器还没准备好，先存起来
            self._spreadsheet_pending_file = file_path
            # 触发编辑器初始化
            self._spreadsheet_page_stack.setCurrentIndex(1)
            return
        # 用后台线程导入
        self._import_btn_set_enabled(False)
        self._import_worker = _ImportWorker(file_path, self)
        self._import_worker.progress.connect(
            lambda msg: self._spreadsheet_status_show(msg)
        )
        self._import_worker.finished.connect(self._on_import_xlsx_done)
        self._import_worker.error.connect(self._on_import_xlsx_error)
        self._import_worker.start()

    # ===== JS Action 处理 =====

    def _on_spreadsheet_js_action(self, action: str, payload_json: str):
        """处理 JS 侧回调"""
        if action == 'page_ready':
            self._spreadsheet_ready = True
            # 加载挂起的数据
            if self._spreadsheet_pending_data is not None:
                self._spreadsheet_bridge.load_workbook_data(
                    self._spreadsheet_pending_data
                )
                self._spreadsheet_pending_data = None
            # 加载挂起的文件
            if hasattr(self, '_spreadsheet_pending_file') and self._spreadsheet_pending_file:
                pf = self._spreadsheet_pending_file
                self._spreadsheet_pending_file = None
                self._import_file_to_editor(pf)
            logger.info("[Spreadsheet] Page ready, Luckysheet initialized.")

        elif action == 'cell_changed':
            # 可选：实时跟踪单元格变更
            logger.debug(f"[Spreadsheet] Cell changed: {payload_json}")

        elif action == 'cell_data_response':
            try:
                raw = json.loads(payload_json)
                # JS 侧返回的错误
                if isinstance(raw, dict) and raw.get('error'):
                    err_msg = raw['error']
                    logger.error(f"[Spreadsheet] 读取数据失败: {err_msg}")
                    self._spreadsheet_status_show(f"❌ 读取失败: {err_msg}")
                    return
                if isinstance(raw, list):
                    # Luckysheet 返回数组格式，转为内部格式供保存/导出使用
                    self._spreadsheet_last_data = _luckysheet_to_univer(raw)
                else:
                    self._spreadsheet_last_data = raw
                sheet_count = 0
                if isinstance(self._spreadsheet_last_data, dict):
                    sheet_count = len(self._spreadsheet_last_data.get('sheets', {}))
                cell_total = 0
                for s in (self._spreadsheet_last_data or {}).get('sheets', {}).values():
                    cell_total += sum(len(r) for r in s.get('cellData', {}).values())
                logger.info(f"[Spreadsheet] 读取数据成功: {sheet_count} 个工作表, {cell_total} 个单元格")
                self._spreadsheet_status_show(f"✅ 读取完成: {sheet_count} 表, {cell_total} 单元格")
            except json.JSONDecodeError:
                self._spreadsheet_last_data = payload_json
                logger.error(f"[Spreadsheet] 读取数据失败: 返回的不是有效 JSON")
                self._spreadsheet_status_show("❌ 读取失败: 返回数据格式异常")

        elif action == 'data_loaded':
            logger.info(f"[Spreadsheet] Data loaded: {payload_json}")

        elif action == 'formula_error':
            logger.error(f"[Spreadsheet] Formula error: {payload_json}")

        elif action == 'xlsx_export_ready':
            self._on_spreadsheet_xlsx_exported(payload_json)

    def _on_spreadsheet_js_console(self, message, line, source):
        """捕获 JS 控制台消息并记录到日志"""
        if "error" in message.lower() or "uncaught" in message.lower() or "failed" in message.lower():
            logger.error(f"[JS] {message}")
        else:
            logger.info(f"[JS] {message}")

    # ===== 公开 API: 加载数据（Python → JS）=====

    def spreadsheet_load_data(self, workbook_data: dict):
        """
        加载工作簿数据到电子表格。

        Args:
            workbook_data: {
                "sheets": { "id": { "name": "...", "cellData": {...} } },
                "activeSheet": "id"
            }
        """
        if self._spreadsheet_ready:
            self._spreadsheet_bridge.load_workbook_data(workbook_data)
        else:
            self._spreadsheet_pending_data = workbook_data
            logger.info("[Spreadsheet] Data queued (page not ready yet)")

    def spreadsheet_load_dataframe(self, sheet_name: str, df):
        """
        从 pandas DataFrame 加载数据到电子表格。

        Args:
            sheet_name: Sheet 名称
            df: pandas.DataFrame（第一行为表头）
        """
        import pandas as pd
        if not isinstance(df, pd.DataFrame):
            logger.error("[spreadsheet_load_dataframe] df must be a pandas DataFrame")
            return

        cell_data = self._dataframe_to_cell_data(df)
        workbook_data = {
            "sheets": {
                sheet_name: {
                    "name": sheet_name,
                    "cellData": cell_data,
                }
            },
            "activeSheet": sheet_name,
        }
        self.spreadsheet_load_data(workbook_data)

    def _dataframe_to_cell_data(self, df) -> dict:
        """将 pandas DataFrame 转换为 Univer cellData 格式"""
        return _dataframe_to_cell_data_static(df)

    # ===== 公开 API: 读取数据（JS → Python）=====

    def spreadsheet_read_data(self) -> dict:
        """
        请求读回电子表格所有数据（异步）。

        Returns:
            上次缓存的数据（可能不是最新的，如需最新请在回调中获取）
        """
        self._spreadsheet_bridge.request_cell_data()
        return self._spreadsheet_last_data or {}

    # ===== 导入 XLSX =====

    def _on_spreadsheet_import_xlsx(self):
        """导入 .xlsx 文件到电子表格（后台线程，不卡死 UI）"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "导入 Excel 文件", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if not file_path:
            return

        self._import_btn_set_enabled(False)
        self._import_worker = _ImportWorker(file_path, self)
        self._import_worker.progress.connect(
            lambda msg: self._spreadsheet_status_show(msg)
        )
        self._import_worker.finished.connect(self._on_import_xlsx_done)
        self._import_worker.error.connect(self._on_import_xlsx_error)
        self._import_worker.start()
        logger.info(f"[Spreadsheet] Import started (background): {file_path}")

    def _on_import_xlsx_done(self, workbook_data: dict):
        """XLSX 后台导入完成"""
        self._import_btn_set_enabled(True)
        self.spreadsheet_load_data(workbook_data)
        self._spreadsheet_status_show("✅ 导入完成")

    def _on_import_xlsx_error(self, msg: str):
        """XLSX 后台导入失败"""
        self._import_btn_set_enabled(True)
        self._spreadsheet_status_show(f"❌ 导入失败: {msg}")
        logger.error(f"[Spreadsheet] Import failed: {msg}")

    def _on_spreadsheet_import_mysql(self):
        """从 MySQL 导入数据到电子表格"""
        mysql_config = getattr(self, 'config', {}).get('mysql_config', {})
        if not mysql_config.get('enabled'):
            frameless_message_box(
                self, "MySQL 未配置",
                "请先在 设置 → MySQL配置 中启用并填写数据库连接信息。"
            )
            return

        dlg = MysqlImportDialog(mysql_config, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        sql, sheet_name = dlg.get_import_info()
        if not sql:
            return

        self._import_btn_set_enabled(False)
        self._mysql_worker = _MysqlQueryWorker(mysql_config, sql, sheet_name, self)
        self._mysql_worker.progress.connect(
            lambda msg: self._spreadsheet_status_show(msg)
        )
        self._mysql_worker.finished.connect(self._on_mysql_import_done)
        self._mysql_worker.error.connect(self._on_mysql_import_error)
        self._mysql_worker.start()
        logger.info(f"[Spreadsheet] MySQL import started: {sheet_name}")

    def _on_mysql_import_done(self, workbook_data: dict, sheet_name: str):
        """MySQL 后台导入完成"""
        try:
            self._import_btn_set_enabled(True)
            self.spreadsheet_load_data(workbook_data)
            self._spreadsheet_status_show(f"✅ 从 MySQL 导入 [{sheet_name}] 完成")
        except Exception as e:
            logger.error(f"[Spreadsheet] MySQL import callback error: {e}")
            self._spreadsheet_status_show(f"❌ 数据加载失败: {e}")

    def _on_mysql_import_error(self, msg: str):
        """MySQL 后台导入失败"""
        self._import_btn_set_enabled(True)
        self._spreadsheet_status_show(f"❌ MySQL导入失败: {msg}")
        logger.error(f"[Spreadsheet] MySQL import failed: {msg}")

    def _import_btn_set_enabled(self, enabled: bool):
        """启用/禁用导入按钮（防止重复点击）"""
        if hasattr(self, '_spreadsheet_page'):
            for btn in self._spreadsheet_page.findChildren(QPushButton):
                txt = btn.text()
                if '导入' in txt or 'MySQL' in txt:
                    btn.setEnabled(enabled)

    def _spreadsheet_status_show(self, msg: str):
        """在工具栏标题旁显示临时状态"""
        if hasattr(self, '_spreadsheet_status_label'):
            self._spreadsheet_status_label.setText(msg)
            self._spreadsheet_status_label.repaint()

    # ===== 导出 XLSX =====

    def _on_spreadsheet_export_xlsx(self):
        """导出电子表格为 .xlsx 文件"""
        # 先请求最新数据
        self._spreadsheet_bridge.request_cell_data()
        # 等待异步回传后执行导出
        QTimer.singleShot(800, self._do_export_xlsx)

    def _do_export_xlsx(self):
        """执行实际导出（在 request_cell_data 回调后）"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel 文件", "spreadsheet.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        data = self._spreadsheet_last_data
        if not data or not isinstance(data, dict):
            frameless_message_box(self, "导出失败", "没有可导出的数据")
            return

        try:
            import openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            sheets = data.get('sheets', data)
            for sheet_id, sheet_info in sheets.items():
                if not isinstance(sheet_info, dict):
                    continue
                sheet_name = sheet_info.get('name', str(sheet_id))
                # Excel sheet 名最长 31 字符
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                ws = wb.create_sheet(title=sheet_name)

                cell_data = sheet_info.get('cellData', {})
                for row_key, row_data in cell_data.items():
                    if not isinstance(row_data, dict):
                        continue
                    row_idx = int(row_key) + 1  # openpyxl 是 1-based
                    for col_key, cell_info in row_data.items():
                        if not isinstance(cell_info, dict):
                            continue
                        col_idx = int(col_key) + 1
                        formula = cell_info.get('f')
                        if formula and isinstance(formula, str):
                            # 公式单元格：写入公式字符串（Excel/WPS 打开自动计算）
                            ws.cell(row=row_idx, column=col_idx, value=formula)
                        else:
                            val = cell_info.get('v', None)
                            if val is not None:
                                ws.cell(row=row_idx, column=col_idx, value=val)

            wb.save(file_path)
            wb.close()
            logger.info(f"[Spreadsheet] Exported: {file_path}")
            frameless_message_box(self, "导出成功", f"已保存到:\n{file_path}")

        except Exception as e:
            logger.error(f"[Spreadsheet] Export failed: {e}")
            frameless_message_box(self, "导出失败", str(e))

    # ===== 读取数据按钮 =====

    def _on_spreadsheet_read_data(self):
        """读取数据按钮回调"""
        self._spreadsheet_status_show("⏳ 正在读取表格数据...")
        self._spreadsheet_bridge.request_cell_data()
        logger.info("[Spreadsheet] Requesting cell data from JS...")
