# -*- coding: utf-8 -*-
from core import *
from common import *
from common import FilterPanel  # Explicit import for FilterPanel migration
import common  # 显式导入，用于访问模块级私有函数
choice = None  # 全局功能选择标识，在后台线程中设置

"""
file_generation.py — 文件生成 Mixin（Excel → 合同）
─────────────────────────────────────────────────
负责：MainFrame 中文件生成页面相关方法
  - create_excel_to_pdf_page()  文件生成页面（模板选择 / 数据处理 / 合同生成 / PDF转换）
  - 所有 _data_* / on_data_* / _get_data_* 方法（筛选 / 排序 / 分页 / 搜索）
  - 打印登录用户信息
依赖：core.py / common.py
被导入：主程序（作为 MainFrame 的 Mixin 父类）
"""
"""file_generation Mixin"""

# 导入所需模块
import os

# QtWebEngine 必须在 QApplication 创建前导入
import PyQt6.QtWebEngineWidgets  # noqa: F401

from pathlib import Path  # 路径处理
import copy
import builtins
from functools import lru_cache
from io import BytesIO
import logging  # 日志记录
import platform  # 操作系统平台信息
import shutil  # 文件操作
import json  # JSON处理
import subprocess  # 子进程管理
from contextlib import contextmanager
from datetime import datetime, timedelta  # 日期时间处理
from decimal import Decimal, ROUND_HALF_UP

import sys  # 系统相关
import sqlite3  # SQLite 本地缓存
import os  # 操作系统接口
import ssl
import time  # 时间
import threading  # 线程管理
import urllib.request
import email.utils
import pkgutil
import tempfile
import psutil  # 进程管理
import re  # 正则表达式
import uuid  # UUID 生成（API 配置等）
import hashlib  # 哈希算法
import certifi
import requests  # HTTP请求
from requests.adapters import HTTPAdapter
from urllib3.poolmanager import PoolManager


from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QComboBox, QLineEdit, QHeaderView, QAbstractItemView, QSpinBox,
    QMessageBox, QDialog, QTreeWidget, QFileDialog, QCheckBox, QFrame, QScrollArea,
    QGroupBox, QStackedWidget, QRadioButton, QCalendarWidget, QDateEdit, QListWidget,
    QListWidgetItem, QMenu, QSizePolicy, QInputDialog, QSplitter, QToolButton, QProgressBar,
)
from PyQt6.QtCore import Qt, QTimer, QEvent, pyqtSignal, QDate, QSize, QPoint
from PyQt6.QtGui import QFont, QColor, QPalette, QAction, QIcon

class file_generationMixin:
    """file_generation functionality."""

    def save_user_runtime_state_patch(self, updates, immediate=True):
        """保存当前用户的运行状态补丁"""
        self.runtime_state = deep_merge_dict(getattr(self, 'runtime_state', {}), updates or {})
        # ✅ 关键修复：写入前从文件重新加载 table_column_widths，
        #    避免用 MainFrame 内存中的旧数据覆盖 SettingsDialog 刚保存的列宽
        try:
            file_state = load_user_runtime_state()
            file_widths = file_state.get('table_column_widths', {})
            if file_widths:
                if 'table_column_widths' not in self.runtime_state:
                    self.runtime_state['table_column_widths'] = {}
                self.runtime_state['table_column_widths'] = file_widths
        except Exception:
            pass
        save_user_runtime_state(self.runtime_state, immediate=immediate, merge=False)


    def restore_user_runtime_state(self):
        """恢复当前用户上次运行时的主界面状态"""
        self.runtime_state = load_user_runtime_state()

        # 恢复所有模块的分页大小
        self._restore_page_sizes()

        contract_state = self.runtime_state.get('contract_page_state', {})

        if hasattr(self, 'selected_filter_start_date') and hasattr(self, 'selected_filter_end_date'):
            start_date_str = contract_state.get('date_start')
            end_date_str = contract_state.get('date_end')
            if start_date_str and end_date_str:
                start_date = QDate.fromString(start_date_str, 'yyyy-MM-dd')
                end_date = QDate.fromString(end_date_str, 'yyyy-MM-dd')
                if start_date.isValid() and end_date.isValid():
                    self.selected_filter_start_date = start_date
                    self.selected_filter_end_date = end_date
                    if hasattr(self, 'date_filter_button'):
                        self.date_filter_button.setText(
                            f"{start_date.toString('yyyy-MM-dd')} ~ {end_date.toString('yyyy-MM-dd')}"
                        )
            elif 'date_start' in contract_state or 'date_end' in contract_state:
                self.selected_filter_start_date = None
                self.selected_filter_end_date = None
                if hasattr(self, 'date_filter_button'):
                    self.date_filter_button.setText("请选择日期范围")

        if hasattr(self, 'data_search_input'):
            search_text = contract_state.get('search_text', '')
            if search_text:
                self.data_search_input.blockSignals(True)
                self.data_search_input.setText(search_text)
                self.data_search_input.blockSignals(False)

        if hasattr(self, 'output_frame'):
            self.output_frame.setVisible(self.output_visible)
            self.toggle_output_btn.setText("隐藏" if self.output_visible else "显示")

        selected_operation = contract_state.get('selected_operation')
        if hasattr(self, 'operation_combo') and selected_operation is not None:
            combo_index = self.operation_combo.findData(selected_operation)
            if combo_index >= 0:
                self.operation_combo.setCurrentIndex(combo_index)

        # ✅ 修复：不再在启动时自动应用列宽（避免被后续数据加载覆盖）
        # 列宽现在只在数据加载时（_update_data_table中）直接应用
        # from PyQt6.QtCore import QTimer
        # QTimer.singleShot(500, self._delayed_apply_data_table_column_widths)
        QTimer.singleShot(500, self._delayed_apply_crm_column_widths)


    def persist_contract_page_state(self):
        """保存合同生成页面的用户运行状态"""
        start_date = getattr(self, 'selected_filter_start_date', None)
        end_date = getattr(self, 'selected_filter_end_date', None)
        self.save_user_runtime_state_patch({
            'contract_page_state': {
                'selected_operation': getattr(self, 'selected_operation', None),
                'page_size': self.page_size_combo.currentData() if hasattr(self, 'page_size_combo') else 20,
                'current_page': getattr(self, 'current_page', 1),
                'search_text': self.data_search_input.text().strip() if hasattr(self, 'data_search_input') else "",
                'date_start': start_date.toString('yyyy-MM-dd') if start_date and start_date.isValid() else "",
                'date_end': end_date.toString('yyyy-MM-dd') if end_date and end_date.isValid() else "",
            }
        }, immediate=True)

    def create_excel_to_pdf_page(self):
        """创建Excel to PDF功能页面（页面0）"""
        page = QFrame()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(15, 10, 15, 10)
        layout.setSpacing(8)

        file_generation_group = QGroupBox("文件生成")
        file_generation_layout = QVBoxLayout(file_generation_group)
        file_generation_layout.setSpacing(8)

        # 第一行：下拉选项框 + 日期筛选框
        row1_layout = QHBoxLayout()
        row1_layout.setSpacing(15)

        options_label = QLabel("表格选择：")
        row1_layout.addWidget(options_label)

        # 下拉选项框（替代原来的单选按钮）
        from PyQt6.QtWidgets import QComboBox
        self.operation_combo = QComboBox()
        self.operation_combo.setFixedWidth(120)
        self.operation_combo.setFixedHeight(28)
        self.operation_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #8a8a8a;
                border-radius: 4px;
                padding: 2px 10px;
                background-color: #ffffff;
            }
            QComboBox:focus {
                border: 1px solid #2d8cf0;
            }
        """)
        gui_options = self.config.get('gui_options', [
            {"text": "功能1", "value": 1},
            {"text": "功能2", "value": 2},
            {"text": "功能3", "value": 3},
            {"text": "功能4", "value": 4},
            {"text": "功能5", "value": 5}
        ])

        gui_options = [option for option in gui_options if option.get('text', '') != 'PDF文件整理']
        gui_options.sort(key=lambda x: x['value'])

        self.gui_options = gui_options
        self.operation_map = {i: option['value'] for i, option in enumerate(gui_options)}

        self.operation_combo.addItem("请选择功能", None)

        for i, option in enumerate(gui_options):
            text = option.get("text", "")
            value = option.get("value", 0)
            self.operation_combo.addItem(text, value)

        self.selected_operation = None

        self.operation_combo.currentIndexChanged.connect(self.on_operation_combo_changed)
        row1_layout.addWidget(self.operation_combo)

        row1_layout.addSpacing(10)

        # 筛选方案按钮（点击弹出方案列表）
        self.data_preset_btn = QPushButton("方案 ▼")
        self.data_preset_btn.setFixedHeight(30)
        self.data_preset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.data_preset_btn.setToolTip("选择筛选方案")
        self.data_preset_btn.setStyleSheet("QPushButton { border: 1px solid #D9D9D9; border-radius: 4px; padding: 2px 10px; font-size: 12px; color: #333; background: #FFF; } QPushButton:hover { border-color: #1890FF; color: #1890FF; }")
        self.data_preset_btn.clicked.connect(self._show_data_preset_popup)
        row1_layout.addWidget(self.data_preset_btn)

        # 筛选按钮（带计数）
        self.data_filter_toggle_btn = QPushButton("筛选")
        self.data_filter_toggle_btn.setFixedHeight(30)
        self.data_filter_toggle_btn.setStyleSheet("""
            QPushButton {
                border: 1px solid #D9D9D9; border-radius: 4px;
                padding: 4px 12px; font-size: 13px;
                background-color: #FFFFFF; color: #333;
            }
            QPushButton:hover { border-color: #1890FF; color: #1890FF; }
        """)
        self.data_filter_toggle_btn.clicked.connect(self._toggle_data_filter_panel)
        row1_layout.addWidget(self.data_filter_toggle_btn)

        # 搜索框（字段选择 + 分隔线 + 搜索输入，仿 CRM）
        search_frame = QFrame()
        search_frame.setFixedHeight(30)
        search_frame.setMaximumWidth(260)
        search_frame.setStyleSheet("QFrame { border: 1px solid #D9D9D9; border-radius: 4px; background: #FFF; } QFrame:focus-within { border-color: #1890FF; }")
        search_layout = QHBoxLayout(search_frame)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(0)

        # 搜索输入区域（文本 / 日期范围切换）
        self.data_search_stack = QStackedWidget()

        self.data_search_input = QLineEdit()
        self.data_search_input.setPlaceholderText("搜索")
        self.data_search_input.setFixedHeight(28)
        self.data_search_input.setStyleSheet("QLineEdit { border: none; background: transparent; padding: 2px 8px; font-size: 12px; }")
        self.data_search_input.textChanged.connect(self.on_global_search_changed)
        self.data_search_input.returnPressed.connect(self._apply_data_filters)
        self.data_search_stack.addWidget(self.data_search_input)  # Page 0

        self.data_search_date_btn = QPushButton("选择日期范围")
        self.data_search_date_btn.setFixedHeight(28)
        self.data_search_date_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.data_search_date_btn.setStyleSheet("QPushButton { border: none; background: transparent; padding: 2px 8px; font-size: 11px; color: #999; } QPushButton:hover { color: #1890FF; }")
        self.data_search_date_btn.setProperty('date_range', {'start': None, 'end': None})
        self.data_search_date_btn.clicked.connect(lambda: self._open_data_search_date_range())
        self.data_search_stack.addWidget(self.data_search_date_btn)  # Page 1

        search_layout.addWidget(self.data_search_stack, stretch=1)
        row1_layout.addWidget(search_frame)

        row1_layout.addSpacing(10)
        keep_format = self.config.get('app_settings', {}).get('keep_output_format', 'word_pdf')
        self.keep_format_combo = QComboBox()
        self.keep_format_combo.addItems(["Word+PDF", "Word", "PDF"])
        fmt_map = {"word_pdf": "Word+PDF", "word": "Word", "pdf": "PDF"}
        self.keep_format_combo.setCurrentText(fmt_map.get(keep_format, "Word+PDF"))
        self.keep_format_combo.setFixedWidth(110)
        self.keep_format_combo.setToolTip("生成后保留的文件格式")
        self.keep_format_combo.currentIndexChanged.connect(self.on_keep_format_changed)

        self.refresh_data_btn = QPushButton("⟳ 刷新")
        self.refresh_data_btn.setFixedSize(self._scaled_ui_value(92, minimum=88), self._scaled_ui_value(30, minimum=28)) # 根据UI缩放调整按钮大小
        self.refresh_data_btn.setToolTip("刷新数据")
        self.refresh_data_btn.setStyleSheet(get_compact_refresh_button_style(self.ui_scale_percent))
        self.refresh_data_btn.clicked.connect(self.on_refresh_data)

        row1_layout.addStretch()
        row1_layout.addWidget(self.keep_format_combo)
        row1_layout.addSpacing(10)

        self.confirm_btn = QPushButton("确认执行")
        self.confirm_btn.setFixedWidth(100)
        self.confirm_btn.setFixedHeight(28)
        self.confirm_btn.clicked.connect(self.on_confirm_operation)
        row1_layout.addWidget(self.confirm_btn)

        file_generation_layout.addLayout(row1_layout)

        # ===== 新增：数据选择区域 =====
        data_selection_group = QGroupBox("数据选择")
        data_selection_layout = QVBoxLayout(data_selection_group)
        data_selection_layout.setSpacing(8)

        # 工具栏：筛选 + 搜索 + 字段显示 + 刷新
        # 工具栏：字段显示 + 刷新
        toolbar_layout = QHBoxLayout()
        toolbar_layout.setSpacing(8)
        toolbar_layout.addStretch()

        self.column_settings_btn = QPushButton("字段显示")
        self.column_settings_btn.setFixedWidth(80)
        self.column_settings_btn.clicked.connect(self.open_column_settings_dialog)
        toolbar_layout.addWidget(self.column_settings_btn)

        # 外露标签（放在工具栏内，与按钮同行）
        self.data_exposed_tags = QFrame()
        self.data_exposed_tags.setVisible(False)
        self.data_exposed_tags.setStyleSheet("QFrame { background: transparent; border: none; }")
        data_tags_layout = QHBoxLayout(self.data_exposed_tags)
        data_tags_layout.setContentsMargins(0, 0, 0, 0)
        data_tags_layout.setSpacing(4)

        toolbar_layout.insertWidget(0, self.data_exposed_tags)
        toolbar_layout.addWidget(self.refresh_data_btn)

        data_selection_layout.addLayout(toolbar_layout)

        # ===== 筛选条件面板（弹窗形式，仿 CRM） =====
        self.data_filter_panel = QFrame()
        self.data_filter_panel.setVisible(False)
        self.data_filter_panel.setFixedWidth(500)
        self.data_filter_panel.setStyleSheet("QFrame { border: 1px solid #E8E8E8; border-radius: 6px; background: #FFF; }")
        dp_layout = QVBoxLayout(self.data_filter_panel)
        dp_layout.setContentsMargins(0, 0, 0, 0)
        dp_layout.setSpacing(0)

        # 标题栏
        dp_header = QFrame()
        dp_header.setFixedHeight(36)
        dp_header.setStyleSheet("QFrame { background: #FAFAFA; border-bottom: 1px solid #E8E8E8; border-top-left-radius: 6px; border-top-right-radius: 6px; }")
        dh_layout = QHBoxLayout(dp_header)
        dh_layout.setContentsMargins(12, 0, 8, 0)
        dh_layout.addWidget(QLabel("设置筛选"))
        dh_layout.addStretch()
        dp_close = QPushButton("×")
        dp_close.setFixedSize(24, 24)
        dp_close.setStyleSheet("QPushButton { border: none; font-size: 14px; color: #999; background: transparent; } QPushButton:hover { color: #333; }")
        dp_close.clicked.connect(self._toggle_data_filter_panel)
        dh_layout.addWidget(dp_close)
        dp_layout.addWidget(dp_header)

        # 条件行区域
        self.data_conditions_frame = QFrame()
        self.data_conditions_frame.setStyleSheet("QFrame { background: #FFF; border: none; }")
        self.data_conditions_layout = QVBoxLayout(self.data_conditions_frame)
        self.data_conditions_layout.setContentsMargins(8, 8, 8, 4)
        self.data_conditions_layout.setSpacing(4)
        self.data_conditions_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.data_condition_rows = []

        # 底部按钮行：另存为 + 添加条件 + 清除 + 筛选
        dp_footer = QFrame()
        dp_footer.setFixedHeight(44)
        dp_footer.setStyleSheet("QFrame { background: #FAFAFA; border-top: 1px solid #E8E8E8; border-bottom-left-radius: 6px; border-bottom-right-radius: 6px; }")
        df_layout = QHBoxLayout(dp_footer)
        df_layout.setContentsMargins(12, 6, 12, 6)
        df_layout.setSpacing(8)

        save_btn = QPushButton("另存为")
        save_btn.setFixedHeight(28)
        save_btn.setStyleSheet("QPushButton { border: 1px solid #D9D9D9; border-radius: 4px; font-size: 12px; padding: 2px 12px; background: #FFF; color: #333; } QPushButton:hover { border-color: #1890FF; color: #1890FF; }")
        save_btn.clicked.connect(lambda: self._save_data_filter_preset())
        df_layout.addWidget(save_btn)

        add_cond_btn = QPushButton("+ 添加条件")
        add_cond_btn.setFixedHeight(28)
        add_cond_btn.setStyleSheet("QPushButton { border: 1px dashed #1890FF; border-radius: 4px; color: #1890FF; font-size: 12px; padding: 2px 12px; background: transparent; } QPushButton:hover { background: #E6F7FF; }")
        add_cond_btn.clicked.connect(lambda: self._add_data_condition_row())
        df_layout.addWidget(add_cond_btn)

        clear_cond_btn = QPushButton("清除")
        clear_cond_btn.setFixedHeight(28)
        clear_cond_btn.setStyleSheet("QPushButton { border: 1px solid #D9D9D9; border-radius: 4px; font-size: 12px; padding: 2px 12px; background: #FFF; color: #666; } QPushButton:hover { border-color: #FF4D4F; color: #FF4D4F; }")
        clear_cond_btn.clicked.connect(lambda: [self._clear_all_data_conditions(), self._apply_data_filters()])
        df_layout.addWidget(clear_cond_btn)

        df_layout.addStretch()
        apply_btn = QPushButton("筛选")
        apply_btn.setFixedHeight(28)
        apply_btn.setFixedWidth(70)
        apply_btn.setStyleSheet("QPushButton { background: #1890FF; color: #FFF; border: none; border-radius: 4px; font-size: 12px; } QPushButton:hover { background: #40A9FF; }")
        apply_btn.clicked.connect(self._apply_data_filter_and_close)
        df_layout.addWidget(apply_btn)

        dp_layout.addWidget(self.data_conditions_frame, stretch=1)
        dp_layout.addWidget(dp_footer)
        data_selection_layout.addWidget(self.data_filter_panel)

        # 数据表格（带复选框，支持列拖拽排序，铺满页面）
        from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem, QSplitter
        self.data_table = QTableWidget()
        install_table_edit_context_menu(self.data_table)
        self.table_header = CheckBoxHeader(self.data_table)
        self.table_header.toggled.connect(self.on_header_select_all_toggled)
        self.data_table.setHorizontalHeader(self.table_header)
        self.data_table.setColumnCount(0)
        self.data_table.setRowCount(0)
        self.data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.data_table.cellClicked.connect(self.on_data_table_cell_clicked)
        self.data_table.itemChanged.connect(self.on_data_table_item_changed)
        # ✅ 修复：移除setStretchLastSection，避免最后一列自动拉伸导致保存的列宽被覆盖
        self.table_header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table_header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        # 不再使用 setStretchLastSection(True)，让用户调整的列宽完全保持
        self.data_table.verticalHeader().setVisible(False)
        self.data_table.setAlternatingRowColors(True)
        self.table_header.setSectionsMovable(True)
        self.table_header.sectionMoved.connect(self.on_data_table_header_moved)
        self.table_header.sectionResized.connect(self._on_data_table_column_resized)
        self.table_header.sectionClicked.connect(self.on_data_table_header_clicked)
        self.table_header.setSortIndicatorShown(False)

        # 让表格铺满可用空间
        data_selection_layout.addWidget(self.data_table, stretch=1)

        # 分页栏
        pagination_layout = QHBoxLayout()
        self.data_header_label = QLabel("请选择功能后加载数据")
        self.data_header_label.setStyleSheet("font-weight: bold; color: #666;")
        pagination_layout.addWidget(self.data_header_label)
        pagination_layout.addStretch()
        pagination_layout.addWidget(QLabel("每页显示："))

        self.page_size_combo = QComboBox()
        self.page_size_combo.setEditable(True)
        self.page_size_combo.addItem("自定义", -1)
        for page_size in (20, 50, 80, 100, 200):
            self.page_size_combo.addItem(str(page_size), page_size)
        self.page_size_combo.setCurrentIndex(1)
        self.page_size_combo.setStyleSheet("QComboBox { font-size: 11px; padding: 2px 4px; }")
        self.page_size_combo.currentIndexChanged.connect(self.on_page_size_changed)
        pagination_layout.addWidget(self.page_size_combo)

        pagination_layout.addSpacing(12)
        self.prev_page_btn = QPushButton("<")
        self.prev_page_btn.setFixedWidth(34)
        self.prev_page_btn.clicked.connect(self.on_prev_page)
        pagination_layout.addWidget(self.prev_page_btn)

        self.page_label = QLabel("1/1")
        self.page_label.setFixedWidth(60)
        self.page_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.page_label.setStyleSheet("font-weight: bold; color: #333;")
        pagination_layout.addWidget(self.page_label)

        self.next_page_btn = QPushButton(">")
        self.next_page_btn.setFixedWidth(34)
        self.next_page_btn.clicked.connect(self.on_next_page)
        pagination_layout.addWidget(self.next_page_btn)
        data_selection_layout.addLayout(pagination_layout)

        file_generation_layout.addWidget(data_selection_group, stretch=1)  # stretch=1让数据区铺满
        # ===== 数据选择区域结束 =====

        layout.addWidget(file_generation_group, 1)

        # 保存下拉框和日期筛选框的引用
        self.operation_combo_frame = row1_layout

        # 初始化数据缓存
        self._data_cache = {}  # 缓存格式: {operation_value: (headers, data_rows)}
        self._cache_valid = False
        self._is_updating_data_table = False
        self.current_all_headers = []
        self.current_visible_headers = []
        self.current_filtered_rows = []
        self.current_source_headers = []
        self.current_header_source_index_map = {}
        self.current_page = 1
        self._selected_row_ids_by_operation = {}
        self._date_filter_timer = QTimer(self)
        self._date_filter_timer.setSingleShot(True)
        self._date_filter_timer.timeout.connect(self.apply_date_filter_now)
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self.apply_search_filter_now)
        self.data_sort_header = ""
        self.data_sort_order = Qt.SortOrder.AscendingOrder

        self.content_stack.addWidget(page)


    def _save_data_table_column_widths(self):
        """保存文件生成表格列宽设置 - 按功能选项分别保存"""
        try:
            if not hasattr(self, 'data_table'):
                return

            column_widths = {}
            for col in range(self.data_table.columnCount()):
                width = self.data_table.columnWidth(col)
                column_widths[col] = width

            print(f"[DEBUG-保存] 保存文件生成列宽 | 功能:{getattr(self, 'selected_operation', None)} | 列数:{len(column_widths)} | 数据:{column_widths}")

            config = load_config()  # 强制加载最新配置
            if 'app_settings' not in config:
                config['app_settings'] = {}
            if 'data_table_settings' not in config['app_settings']:
                config['app_settings']['data_table_settings'] = {}

            # ✅ 修复：按功能选项(gui_operation)分别保存列宽，避免不同功能互相覆盖
            operation_value = getattr(self, 'selected_operation', None)
            if operation_value is not None:
                operation_key = str(operation_value)
                if 'column_widths_by_operation' not in config['app_settings']['data_table_settings']:
                    config['app_settings']['data_table_settings']['column_widths_by_operation'] = {}
                config['app_settings']['data_table_settings']['column_widths_by_operation'][operation_key] = column_widths
                self.save_user_runtime_state_patch({
                    'table_column_widths': {
                        'data_table_by_operation': {
                            operation_key: column_widths
                        }
                    }
                }, immediate=True)
                print(f"[DEBUG-保存] 已保存到 column_widths_by_operation[{operation_key}]")

            # 同时保存到通用位置（向后兼容）
            config['app_settings']['data_table_settings']['column_widths'] = column_widths
            save_config(config, immediate=False)
            # 同步更新 self.config
            self.config = config

        except Exception as e:
            print(f"保存文件生成列宽失败: {e}")
            operation_log.error(f"保存文件生成列宽失败: {e}")


    def _debounced_save_data_table_column_widths(self):
        """延迟保存文件生成表格列宽（防抖后的实际保存操作）"""
        self._save_data_table_column_widths()


    def _delayed_apply_data_table_column_widths(self):
        """延迟应用文件生成表格列宽 - 只在首次加载时应用，防止重复覆盖"""
        try:
            print(f"[DEBUG-延迟应用] _delayed_apply_data_table_column_widths 被调用")
            print(f"[DEBUG-延迟应用] _data_table_column_widths_applied={getattr(self, '_data_table_column_widths_applied', False)}")

            # ✅ 如果已经应用过，不再重复设置（保护用户调整的列宽）
            if getattr(self, '_data_table_column_widths_applied', False):
                print(f"[DEBUG-延迟应用] 跳过：已应用过")
                return

            if not hasattr(self, 'data_table') or self.data_table.columnCount() == 0:
                print(f"[DEBUG-延迟应用] 跳过：表格不存在或无列 | 列数:{getattr(self, 'data_table', None).columnCount() if hasattr(self, 'data_table') and self.data_table else 0}")
                return

            config = load_config()
            app_set = config.get('app_settings', {})
            data_set = app_set.get('data_table_settings', {})
            runtime_state = load_user_runtime_state()
            runtime_widths = runtime_state.get('table_column_widths', {})

            # ✅ 修复：优先从按功能选项保存的列宽中读取
            operation_value = getattr(self, 'selected_operation', None)
            saved_widths_raw = {}
            if operation_value is not None:
                runtime_by_operation = runtime_widths.get('data_table_by_operation', {})
                op_key = str(operation_value)
                if op_key in runtime_by_operation:
                    saved_widths_raw = runtime_by_operation[op_key]
                    print(f"[DEBUG-延迟应用] [OK] 找到用户运行态功能专属列宽: {saved_widths_raw}")

            if not saved_widths_raw and operation_value is not None:
                by_operation = data_set.get('column_widths_by_operation', {})
                op_key = str(operation_value)
                print(f"[DEBUG-延迟应用] 查找功能{op_key}的列宽 | column_widths_by_operation存在:{'column_widths_by_operation' in data_set}")
                if op_key in by_operation:
                    saved_widths_raw = by_operation[op_key]
                    print(f"[DEBUG-延迟应用] [OK] 找到功能专属列宽: {saved_widths_raw}")

            # 如果没有找到按功能选项保存的列宽，则使用通用位置（向后兼容）
            if not saved_widths_raw:
                saved_widths_raw = runtime_widths.get('data_table', {})
                if saved_widths_raw:
                    print(f"[DEBUG-延迟应用] 使用用户运行态通用列宽: {saved_widths_raw}")

            if not saved_widths_raw:
                saved_widths_raw = data_set.get('column_widths', {})
                print(f"[DEBUG-延迟应用] 使用通用列宽: {saved_widths_raw}")

            # ✅ 关键修复：统一键类型为整数（配置文件可能存储为字符串）
            saved_widths = {}
            for k, v in saved_widths_raw.items():
                try:
                    saved_widths[int(k)] = v
                except (ValueError, TypeError):
                    pass

            print(f"[DEBUG-延迟应用] 解析后列宽: {saved_widths} | 表格当前列数: {self.data_table.columnCount()}")

            if saved_widths:
                applied_count = 0
                for col_idx in range(self.data_table.columnCount()):
                    if col_idx in saved_widths:
                        self.data_table.setColumnWidth(col_idx, saved_widths[col_idx])
                        applied_count += 1
                print(f"[DEBUG-延迟应用] 已应用 {applied_count} 个列宽")

                # 验证：打印实际设置的值
                verify_widths = {}
                for col_idx in range(min(5, self.data_table.columnCount())):
                    verify_widths[col_idx] = self.data_table.columnWidth(col_idx)
                print(f"[DEBUG-延迟应用] 验证前5列实际宽度: {verify_widths}")

            # ✅ 标记为已应用
            self._data_table_column_widths_applied = True
            print(f"[DEBUG-延迟应用] 完成，标记为已应用")

        except Exception as e:
            print(f"[ERROR] 延迟应用文件生成列宽失败: {e}")
            import traceback
            traceback.print_exc()


    def _on_data_table_column_resized(self, logicalIndex, oldSize, newSize):
        """文件生成表格列宽调整时使用防抖机制保存（避免频繁IO）"""
        try:
            print(f"[DEBUG-列宽调整] 文件生成表格 | 列:{logicalIndex} | 旧宽度:{oldSize} -> 新宽度:{newSize}")

            # 检查定时器是否存在
            if not hasattr(self, '_data_column_width_save_timer'):
                print(f"[DEBUG-列宽调整] [ERROR] 定时器不存在！")
                return

            # 复用定时器，重置防抖计时
            if self._data_column_width_save_timer.isActive():
                self._data_column_width_save_timer.stop()

            # 启动防抖定时器（1秒后保存）
            self._data_column_width_save_timer.start(1000)
            print(f"[DEBUG-列宽调整] [OK] 已启动防抖定时器（1秒后保存）")
        except Exception as e:
            print(f"文件生成列宽调整事件处理失败: {e}")
            import traceback
            traceback.print_exc()


    def on_radio_select(self, value):
        """单选按钮选择事件处理（兼容旧代码）"""
        self.selected_operation = value
        self._on_operation_changed(value)


    def on_operation_combo_changed(self, index):
        """下拉选项框变化事件处理"""
        if index < 0:
            return
        value = self.operation_combo.currentData()
        if value is None:
            self.selected_operation = None
            self.persist_contract_page_state()
            self._clear_data_table("请选择功能后加载数据")
            return
        self.selected_operation = value
        self.persist_contract_page_state()
        self._on_operation_changed(value)


    def _on_operation_changed(self, value):
        """功能选择变化的公共处理（使用缓存，不强制刷新）"""
        # ✅ 修复：切换功能时重置列宽应用标志，让新功能能正确加载自己的列宽
        self._data_table_column_widths_applied = False

        # 更新状态显示
        for option in self.gui_options:
            if option['value'] == value:
                operation_name = option['text']
                break
        else:
            operation_name = "未知操作"

        status_text = f"已选择: {operation_name}"
        self.append_output(status_text)

        # 加载Excel数据到表格（使用缓存，不强制刷新）
        self.load_excel_data_to_table(value, force_reload=False)


    def on_refresh_data(self):
        """刷新按钮 - 重新加载Excel数据（强制刷新，清除缓存）"""
        self._reset_contract_filters_and_selection()
        if hasattr(self, 'selected_operation') and self.selected_operation:
            # 清除该功能的缓存
            cache_key = str(self.selected_operation)
            if cache_key in getattr(self, '_data_cache', {}):
                del self._data_cache[cache_key]
            self.load_excel_data_to_table(self.selected_operation, force_reload=True)
            self.append_output("数据已刷新")


    def _reset_contract_filters_and_selection(self):
        """刷新文件生成时重置搜索、筛选和勾选状态"""
        if hasattr(self, '_search_timer') and self._search_timer.isActive():
            self._search_timer.stop()
        if hasattr(self, '_date_filter_timer') and self._date_filter_timer.isActive():
            self._date_filter_timer.stop()

        if hasattr(self, 'data_search_input'):
            self.data_search_input.blockSignals(True)
            self.data_search_input.clear()
            self.data_search_input.blockSignals(False)

        self.selected_filter_start_date = None
        self.selected_filter_end_date = None
        if hasattr(self, 'date_filter_button'):
            self.date_filter_button.setText("请选择日期范围")
        self._clear_all_data_conditions()
        self._refresh_data_exposed_tags()

        operation_value = getattr(self, 'selected_operation', None)
        if operation_value is not None:
            self._selected_row_ids_by_operation[str(operation_value)] = set()

        if hasattr(self, 'data_table'):
            self.data_table.clearSelection()
        if hasattr(self, 'table_header'):
            self.table_header.set_check_state(Qt.CheckState.Unchecked)

        self.current_page = 1
        self.persist_contract_page_state()

    # ===== 文件生成多条件筛选 =====


    def _get_data_filter_headers(self):
        """获取可筛选的字段列表"""
        # 优先从 current_source_headers 获取
        headers = getattr(self, 'current_source_headers', [])
        if headers:
            return [str(h).strip() for h in headers if h and str(h).strip()]
        # 回退：从表格列头获取
        result = []
        if hasattr(self, 'data_table') and self.data_table.columnCount() > 0:
            for col in range(self.data_table.columnCount()):
                item = self.data_table.horizontalHeaderItem(col)
                if item:
                    txt = item.text().strip()
                    if txt:
                        result.append(txt)
        return result


    def _is_data_date_field(self, field_label):
        """判断是否为日期/时间字段（模糊匹配：包含 日期/时间）"""
        if not field_label:
            return False
        label = str(field_label)
        return '日期' in label or '时间' in label


    def _toggle_data_filter_panel(self):
        """切换文件生成筛选面板（弹窗形式）"""
        if not hasattr(self, 'data_filter_panel'):
            return
        visible = not self.data_filter_panel.isVisible()
        if visible:
            self.data_filter_panel.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.FramelessWindowHint)
            self.data_filter_panel.setVisible(True)
            self.data_filter_panel.raise_()
            self.data_filter_panel.activateWindow()
            # 更新字段下拉选项，调整面板高度
            self._refresh_data_search_field_combo()
            self._adjust_data_filter_panel_size()
            # 定位在搜索框下方
            search_frame = self.data_search_stack.parent()
            panel_w = self.data_filter_panel.width()
            panel_h = self.data_filter_panel.height()
            if search_frame:
                pos = search_frame.mapToGlobal(QPoint(0, search_frame.height()))
                x = pos.x()
                y = pos.y() + 4
            else:
                x = self.geometry().x() + 100
                y = self.geometry().y() + 80
            screen = self.screen()
            if screen:
                geo = screen.availableGeometry()
                x = max(geo.x() + 10, min(x, geo.right() - panel_w - 10))
                y = max(geo.y() + 10, min(y, geo.bottom() - panel_h - 10))
            self.data_filter_panel.move(x, y)
            # 外部点击关闭
            self.data_filter_panel.reject = lambda: self.data_filter_panel.setVisible(False)
            self.data_filter_panel._outside_close_armed = False
            pf = common._DialogOutsideCloseFilter(self.data_filter_panel)
            self.data_filter_panel._outside_filter = pf
            QApplication.instance().installEventFilter(pf)
            QTimer.singleShot(0, lambda p=self.data_filter_panel: setattr(p, '_outside_close_armed', True))
            self.data_filter_panel.destroyed.connect(lambda obj, f=pf: QApplication.instance().removeEventFilter(f))
        else:
            self.data_filter_panel.setVisible(False)
        self._update_data_filter_toggle_btn()


    def _refresh_data_search_field_combo(self):
        """[兼容层] 搜索字段下拉已移除，保留方法避免调用错误。"""
        pass


    def _update_data_search_field(self):
        """[兼容层] 搜索字段下拉已移除，保留方法避免调用错误。"""
        pass


    def _open_data_search_date_range(self):
        """打开搜索栏日期范围选择"""
        btn = self.data_search_date_btn
        dr = btn.property('date_range') or {'start': None, 'end': None}
        dlg = QuickDatePickerDialog(start_date=dr.get('start'), end_date=dr.get('end'), parent=self)
        dlg.set_popup_anchor_widget(btn)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.start_date and dlg.end_date:
            new_range = {'start': dlg.start_date, 'end': dlg.end_date}
            btn.setProperty('date_range', new_range)
            btn.setText(f"{dlg.start_date.toString('yyyy-MM-dd')} ~ {dlg.end_date.toString('yyyy-MM-dd')}")
            self.apply_search_filter_now()


    def _update_data_filter_toggle_btn(self):
        """更新筛选按钮计数"""
        count = len(getattr(self, 'data_condition_rows', []))
        if count > 0:
            self.data_filter_toggle_btn.setText(f"筛选({count})")
        else:
            self.data_filter_toggle_btn.setText("筛选")


    def _add_data_condition_row(self, condition=None):
        """Add a filter condition row (delegated to FilterPanel)."""
        if hasattr(self, "_data_filter_panel"):
            return self._data_filter_panel.add_row(condition)
        return None

    def _update_data_condition_input_mode(self, row_info):
        """No-op: input mode is automatically handled by FilterPanel."""
        pass

    def _open_data_date_range(self, btn):
        """打开日期范围选择弹窗"""
        dr = btn.property('date_range') or {'start': None, 'end': None}
        dlg = QuickDatePickerDialog(start_date=dr.get('start'), end_date=dr.get('end'), parent=self)
        dlg.set_popup_anchor_widget(btn)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.start_date and dlg.end_date:
            new_range = {'start': dlg.start_date, 'end': dlg.end_date}
            btn.setProperty('date_range', new_range)
            btn.setText(f"{dlg.start_date.toString('yyyy-MM-dd')} ~ {dlg.end_date.toString('yyyy-MM-dd')}")
            self._apply_data_filters()


    def _adjust_data_filter_panel_size(self):
        """动态调整筛选面板高度"""
        if not hasattr(self, 'data_filter_panel'):
            return
        row_count = len(getattr(self, 'data_condition_rows', []))
        h = 36 + row_count * 32 + max(0, row_count - 1) * 4 + 44 + 24
        h = max(140, min(h, 500))
        self.data_filter_panel.setFixedHeight(h)


    def _apply_data_filter_and_close(self):
        """应用筛选并关闭弹窗"""
        self._apply_data_filters()
        self._refresh_data_exposed_tags()
        if hasattr(self, 'data_filter_panel'):
            self.data_filter_panel.setVisible(False)
        self._update_data_filter_toggle_btn()


    def _remove_data_condition_row(self, row_info):
        """Remove a filter condition row (delegated to FilterPanel)."""
        if hasattr(self, "_data_filter_panel"):
            if isinstance(row_info, dict) and "row" in row_info:
                self._data_filter_panel.remove_row(row_info["row"])
            else:
                self._data_filter_panel.remove_row(row_info)

    def _collect_data_filter_conditions(self):
        """Collect all filter conditions (delegated to FilterPanel)."""
        if hasattr(self, "_data_filter_panel"):
            return self._data_filter_panel.get_all_conditions()
        return []

    def _apply_data_filters(self):
        """条件变化时重新刷新表格"""
        data_rows = getattr(self, '_current_raw_data_rows', None)
        headers = getattr(self, 'current_source_headers', None)
        if not data_rows or not headers:
            return
        self.current_page = 1
        self._update_data_table(headers, data_rows)
        self._update_pagination_controls()


    def _clear_all_data_conditions(self):
        """Clear all filter conditions (delegated to FilterPanel)."""
        if hasattr(self, "_data_filter_panel"):
            self._data_filter_panel.clear_all()
    def _show_data_preset_popup(self):
        """弹出方案选择面板"""
        presets = self.config.get('file_gen_filter_presets', [])
        if not isinstance(presets, list):
            presets = []
        default_name = self.config.get('file_gen_filter_default_preset', '')

        popup = QFrame()
        popup.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        popup.setStyleSheet("QFrame { border: 1px solid #D9D9D9; border-radius: 4px; background: #FFF; }")
        popup_layout = QVBoxLayout(popup)
        popup_layout.setContentsMargins(4, 4, 4, 4)
        popup_layout.setSpacing(2)

        none_btn = QPushButton("(未选择)")
        none_btn.setFixedHeight(28)
        none_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        none_btn.setStyleSheet("QPushButton { border: none; text-align: left; padding: 4px 8px; font-size: 12px; color: #999; background: transparent; } QPushButton:hover { background: #F5F5F5; }")
        none_btn.clicked.connect(lambda: [self._clear_all_data_conditions(), self._apply_data_filters(), popup.close()])
        popup_layout.addWidget(none_btn)

        icon_style = """
            QPushButton { border: none; border-radius: 2px; font-size: 13px; background: transparent; color: #999; }
            QPushButton:hover { background: #E6F7FF; color: #1890FF; }
        """
        for preset in presets:
            name = preset.get('name', '')
            is_default = (name == default_name)
            display = f"★ {name}" if is_default else name
            row = QFrame()
            row.setFixedHeight(30)
            row.setStyleSheet("QFrame { border: none; background: transparent; }")
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(4)

            load_btn = QPushButton(display)
            load_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            if is_default:
                load_btn.setStyleSheet("QPushButton { border: none; text-align: left; padding: 4px 8px; font-size: 12px; color: #333; background: transparent; } QPushButton:hover { background: #E6F7FF; color: #1890FF; }")
            else:
                load_btn.setStyleSheet("QPushButton { border: none; text-align: left; padding: 4px 8px; font-size: 12px; color: #333; background: transparent; } QPushButton:hover { background: #E6F7FF; color: #1890FF; }")
            load_btn.clicked.connect(lambda checked, n=name: [self._load_data_filter_preset(n), popup.close()])
            row_layout.addWidget(load_btn, stretch=1)

            def_btn = QPushButton("★")
            def_btn.setFixedSize(22, 22)
            if is_default:
                def_btn.setStyleSheet("QPushButton { border: none; border-radius: 2px; font-size: 13px; background: transparent; color: #FA8C16; } QPushButton:hover { background: #FFF7E6; color: #FA8C16; }")
            else:
                def_btn.setStyleSheet(icon_style)
            def_btn.setToolTip(f"设置/取消「{name}」为默认方案")
            def_btn.clicked.connect(lambda checked, n=name: [self._toggle_data_filter_default_preset(n), popup.close()])
            row_layout.addWidget(def_btn)

            del_btn = QPushButton("✕")
            del_btn.setFixedSize(22, 22)
            del_btn.setStyleSheet(icon_style + "QPushButton:hover { background: #FFF2F0; color: #FF4D4F; }")
            del_btn.setToolTip(f"删除「{name}」")
            del_btn.clicked.connect(lambda checked, n=name: [self._delete_data_filter_preset(n), popup.close()])
            row_layout.addWidget(del_btn)
            popup_layout.addWidget(row)

        popup.adjustSize()
        btn_pos = self.data_preset_btn.mapToGlobal(QPoint(0, self.data_preset_btn.height()))
        popup.move(btn_pos.x(), btn_pos.y() + 2)
        popup.reject = popup.close
        popup._outside_close_armed = False
        pf = common._DialogOutsideCloseFilter(popup)
        QApplication.instance().installEventFilter(pf)
        QTimer.singleShot(0, lambda p=popup: setattr(p, '_outside_close_armed', True))
        popup.destroyed.connect(lambda obj, f=pf: QApplication.instance().removeEventFilter(f))
        popup.show()


    def _load_data_filter_preset(self, name):
        """加载保存的筛选方案"""
        presets = self.config.get('file_gen_filter_presets', [])
        if not isinstance(presets, list):
            return
        target = next((p for p in presets if p.get('name') == name), None)
        if not target:
            return
        conditions = target.get('conditions', [])
        self._clear_all_data_conditions()
        for cond in conditions:
            self._add_data_condition_row(cond)
        self._apply_data_filters()


    def _delete_data_filter_preset(self, name):
        """删除筛选方案"""
        presets = self.config.get('file_gen_filter_presets', [])
        if not isinstance(presets, list):
            return
        reply = QMessageBox.question(self, '确认删除', f'确定删除方案「{name}」吗？')
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.config['file_gen_filter_presets'] = [p for p in presets if p.get('name') != name]
        if self.config.get('file_gen_filter_default_preset') == name:
            self.config['file_gen_filter_default_preset'] = ''
        save_config(self.config)


    def _toggle_data_filter_default_preset(self, preset_name):
        """切换默认方案：已是默认则取消，否则设为默认"""
        if not preset_name:
            return
        current_default = self.config.get('file_gen_filter_default_preset', '')
        if current_default == preset_name:
            self.config['file_gen_filter_default_preset'] = ''
        else:
            presets = self.config.get('file_gen_filter_presets', [])
            if not isinstance(presets, list) or not any(p.get('name') == preset_name for p in presets):
                return
            self.config['file_gen_filter_default_preset'] = preset_name
        save_config(self.config)


    def _save_data_filter_preset(self):
        """保存文件生成筛选条件为方案（存储在 config['file_gen_filter_presets']）"""
        conditions = self._collect_data_filter_conditions()
        name, ok = QInputDialog.getText(self, "保存筛选方案", "请输入方案名称：")
        if not ok or not name or not name.strip():
            return
        name = name.strip()
        presets = self.config.get('file_gen_filter_presets', [])
        if not isinstance(presets, list):
            presets = []
        existing = next((i for i, p in enumerate(presets) if p.get('name') == name), None)
        if existing is not None:
            reply = QMessageBox.question(self, '确认覆盖', f'方案「{name}」已存在，是否覆盖？')
            if reply != QMessageBox.StandardButton.Yes:
                return
            presets[existing] = {'name': name, 'conditions': conditions}
        else:
            presets.append({'name': name, 'conditions': conditions})
        self.config['file_gen_filter_presets'] = presets
        save_config(self.config)
        QMessageBox.information(self, "成功", f"方案「{name}」已保存。")


    def _refresh_data_exposed_tags(self):
        """Refresh exposed tags (delegated to FilterPanel)."""
        if hasattr(self, "_data_filter_panel"):
            self._data_filter_panel.refresh_exposed_tags()
    def _clear_data_table(self, message="请选择功能后加载数据"):
        """清空数据表格"""
        self._is_updating_data_table = True

        # ✅ 保存当前列宽（如果有）
        current_widths = {}
        if hasattr(self, 'data_table') and self.data_table.columnCount() > 0:
            for col in range(self.data_table.columnCount()):
                current_widths[col] = self.data_table.columnWidth(col)

        self.data_table.clear()
        self.data_table.setColumnCount(0)
        self.data_table.setRowCount(0)

        # ✅ 如果之前有列宽且现在列数为0，暂存起来用于后续恢复
        if current_widths and not hasattr(self, '_saved_data_table_widths'):
            self._saved_data_table_widths = current_widths

        self._is_updating_data_table = False
        self.table_raw_data = []
        self.current_all_headers = []
        self.current_visible_headers = []
        self.current_filtered_rows = []
        self.current_source_headers = []
        self.current_header_source_index_map = {}
        self.current_page = 1
        self.data_header_label.setText(message)
        if hasattr(self, 'table_header'):
            self.table_header.set_check_state(Qt.CheckState.Unchecked)
        self._update_pagination_controls()
        self._update_data_table_sort_indicator()


    def _get_excel_table_settings(self, operation_value):
        """读取当前功能的表格字段设置"""
        if not operation_value:
            return {}
        config = getattr(self, 'config', None) or load_config() or {}
        return (
            config.get('app_settings', {})
            .get('excel_table_settings', {})
            .get(str(operation_value), {})
        )


    def _save_excel_table_settings(self, operation_value, visible_headers=None, header_order=None):
        """保存当前功能的表格字段设置"""
        if not operation_value:
            return

        # ✅ 使用 load_config() 获取最新配置，避免使用旧的 self.config 缓存
        config = load_config()
        app_settings = config.setdefault('app_settings', {})
        table_settings = app_settings.setdefault('excel_table_settings', {})
        operation_settings = table_settings.setdefault(str(operation_value), {})

        changed = False
        if visible_headers is not None and operation_settings.get('visible_headers') != visible_headers:
            operation_settings['visible_headers'] = visible_headers
            changed = True
        if header_order is not None and operation_settings.get('header_order') != header_order:
            operation_settings['header_order'] = header_order
            changed = True

        if changed:
            save_config(config, immediate=False)
            # 同步更新 self.config
            self.config = config


    def _compose_full_header_order(self, visible_headers):
        """根据当前完整表头补齐隐藏字段顺序"""
        all_headers = list(getattr(self, 'current_all_headers', []))
        return list(visible_headers) + [header for header in all_headers if header not in visible_headers]


    def _build_sort_key(self, value):
        """将不同类型的值标准化为可比较的排序键"""
        if value is None:
            return None

        if isinstance(value, datetime):
            return (1, value.timestamp())

        if isinstance(value, (int, float)):
            return (0, float(value))

        text = str(value).strip()
        if not text:
            return None

        numeric_text = text.replace(',', '')
        if numeric_text.endswith('%'):
            numeric_text = numeric_text[:-1]
        if re.fullmatch(r'-?\d+(?:\.\d+)?', numeric_text):
            return (0, float(numeric_text))

        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return (1, datetime.strptime(text, fmt).timestamp())
            except ValueError:
                continue

        return (2, text.casefold())


    def _sort_records(self, records, value_getter, sort_order):
        """按指定取值函数对记录排序，空值始终排在末尾"""
        sortable_records = []
        empty_records = []

        for record in records:
            sort_key = self._build_sort_key(value_getter(record))
            if sort_key is None:
                empty_records.append(record)
            else:
                sortable_records.append((sort_key, record))

        sortable_records.sort(
            key=lambda item: item[0],
            reverse=(sort_order == Qt.SortOrder.DescendingOrder)
        )
        return [record for _, record in sortable_records] + empty_records


    def _get_excel_sort_value(self, entry, header_label, source_index_map):
        """内部方法：获取Excel排序值。"""
        row_id, row_data = entry
        source_index = source_index_map.get(header_label, -1)
        if 0 <= source_index < len(row_data):
            return row_data[source_index]
        return ''


    def _sort_excel_data_entries(self, data_entries, visible_headers, source_index_map):
        """内部方法：处理排序Excel数据entries逻辑。"""
        sort_header = str(getattr(self, 'data_sort_header', '') or '').strip()
        if not sort_header:
            return list(data_entries)
        if sort_header not in visible_headers:
            self.data_sort_header = ''
            return list(data_entries)

        return self._sort_records(
            list(data_entries),
            lambda entry: self._get_excel_sort_value(entry, sort_header, source_index_map),
            getattr(self, 'data_sort_order', Qt.SortOrder.AscendingOrder)
        )


    def _update_data_table_sort_indicator(self):
        """内部方法：更新数据表格排序indicator。"""
        if not hasattr(self, 'table_header'):
            return

        sort_header = str(getattr(self, 'data_sort_header', '') or '').strip()
        visible_headers = list(getattr(self, 'current_visible_headers', []))
        if not sort_header or sort_header not in visible_headers:
            self.table_header.setSortIndicatorShown(False)
            return

        logical_index = visible_headers.index(sort_header) + 1
        self.table_header.setSortIndicatorShown(True)
        self.table_header.setSortIndicator(logical_index, getattr(self, 'data_sort_order', Qt.SortOrder.AscendingOrder))


    def on_data_table_header_clicked(self, logical_index):
        """响应数据表头clicked相关操作。"""
        if logical_index <= 0 or logical_index >= self.data_table.columnCount():
            return

        header_item = self.data_table.horizontalHeaderItem(logical_index)
        if header_item is None:
            return

        header_text = header_item.text().strip()
        if not header_text:
            return

        if getattr(self, 'data_sort_header', '') == header_text:
            if getattr(self, 'data_sort_order', Qt.SortOrder.AscendingOrder) == Qt.SortOrder.AscendingOrder:
                self.data_sort_order = Qt.SortOrder.DescendingOrder
            else:
                self.data_sort_header = ''
                self.data_sort_order = Qt.SortOrder.AscendingOrder
        else:
            self.data_sort_header = header_text
            self.data_sort_order = Qt.SortOrder.AscendingOrder

        self.current_page = 1
        self._refresh_current_table_page()


    def _get_current_visual_headers(self):
        """读取当前表格表头的视觉顺序"""
        if not hasattr(self, 'data_table') or self.data_table.columnCount() <= 1:
            return list(getattr(self, 'current_visible_headers', []))

        header = self.data_table.horizontalHeader()
        visual_headers = []
        for visual_index in range(1, self.data_table.columnCount()):
            logical_index = header.logicalIndex(visual_index)
            header_item = self.data_table.horizontalHeaderItem(logical_index)
            if not header_item:
                continue
            header_text = header_item.text().strip()
            if header_text:
                visual_headers.append(header_text)

        return visual_headers or list(getattr(self, 'current_visible_headers', []))


    def on_data_table_header_moved(self, logical_index, old_visual_index, new_visual_index):
        """拖动表头后保存字段顺序"""
        if getattr(self, '_is_updating_data_table', False):
            return
        if logical_index <= 0 or old_visual_index == new_visual_index:
            return

        visual_headers = self._get_current_visual_headers()
        if not visual_headers:
            return

        self.current_visible_headers = visual_headers
        self._save_excel_table_settings(
            getattr(self, 'selected_operation', None),
            visible_headers=visual_headers,
            header_order=self._compose_full_header_order(visual_headers)
        )

        self._force_rebuild_data_table = True
        self._refresh_current_table_page()


    def _get_operation_selected_row_ids(self, operation_value=None):
        """获取当前功能勾选的行ID集合"""
        target_operation = operation_value if operation_value is not None else getattr(self, 'selected_operation', None)
        if target_operation is None:
            return set()

        cache_key = str(target_operation)
        if cache_key not in self._selected_row_ids_by_operation:
            self._selected_row_ids_by_operation[cache_key] = set()
        return self._selected_row_ids_by_operation[cache_key]


    def _get_current_page_size(self):
        """获取当前每页显示数量"""
        return self._get_combo_page_size(getattr(self, 'page_size_combo', None))


    def _get_total_pages(self):
        """计算总页数"""
        total_rows = len(getattr(self, 'current_filtered_rows', []))
        page_size = max(1, self._get_current_page_size())
        return max(1, (total_rows + page_size - 1) // page_size)


    def _update_pagination_controls(self):
        """更新分页控件状态"""
        if not hasattr(self, 'page_label'):
            return

        total_pages = self._get_total_pages()
        if getattr(self, 'current_page', 1) < 1:
            self.current_page = 1
        if self.current_page > total_pages:
            self.current_page = total_pages

        self.page_label.setText(f"{self.current_page}/{total_pages}")

        if hasattr(self, 'prev_page_btn'):
            self.prev_page_btn.setEnabled(self.current_page > 1)
        if hasattr(self, 'next_page_btn'):
            self.next_page_btn.setEnabled(self.current_page < total_pages)


    def on_page_size_changed(self, index):
        """切换每页显示数量"""
        if index < 0:
            return
        self.current_page = 1
        self._save_page_sizes()
        self.persist_contract_page_state()
        self._refresh_current_table_page()


    def on_page_spin_changed(self, page):
        """切换页码"""
        if self._is_updating_data_table:
            return
        self.current_page = max(1, page)
        self.persist_contract_page_state()
        self._refresh_current_table_page()


    def on_prev_page(self):
        """上一页"""
        if self.current_page > 1:
            self.current_page -= 1
            self._refresh_current_table_page()


    def on_next_page(self):
        """下一页"""
        total_pages = self._get_total_pages()
        if self.current_page < total_pages:
            self.current_page += 1
            self._refresh_current_table_page()


    def on_row_checkbox_toggled(self, row_id, checked):
        """同步跨页勾选状态"""
        if self._is_updating_data_table or row_id is None or not getattr(self, 'selected_operation', None):
            return

        selected_ids = self._get_operation_selected_row_ids()
        if checked:
            selected_ids.add(row_id)
        else:
            selected_ids.discard(row_id)

        self._update_header_checkbox_state()
        self._update_data_header_summary()


    def on_data_table_item_changed(self, item):
        """同步主表首列复选框和跨页勾选状态"""
        if (
            self._is_updating_data_table
            or item is None
            or item.column() != 0
            or not getattr(self, 'selected_operation', None)
        ):
            return

        row_id = item.data(Qt.ItemDataRole.UserRole)
        if row_id is None:
            return

        selected_ids = self._get_operation_selected_row_ids()
        if item.checkState() == Qt.CheckState.Checked:
            selected_ids.add(row_id)
        else:
            selected_ids.discard(row_id)

        self._update_header_checkbox_state()
        self._update_data_header_summary()


    def _update_header_checkbox_state(self):
        """根据当前页面显示结果更新表头复选框状态"""
        if not hasattr(self, 'table_header'):
            return

        current_page_ids = [row_id for row_id, _ in getattr(self, '_current_page_rows', [])]
        if not current_page_ids:
            self.table_header.set_check_state(Qt.CheckState.Unchecked)
            return

        selected_ids = self._get_operation_selected_row_ids()
        checked_count = sum(1 for row_id in current_page_ids if row_id in selected_ids)
        if checked_count == 0:
            state = Qt.CheckState.Unchecked
        elif checked_count == len(current_page_ids):
            state = Qt.CheckState.Checked
        else:
            state = Qt.CheckState.PartiallyChecked
        self.table_header.set_check_state(state)


    def _update_data_header_summary(self):
        """更新数据区摘要文本"""
        if not hasattr(self, 'data_header_label'):
            return

        total_rows = len(getattr(self, 'current_filtered_rows', []))
        visible_headers = list(getattr(self, 'current_visible_headers', []))
        selected_row_ids = self._get_operation_selected_row_ids()
        self.data_header_label.setText(
            f"已加载 {total_rows} 条记录，显示 {len(visible_headers)} 个字段，已勾选 {len(selected_row_ids)} 条"
        )


    def on_data_table_cell_clicked(self, row, col):
        """文件生成表格单元格点击处理：第0列（复选框）选中整行，其他列只选中单元格"""
        if col == 0:
            self.data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            self.data_table.selectRow(row)
            self.data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)


    def on_header_select_all_toggled(self, checked):
        """点击表头复选框，全选或全不选当前页面显示结果"""
        selected_ids = self._get_operation_selected_row_ids()
        current_page_ids = [row_id for row_id, _ in getattr(self, '_current_page_rows', [])]
        if checked:
            selected_ids.update(current_page_ids)
        else:
            for row_id in current_page_ids:
                selected_ids.discard(row_id)

        self._is_updating_data_table = True
        try:
            for row in range(self.data_table.rowCount()):
                checkbox_container = self.data_table.cellWidget(row, 0)
                checkbox = checkbox_container.findChild(QCheckBox) if checkbox_container else None
                if checkbox:
                    checkbox.setChecked(checked)
        finally:
            self._is_updating_data_table = False

        self.data_table.viewport().update()
        self._update_header_checkbox_state()
        self._update_data_header_summary()


    def on_global_search_changed(self, text):
        """搜索框输入变化 → 重新筛选"""
        self._apply_data_filters()


    def _do_data_search(self):
        self._apply_data_filters()


    def apply_search_filter_now(self):
        """应用全局搜索"""
        self._apply_data_filters()


    def load_excel_data_to_table(self, operation_value, force_reload=False):
        """
        加载选中功能的Excel数据到表格（支持公式计算结果和缓存机制）

        参数:
            operation_value: 功能值
            force_reload: 是否强制重新加载（忽略缓存），默认False
        """
        try:
            config = getattr(self, 'config', None) or load_config()
            gui_options = config.get('gui_options', [])
            excel_paths = config.get('path_config', {}).get('excel_paths', {})

            selected_option = None
            for opt in gui_options:
                if opt.get('value') == operation_value:
                    selected_option = opt
                    break

            if not selected_option:
                self.data_header_label.setText("未找到功能配置")
                self.data_table.setRowCount(0)
                return

            excel_path = AppConfig.get_excel_template_path(operation_value)
            if not excel_path or not excel_path.exists():
                self.data_header_label.setText(f"Excel文件不存在: {excel_path}")
                self.data_table.setRowCount(0)
                return

            # ===== 缓存机制 =====
            cache_key = str(operation_value)

            # 检查是否有有效缓存
            if not force_reload and cache_key in getattr(self, '_data_cache', {}):
                cached_headers, cached_rows = self._data_cache[cache_key]
                self.current_page = 1
                self._update_data_table(cached_headers, cached_rows)
                return

            # 缓存无效或强制刷新，从Excel加载
            from openpyxl import load_workbook
            # 使用data_only=True获取公式的计算结果
            workbook = load_workbook(excel_path, data_only=True, read_only=True)
            sheet_name = selected_option.get('sheet', '')

            if sheet_name and sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            data_rows = []
            for row in sheet.iter_rows(min_row=2):
                row_data = [cell.value for cell in row]

                # 过滤空行
                if all(cell is None or cell == '' for cell in row_data):
                    continue

                data_rows.append(row_data)

            workbook.close()

            # 保存到缓存（保存原始未筛选的数据）
            if not hasattr(self, '_data_cache'):
                self._data_cache = {}
            self._data_cache[cache_key] = (headers, data_rows.copy())

            self.current_page = 1
            self._update_data_table(headers, data_rows)

            # 首次加载数据后自动应用默认筛选方案
            if not getattr(self, '_data_default_preset_applied', False):
                self._data_default_preset_applied = True
                default_name = self.config.get('file_gen_filter_default_preset', '')
                if default_name:
                    self._load_data_filter_preset(default_name)

        except Exception as e:
            logging.error(f"加载数据失败: {str(e)}")
            self.data_header_label.setText(f"加载数据失败: {str(e)}")


    def _update_data_table(self, headers, data_rows):
        """更新数据表格显示（与 CRM _populate_crm_table 逻辑一致）"""
        self.current_source_headers = list(headers)
        self._current_raw_data_rows = list(data_rows)
        # 刷新搜索字段下拉框，加载表格列名
        self._refresh_data_search_field_combo()

        # 1. 日期筛选
        date_filtered_rows = self._apply_date_filter(headers, data_rows)

        # 2. 多条件筛选
        condition_filtered_rows = self._apply_condition_filter(headers, date_filtered_rows)

        # 3. 搜索筛选（全字段模糊搜索）
        if hasattr(self, 'data_search_stack') and self.data_search_stack.currentIndex() == 1:
            # 日期范围搜索
            dr = self.data_search_date_btn.property('date_range') or {}
            if dr.get('start') and dr.get('end'):
                start_str = dr['start'].toString('yyyy-MM-dd')
                end_str = dr['end'].toString('yyyy-MM-dd')
                # 日期范围搜索：在全字段中模糊匹配日期字符串
                filtered = []
                for entry in condition_filtered_rows:
                    row_data = entry[1] if isinstance(entry, tuple) else entry
                    for cell in row_data:
                        cell_str = str(cell or '')[:10]
                        if start_str <= cell_str <= end_str:
                            filtered.append(entry)
                            break
                condition_filtered_rows = filtered
            self.current_filtered_rows = condition_filtered_rows
        else:
            search = self.data_search_input.text().strip().lower() if hasattr(self, 'data_search_input') else ''
            if search:
                # 全字段搜索
                filtered = []
                for entry in condition_filtered_rows:
                    row_data = entry[1] if isinstance(entry, tuple) else entry
                    if any(search in str(v or '').lower() for v in row_data):
                        filtered.append(entry)
                condition_filtered_rows = filtered
            self.current_filtered_rows = condition_filtered_rows

        self._refresh_current_table_page()


    def _refresh_current_table_page(self):
        """按当前筛选条件和分页设置刷新表格页"""
        headers = list(getattr(self, 'current_source_headers', []))
        data_entries = list(getattr(self, 'current_filtered_rows', []))

        # ✅ 优先使用内存中的 current_visible_headers（避免从配置文件重新解析导致旧值）
        current_visible = getattr(self, 'current_visible_headers', None)
        if current_visible and len(current_visible) > 0:
            print(f"[DEBUG-文件生成] ✅ 使用内存中的visible_headers: {len(current_visible)}个字段")
            all_headers, ordered_headers, visible_headers, source_index_map = self._resolve_excel_table_headers(headers)
            visible_headers = [h for h in current_visible if h in all_headers]
            if not visible_headers:
                visible_headers = all_headers[1:] if len(all_headers) > 1 else all_headers
            ordered_headers = list(visible_headers) + [h for h in all_headers if h not in visible_headers]
            print(f"[DEBUG-文件生成] 实际使用visible_headers: {len(visible_headers)}个字段 | 列表: {visible_headers[:10]}")
        else:
            print(f"[DEBUG-文件生成] 内存中无visible_headers，从配置文件解析")
            all_headers, _, visible_headers, source_index_map = self._resolve_excel_table_headers(headers)

        data_entries = self._sort_excel_data_entries(data_entries, visible_headers, source_index_map)
        display_headers = [""] + visible_headers
        total_rows = len(data_entries)
        page_size = max(1, self._get_current_page_size())
        total_pages = max(1, (total_rows + page_size - 1) // page_size)
        if self.current_page > total_pages:
            self.current_page = total_pages
        if self.current_page < 1:
            self.current_page = 1

        start_index = (self.current_page - 1) * page_size
        end_index = start_index + page_size
        page_rows = data_entries[start_index:end_index]
        self._current_page_rows = list(page_rows)
        selected_row_ids = self._get_operation_selected_row_ids()

        # ✅ 检查是否需要重建表格结构（列数变化时才重建）
        force_rebuild = getattr(self, '_force_rebuild_data_table', False)
        need_rebuild = force_rebuild or (
            self.data_table.columnCount() != len(display_headers) or
            self.data_table.horizontalHeader() is None or
            len(display_headers) == 0
        )

        if need_rebuild:
            print(f"[DEBUG-列宽] 文件生成表格 | 需要重建 | _data_table_column_widths_applied={getattr(self, '_data_table_column_widths_applied', False)}")
            # ✅ 只在结构变化时才完全重建
            self._is_updating_data_table = True

            # 保存当前列宽（如果有）
            current_widths = {}
            if self.data_table.columnCount() > 0:
                for col in range(self.data_table.columnCount()):
                    current_widths[col] = self.data_table.columnWidth(col)

            self.data_table.clear()
            self.data_table.setColumnCount(len(display_headers))
            self.data_table.setHorizontalHeaderLabels(display_headers)
            self.data_table.setRowCount(len(page_rows))

            # ✅ 关键修复：重建表格后必须重新设置ResizeMode，否则clear()/setColumnCount()会重置为默认值
            header = self.data_table.horizontalHeader()
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
            for ci in range(1, len(display_headers)):
                header.setSectionResizeMode(ci, QHeaderView.ResizeMode.Interactive)
            header.setSectionsMovable(True)

            # ✅ 关键修复：重建表格后必须重新连接sectionResized信号！
            # 否则 setHorizontalHeaderLabels() 可能会导致原信号连接失效
            try:
                header.sectionResized.disconnect(self._on_data_table_column_resized)
            except Exception:
                pass  # 可能未连接
            header.sectionResized.connect(self._on_data_table_column_resized)
            try:
                header.sectionMoved.disconnect(self.on_data_table_header_moved)
            except Exception:
                pass
            header.sectionMoved.connect(self.on_data_table_header_moved)
            print(f"[DEBUG-列宽] [OK] 已重新连接sectionResized信号")

            config = load_config()  # 强制加载最新配置
            app_set = config.get('app_settings', {})
            data_set = app_set.get('data_table_settings', {})

            # ✅ 修复：优先从按功能选项保存的列宽中读取
            operation_value = getattr(self, 'selected_operation', None)
            saved_widths_raw = {}
            if operation_value is not None:
                by_operation = data_set.get('column_widths_by_operation', {})
                op_key = str(operation_value)
                if op_key in by_operation:
                    saved_widths_raw = by_operation[op_key]

            # 如果没有找到按功能选项保存的列宽，则使用通用位置（向后兼容）
            if not saved_widths_raw:
                saved_widths_raw = data_set.get('column_widths', {})

            # ✅ 关键修复：统一键类型为整数
            saved_widths = {}
            for k, v in saved_widths_raw.items():
                try:
                    saved_widths[int(k)] = v
                except (ValueError, TypeError):
                    pass

            print(f"[DEBUG-列宽] 文件生成表格 | 保存的列宽数量: {len(saved_widths)} | 显示列数: {len(display_headers)}")

            if saved_widths and 0 in saved_widths:
                self.data_table.setColumnWidth(0, saved_widths.get(0, 22))
            else:
                self.data_table.setColumnWidth(0, 22)

            for col_idx in range(1, len(display_headers)):
                # ✅ 如果已经应用过保存的列宽，只设置有保存值的列，不强制覆盖
                if getattr(self, '_data_table_column_widths_applied', False) and saved_widths and col_idx in saved_widths:
                    self.data_table.setColumnWidth(col_idx, saved_widths[col_idx])
                elif saved_widths and col_idx in saved_widths:
                    # 首次加载：使用保存的宽度
                    self.data_table.setColumnWidth(col_idx, saved_widths[col_idx])
                elif current_widths and col_idx in current_widths:
                    self.data_table.setColumnWidth(col_idx, current_widths[col_idx])
                else:
                    # 只在首次加载时设置默认值
                    if not getattr(self, '_data_table_column_widths_applied', False):
                        if col_idx == 1:
                            self.data_table.setColumnWidth(col_idx, 120)
                        elif col_idx <= 5:
                            self.data_table.setColumnWidth(col_idx, 100)
                        else:
                            self.data_table.setColumnWidth(col_idx, 80)

            print(f"[DEBUG-列宽] 文件生成表格 | 列宽设置完成")

            # ✅ 验证：打印实际列宽
            actual_widths = {}
            for ci in range(min(5, self.data_table.columnCount())):
                actual_widths[ci] = self.data_table.columnWidth(ci)
            print(f"[DEBUG-列宽] 验证-重建后前5列实际宽度: {actual_widths}")
        else:
            # ✅ 结构不变，只更新行数和数据内容（保持列宽不变）
            self._is_updating_data_table = True
            old_row_count = self.data_table.rowCount()
            new_row_count = len(page_rows)

            if new_row_count > old_row_count:
                self.data_table.setRowCount(new_row_count)
            elif new_row_count < old_row_count:
                for row in range(old_row_count - 1, new_row_count - 1, -1):
                    self.data_table.removeRow(row)

        self._force_rebuild_data_table = False

        self.table_raw_data = []
        self.current_all_headers = all_headers
        self.current_visible_headers = visible_headers
        self.current_header_source_index_map = dict(source_index_map)

        for row_idx, entry in enumerate(page_rows):
            row_id, row_data = entry
            checkbox_item = QTableWidgetItem()
            checkbox_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            checkbox_item.setData(Qt.ItemDataRole.UserRole, row_id)
            self.data_table.setItem(row_idx, 0, checkbox_item)

            checkbox_widget = TableRowCheckBox(row_id=row_id)
            checkbox_widget.blockSignals(True)
            checkbox_widget.setChecked(row_id in selected_row_ids)
            checkbox_widget.blockSignals(False)
            checkbox_widget.toggled_with_row_id.connect(self.on_row_checkbox_toggled)

            checkbox_container = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_container)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            checkbox_layout.addWidget(checkbox_widget)
            self.data_table.setCellWidget(row_idx, 0, checkbox_container)

            for col_idx, header in enumerate(visible_headers):
                source_index = source_index_map.get(header, -1)
                value = row_data[source_index] if 0 <= source_index < len(row_data) else None
                # 处理None值和公式结果
                if value is None:
                    display_value = ""
                elif isinstance(value, float):
                    # 处理浮点数，避免过长小数
                    if value == int(value):
                        display_value = str(int(value))
                    else:
                        display_value = f"{value:.2f}"
                else:
                    display_value = str(value)

                item = QTableWidgetItem(display_value)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.data_table.setItem(row_idx, col_idx + 1, item)

            row_dict = {}
            for header in all_headers:
                source_index = source_index_map.get(header, -1)
                row_dict[header] = row_data[source_index] if 0 <= source_index < len(row_data) else None
            self.table_raw_data.append(row_dict)

        self._is_updating_data_table = False
        self._update_data_header_summary()
        self._update_header_checkbox_state()
        self._update_pagination_controls()
        self._update_data_table_sort_indicator()

        # ✅ 延迟应用列宽 - 确保在所有其他操作完成后执行，不会被覆盖
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(10, self._delayed_apply_data_table_column_widths)


    def _get_active_filter_date_range(self):
        """获取当前日期范围，未选择时返回(None, None)"""
        start_date = getattr(self, 'selected_filter_start_date', None)
        end_date = getattr(self, 'selected_filter_end_date', None)
        if not start_date or not end_date:
            return None, None
        if not start_date.isValid() or not end_date.isValid():
            return None, None
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date


    def _apply_date_filter(self, headers, data_rows):
        """应用日期筛选，返回筛选后的数据行"""
        start_date, end_date = self._get_active_filter_date_range()
        if not start_date or not end_date or not headers:
            return list(enumerate(data_rows))

        date_fields = ['日期', '开始日期', '结束日期', '合同日期', '签署日期', '签订日期']
        date_field_indexes = [i for i, header in enumerate(headers) if header in date_fields]
        if not date_field_indexes:
            return list(enumerate(data_rows))

        filtered_rows = []
        for row_index, row_data in enumerate(data_rows):
            for index in date_field_indexes:
                if index < len(row_data):
                    cell_val = row_data[index]
                    if cell_val:
                        try:
                            if isinstance(cell_val, datetime):
                                cell_date = QDate(cell_val.year, cell_val.month, cell_val.day)
                            else:
                                cell_text = str(cell_val)[:10].replace("/", "-").replace(".", "-")
                                cell_date = QDate.fromString(cell_text, "yyyy-MM-dd")
                            if cell_date.isValid() and start_date <= cell_date <= end_date:
                                filtered_rows.append((row_index, row_data))
                                break
                        except Exception:
                            continue

        return filtered_rows


    def _apply_search_filter(self, headers, data_entries):
        """应用搜索过滤（全字段模糊搜索）"""
        if not hasattr(self, 'data_search_input'):
            return list(data_entries)

        # 日期范围搜索
        if hasattr(self, 'data_search_stack') and self.data_search_stack.currentIndex() == 1:
            dr = self.data_search_date_btn.property('date_range') or {}
            if not dr.get('start') or not dr.get('end'):
                return list(data_entries)
            start_str = dr['start'].toString('yyyy-MM-dd')
            end_str = dr['end'].toString('yyyy-MM-dd')
            # 日期范围搜索：在所有字段中匹配
            filtered = []
            for entry in data_entries:
                if isinstance(entry, tuple) and len(entry) == 2:
                    row_id, row_data = entry
                else:
                    row_data = entry
                for cell in row_data:
                    cell_str = str(cell or '')[:10]
                    if start_str <= cell_str <= end_str:
                        filtered.append(entry)
                        break
            return filtered

        # 文本搜索（全字段模糊搜索）
        keyword = self.data_search_input.text().strip().lower()
        if not keyword:
            return list(data_entries)

        filtered_entries = []
        for entry in data_entries:
            # 兼容 tuple (row_id, row_data) 和 list [cell1, cell2, ...] 格式
            if isinstance(entry, tuple) and len(entry) == 2:
                row_id, row_data = entry
            else:
                row_data = entry
            # 全字段搜索
            values = row_data
            if any(keyword in str(v or '').lower() for v in values):
                filtered_entries.append(entry)
        return filtered_entries


    def _trigger_date_filter_update(self):
        """延迟触发日期筛选，避免连续重绘"""
        if hasattr(self, '_date_filter_timer'):
            self._date_filter_timer.start(180)


    def apply_date_filter_now(self):
        """应用日期筛选到当前缓存数据"""
        if hasattr(self, 'selected_operation') and self.selected_operation:
            cache_key = str(self.selected_operation)
            if cache_key in getattr(self, '_data_cache', {}):
                cached_headers, cached_rows = self._data_cache[cache_key]
                self.current_page = 1
                self._update_data_table(cached_headers, cached_rows)
            else:
                self.load_excel_data_to_table(self.selected_operation)


    def on_date_filter_changed(self):
        """日期变化后延迟应用筛选"""
        self._trigger_date_filter_update()


    def get_selected_rows_from_table(self):
        """从缓存中获取用户勾选的全部行数据（支持跨页）"""
        selected_rows = []
        if not getattr(self, 'selected_operation', None):
            return selected_rows

        cache_key = str(self.selected_operation)
        selected_ids = self._get_operation_selected_row_ids()
        if cache_key in getattr(self, '_data_cache', {}) and selected_ids:
            _, cached_rows = self._data_cache[cache_key]
            header_source_index_map = dict(getattr(self, 'current_header_source_index_map', {}))
            for row_id in sorted(selected_ids):
                if 0 <= row_id < len(cached_rows):
                    row_data = cached_rows[row_id]
                    row_dict = {}
                    for header in self.current_all_headers:
                        source_index = header_source_index_map.get(header, -1)
                        row_dict[header] = row_data[source_index] if 0 <= source_index < len(row_data) else None
                    selected_rows.append(row_dict)
            return selected_rows

        if not hasattr(self, 'data_table') or not hasattr(self, 'table_raw_data'):
            return selected_rows

        for row in range(self.data_table.rowCount()):
            checkbox_container = self.data_table.cellWidget(row, 0)
            checkbox = checkbox_container.findChild(QCheckBox) if checkbox_container else None
            if checkbox and checkbox.isChecked():
                if row < len(self.table_raw_data):
                    selected_rows.append(self.table_raw_data[row])

        return selected_rows


    def on_confirm_operation(self):
        """确认执行选中的操作"""
        if self.selected_operation is None:
            messagebox.showwarning("提示", "请先选择一个操作类型！")
            return

        # 获取用户勾选的行数据
        selected_rows = self.get_selected_rows_from_table()

        if len(selected_rows) == 0:
            messagebox.showwarning("提示", "请在列表中选择要生成的合同！")
            return

        # 更新状态
        status_text = f"处理中... (已选择 {len(selected_rows)} 条记录)"
        self.append_output(status_text)

        # 在后台线程中执行处理，传递选中的行数据
        thread = threading.Thread(target=self.run_process_with_choice, args=(self.selected_operation, selected_rows))
        thread.daemon = True
        thread.start()

        # 禁用按钮
        self.confirm_btn.setEnabled(False)
        self.confirm_btn.setText("执行中>>>")


    def append_output(self, text):
        """追加文本到输出窗口"""
        append_user_operation_record(text)
        # 副窗口：排除数据加载日志和过程细节，仅显示生成结果
        t = str(text)
        _show_in_sub = True
        if t.startswith('[') or t.startswith('  ') or t.startswith('\t'):
            _show_in_sub = False
        elif any(t.startswith(p) for p in ('处理页面', '原文件名', '新文件名', '页面',
                                             '水印', '当前字体', '成功设置', '尝试加载',
                                             '字体 ', '使用备用', 'MediaBox', 'CropBox',
                                             '旋转角度', '选中的文件')):
            _show_in_sub = False
        def safe_append():
            if hasattr(self, 'output_text'):
                self.output_text.append(text)
                self.output_text.ensureCursorVisible()
            if hasattr(self, 'sub_output_text') and not self.output_visible and _show_in_sub:
                self.sub_output_text.append(text)
                self.sub_output_text.ensureCursorVisible()
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(0, safe_append)


    def on_clear_output(self):
       """清除输出窗口内容"""
       self.output_text.clear()
       if hasattr(self, 'sub_output_text'):
           self.sub_output_text.clear()

    def on_toggle_output(self):
        """切换输出窗口显隐"""
        self.output_visible = not getattr(self, 'output_visible', True)
        if hasattr(self, 'output_frame'):
            self.output_frame.setVisible(self.output_visible)
        if hasattr(self, 'sub_output_frame'):
            self.sub_output_frame.setVisible(not self.output_visible)
        if hasattr(self, 'toggle_output_btn'):
            self.toggle_output_btn.setText("隐藏" if self.output_visible else "显示")

   # ============== 后台任务状态栏处理 ==============


    def _on_task_status_changed(self, task_id, name, status):
        """任务状态变化时刷新状态栏"""
        self._refresh_task_bar()


    def _on_task_completed(self, task_id, success, message):
        """任务完成：刷新状态栏 + 处理特定任务的 UI 恢复"""
        self._refresh_task_bar()
        # MySQL 同步完成：恢复按钮和标签
        if 'mysql_sync' in task_id:
            self._on_mysql_sync_completed(task_id, success, message)
        # MySQL 连接测试完成
        if task_id == 'mysql_test_connection':
            self._on_mysql_test_completed(task_id, success, message)


    def _on_task_error_occurred(self, task_id, error_message):
        """任务异常"""
        self.append_output(f"[后台任务] {task_id} 失败: {error_message}")
        self._refresh_task_bar()


    def _refresh_task_bar(self):
        """根据当前活跃任务刷新状态栏显示"""
        count = self.task_manager.active_count
        names = self.task_manager.active_task_names()
        if count == 0:
            self.task_status_bar.setMaximumHeight(0)
            self.task_status_label.setText("")
            self.task_count_label.setText("")
            self.task_status_icon_label.setText("")
        else:
            self.task_status_bar.setMaximumHeight(80)
            frames = ["◐", "◓", "◑", "◒"]
            idx = (int(time.time() * 2)) % len(frames)
            summary = "、".join(names[:3])
            if count > 3:
                summary += f" 等{count}个任务"
            self.task_status_icon_label.setText(f"{frames[idx]} 后台任务")
            self.task_status_label.setText(summary)
            self.task_count_label.setText(f"{count} 个运行中")


    def _on_mysql_sync_completed(self, task_id, success, message):
        """MySQL 同步任务完成：恢复按钮和更新状态标签"""
        if hasattr(self, 'obj_query_sync_btn'):
            self.obj_query_sync_btn.setText("同步")
            self.obj_query_sync_btn.setEnabled(True)
        data = getattr(self, 'obj_query_all_data', [])
        if hasattr(self, 'obj_query_status_label') and data:
            self.obj_query_status_label.setText(f"✅ 已同步 {len(data)} 条到 MySQL")


    def _on_mysql_test_completed(self, task_id, success, message):
        """MySQL 连接测试完成：更新状态标签和按钮"""
        # 查找 SettingsDialog 实例以更新其 UI
        settings_dialog = None
        if hasattr(self, 'settings_dialog') and self.settings_dialog:
            settings_dialog = self.settings_dialog

        if settings_dialog is None:
            return

        # 恢复按钮
        if hasattr(settings_dialog, 'mysql_test_btn'):
            settings_dialog.mysql_test_btn.setEnabled(True)

        if not success:
            # 超时或连接失败
            error_text = message or '连接失败'
            if hasattr(settings_dialog, 'mysql_status_label'):
                settings_dialog.mysql_status_label.setText(f"❌ {error_text}")
                settings_dialog.mysql_status_label.setStyleSheet(
                    "QLabel { color: #d13438; padding: 8px; font-size: 13px; }")
            return

        result = getattr(settings_dialog, '_mysql_test_result', None)
        if not result:
            return
        if hasattr(settings_dialog, 'mysql_status_label'):
            if result['success']:
                settings_dialog.mysql_status_label.setText(
                    f"✅ 连接成功！服务器: {result['host']}:{result['port']}\n"
                    f"数据库: {result['database']} | MySQL版本: {result['version']}")
                settings_dialog.mysql_status_label.setStyleSheet(
                    "QLabel { color: #107c10; padding: 8px; font-size: 13px; }")
            else:
                settings_dialog.mysql_status_label.setText(f"❌ 连接失败: {result['error']}")
                settings_dialog.mysql_status_label.setStyleSheet(
                    "QLabel { color: #d13438; padding: 8px; font-size: 13px; }")

    # ============== 后台任务状态栏处理结束 ==============


    def on_toggle_output(self):
        """切换输出窗口的显示/隐藏状态"""
        self.output_visible = not self.output_visible
        self.save_user_runtime_state_patch({'output_visible': self.output_visible}, immediate=True)

        if self.output_visible:
            self.output_frame.show()
            self.toggle_output_btn.setText("隐藏")
            if hasattr(self, 'sub_output_frame'):
                self.sub_output_frame.hide()
        else:
            self.output_frame.hide()
            self.toggle_output_btn.setText("显示")
            if hasattr(self, 'sub_output_frame'):
                self.sub_output_frame.show()


    def update_operation_buttons(self, preferred_value=None):
        """更新主界面的功能选项下拉框"""
        # 从配置文件中重新加载操作选项
        config = load_config()
        self.config = config
        gui_options = config.get('gui_options', [
            {"text": "SAAS合同", "value": 1},
            {"text": "招投标授权", "value": 2},
            {"text": "服务协议", "value": 3},
            {"text": "CRM合同", "value": 4},
            {"text": "经销协议", "value": 5}
        ])

        # 过滤掉PDF文件整理选项
        gui_options = [option for option in gui_options if option.get('text', '') != 'PDF文件整理']

        # 按value排序，确保顺序一致
        gui_options.sort(key=lambda x: x['value'])

        # 更新实例变量
        self.gui_options = gui_options
        self.operation_map = {i: option['value'] for i, option in enumerate(gui_options)}

        # 更新下拉框
        if hasattr(self, 'operation_combo'):
            # 阻止信号触发
            self.operation_combo.blockSignals(True)

            # 记录当前选中的值
            current_value = preferred_value if preferred_value is not None else self.operation_combo.currentData()

            # 清空并重新填充
            self.operation_combo.clear()

            self.operation_combo.addItem("请选择功能", None)

            for option in gui_options:
                text = option.get("text", "")
                value = option.get("value", 0)
                self.operation_combo.addItem(text, value)

            # 尝试恢复之前的选择
            if current_value is not None:
                index = self.operation_combo.findData(current_value)
                if index >= 0:
                    self.operation_combo.setCurrentIndex(index)
                    self.selected_operation = current_value
                elif self.operation_combo.count() > 0:
                    self.operation_combo.setCurrentIndex(0)
                    self.selected_operation = self.operation_combo.currentData()
            elif self.operation_combo.count() > 0:
                self.operation_combo.setCurrentIndex(0)
                self.selected_operation = self.operation_combo.currentData()

            # 恢复信号
            self.operation_combo.blockSignals(False)

            self.persist_contract_page_state()


    def run_process_with_choice(self, selected_choice, selected_rows=None):
        """在后台线程中执行处理，使用预选择的选项和选中的行数据"""
        try:
            global choice
            choice = selected_choice

            # 清空缓存文件夹
            self.update_output.emit("清空缓存文件夹...")
            print("清空缓存文件夹...")
            try:
                # 加载配置获取缓存文件夹路径
                config = load_config()
                log_dir = config.get('path_config', {}).get('log_dir', 'Cache')
                cache_path = Path(log_dir)
                if cache_path.exists() and cache_path.is_dir():
                    deleted_count = 0
                    skipped_count = 0
                    for f in cache_path.iterdir():
                        if f.is_file():
                            try:
                                f.unlink()
                                deleted_count += 1
                            except Exception as e:
                                print(f"跳过被占用的文件: {f.name} - {str(e)}")
                                skipped_count += 1
                    if deleted_count > 0:
                        self.update_output.emit(f"缓存文件夹已清空")
                        print(f"缓存文件夹已清空，删除了 {deleted_count} 个文件，跳过了 {skipped_count} 个被占用的文件: {cache_path}")
                    else:
                        self.update_output.emit(f"缓存文件夹为空或所有文件都被占用: {cache_path}")
                        print(f"缓存文件夹为空或所有文件都被占用: {cache_path}")
                else:
                    self.update_output.emit(f"缓存文件夹不存在: {cache_path}")
                    print(f"缓存文件夹不存在: {cache_path}")
            except Exception as e:
                self.update_output.emit(f"清空缓存文件夹失败: {str(e)}")
                print(f"清空缓存文件夹失败: {str(e)}")

            config = AppConfig()
            config.get_target_root(choice).mkdir(parents=True, exist_ok=True)

            # 核心处理流程
            self.update_output.emit("开始处理Excel模板...")
            # 检查Excel模板是否存在
            excel_template_path = AppConfig.get_excel_template_path(choice)
            if excel_template_path.exists():
                self.update_output.emit(f"Excel模板存在，开始处理: {excel_template_path}")
            else:
                self.update_output.emit(f"错误: Excel模板不存在，请检查路径: {excel_template_path}")
            process_excel_template(choice)
            self.update_output.emit("Excel模板处理完成")

            self.update_output.emit("开始生成合同...")
            generated_contracts = generate_contracts(AppConfig.get_field_mapping(), selected_rows)
            self.update_output.emit(f"✓ 合同生成完成（共 {len(generated_contracts)} 个）")

            self.update_output.emit("开始转换为PDF格式...")
            import os
            output_dir = AppConfig.OUTPUT_DIR
            pdf_count = 0
            word_count = 0
            for contract_name in generated_contracts:
                word_file = output_dir / f"{contract_name}{AppConfig.FILE_EXTENSION}"
                pdf_file = output_dir / f"{contract_name}.pdf"
                if pdf_file.exists() and pdf_file.stat().st_size > 0:
                    pdf_count += 1
                if word_file.exists():
                    word_count += 1

            if pdf_count > 0:
                self.update_output.emit(f"✓ PDF转换完成（{pdf_count}/{word_count} 个成功）")
            else:
                self.update_output.emit(f"⚠ PDF转换未成功，请检查：")
                self.update_output.emit("   1. 是否安装了Microsoft Word或LibreOffice")
                self.update_output.emit("   2. 是否安装了comtypes库（pip install comtypes）")
                self.update_output.emit("   3. 查看控制台日志了解详细错误信息")

            self.update_output.emit("开始整理输出文件...")
            organize_output_files()
            self.update_output.emit("文件整理完成")

            self.update_output.emit("打开输出文件夹...")
            open_output_folder()
            self.update_output.emit("输出文件夹已打开")

            # 检查SVN提交设置，根据设置决定是否提交SVN
            config = load_config()
            app_settings = config.get('app_settings', {})
            svn_submission = app_settings.get('svn_submission', True)

            if svn_submission:
                # 延时0.8秒后再提交SVN
                self.update_output.emit("准备提交SVN...")
                import time
                time.sleep(0.8)
                commit_to_svn()
                self.update_output.emit("SVN提交完成")
            else:
                self.update_output.emit("SVN提交功能已关闭，跳过提交")

            # 打印生成的合同编号
            if generated_contracts:
                self.update_output.emit("\n生成的合同编号:")
                for contract in generated_contracts:
                    self.update_output.emit(f"{contract}")
            else:
                self.update_output.emit("\n未生成任何合同")

            # 恢复确认按钮状态
            self.enable_button.emit()

            # 执行完成后保持上次选择的功能
        except Exception as e:
            # 处理错误
            error_msg = f"处理失败: {str(e)}"
            self.update_output.emit(error_msg)
            print(error_msg)
            # 在主线程中显示错误对话框
            from PyQt6.QtCore import QTimer
            error_str = str(e)
            QTimer.singleShot(0, lambda: self.show_error_dialog(error_str))
            # 恢复确认按钮状态
            self.enable_button.emit()
            # 错误处理后保持上次选择的功能
            logging.error(error_msg)


    def reset_to_default_selection(self):
        """重置为默认选择"""
        self.selected_operation = 1
        for i, radio_btn in enumerate(self.radio_buttons):
            if i == 0:
                radio_btn.setChecked(True)
                break

        # 更新状态显示
        for option in self.gui_options:
            if option['value'] == 1:
                operation_name = option['text']
                break
        else:
            operation_name = "SAAS合同"

        status_text = f"已选择: {operation_name}"
        self.append_output(status_text)
        self.confirm_btn.setEnabled(True)


    def show_error_dialog(self, error_msg):
        """显示错误对话框"""
        messagebox.showerror("错误", f"处理失败: {error_msg}")


    def enable_confirm_button(self):
        """启用确认按钮"""
        self.confirm_btn.setEnabled(True)
        self.confirm_btn.setText("确认执行")


    def update_output_window_size(self, width, height):
        """更新输出窗口大小 - 新布局中自动适配，无需手动调整"""
        pass


    def print_template_info(self):
        """打印模板信息到运行窗口"""
        try:
            # 打印登录用户信息
            # self.append_output("\n=== 登录用户信息 ===")

            # 获取全局变量中的用户信息（从主模块读取）
            import sys
            _main = sys.modules.get('__main__')
            _cu = getattr(_main, 'current_user', None) if _main else None
            _ut = getattr(_main, 'user_type', None) if _main else None

            if _cu:
                self.append_output(f"用户名: {_cu}")
                # self.append_output(f"用户类型: {_ut if _ut else '普通用户'}")
            else:
                self.append_output("用户名: 未登录")
                self.append_output("用户类型: 未登录")

            # self.append_output("====================")

            # 打印模板信息
            # self.append_output("\n=== 模板信息 ===")

            # 加载配置
            config = load_config()

            # 打印Word模板信息（只显示检查状态）
            word_templates = config.get('file_config', {}).get('word_templates', {})

            # 打印模板文件夹信息
            template_dir = config.get('path_config', {}).get('template_dir', 'template')

            # 检查模板文件是否存在
            template_dir_path = resolve_app_path(template_dir)

            # 只检查Word模板状态，不显示文件夹路径
            if word_templates:
                # 检查并只打印不存在的模板
                missing_templates = []
                for name, path in word_templates.items():
                    template_path = template_dir_path / path
                    if not template_path.exists():
                        missing_templates.append(name)

                if missing_templates:
                    self.append_output("异常模板（不存在）:")
                    for name in missing_templates:
                        self.append_output(f"  - {name}")
                else:
                    self.append_output("所有模板均存在")

            # self.append_output("====================")

        except Exception as e:
            self.append_output(f"打印模板信息失败: {str(e)}")


